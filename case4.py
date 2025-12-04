# Archivo: case4.py
# Ubicación: raíz del proyecto
# Descripción: Caso 4 - Rediseña el formato del estado de cuenta con nuevo encabezado, tabla y archivo resumen

import io
import os
import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from case1 import Case as BaseCase
from config_manager import ConfigManager


class Case(BaseCase):
    """Caso 4 - Aplica un rediseño al archivo del estado de cuenta y genera archivo resumen."""

    HEADERS: Tuple[str, ...] = (
        "Fecha",
        "Código",
        "Descripción",
        "Ref.",
        "Débitos (DR)",
        "Créditos (CR)",
        "Revisar",
        "Saldo Contable",
        "Ref2",
        "Tipo Tran",
        "Causa",
        "Sucursal",
        "D/C",
        "Cuenta",
    )

    INFO_FIELDS: Tuple[Tuple[str, str], ...] = (
        ("Titular de la cuenta", "titulardelacuenta"),
        ("Número de Cuenta", "numerodecuenta"),
        ("Moneda", "moneda"),
        ("Rango de fechas de movimientos", "rangodefechasdemovimientos"),
        ("Usuario que generó el reporte", "usuarioquegeneroelreporte"),
        ("Fecha del día y hora que se generó el reporte", "fechadeldiayhoraquesegeneroelreporte"),
    )

    def __init__(self):
        super().__init__()
        self.name = "Caso 4"
        self.description = (
            "Recibe archivos Excel del estado de cuenta y genera una versión con el nuevo diseño "
            "que reubica los datos del encabezado y actualiza los títulos de la tabla de movimientos. "
            "Además genera un archivo resumen contable."
        )
        self.response_message = (
            "Hola,\n\nSe adjuntan los archivos con el formato actualizado del estado de cuenta "
            "correspondiente al Caso 4. Quedo atento a cualquier comentario.\n\nSaludos cordiales."
        )
        self.config_manager = ConfigManager()
        self.config_case_key = 'case4'

    def get_search_keywords(self) -> List[str]:
        """Obtiene la palabra clave configurada para el Caso 4."""
        try:
            config = self.config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso4', '').strip()
            if keyword:
                return [keyword]
            return []
        except Exception as exc:
            print(f"Error al cargar palabras clave para caso4: {exc}")
            return []

    def process_email(self, email_data: Dict[str, Any], logger) -> Optional[Dict[str, Any]]:
        """Procesa los adjuntos Excel y devuelve el rediseño solicitado con archivo resumen."""
        try:
            sender = email_data.get('sender', '')
            subject = email_data.get('subject', '')
            attachments = email_data.get('attachments', [])

            logger.log(
                f"Procesando {self.name} para email de {sender} con asunto: {subject}",
                level="INFO",
            )

            date_range = self._extract_date_range(subject)
            if date_range:
                start, end = date_range
                logger.log(
                    (
                        "Se aplicará un filtrado de fechas desde "
                        f"{start.strftime('%d/%m/%Y')} hasta {end.strftime('%d/%m/%Y')}"
                    ),
                    level="INFO",
                )
            else:
                logger.log(
                    "No se encontró un rango de fechas válido en el asunto. Se conservarán "
                    "todos los movimientos.",
                    level="INFO",
                )

            excel_attachments = [
                attachment
                for attachment in attachments
                if self._is_excel_file(attachment.get('filename'))
            ]

            if not excel_attachments:
                logger.log(
                    "No se encontraron adjuntos de Excel para procesar en el correo recibido.",
                    level="WARNING",
                )
                return None

            processed_files: List[Dict[str, Any]] = []

            for attachment in excel_attachments:
                redesigned_files = self._redesign_excel_attachment(attachment, logger, date_range)
                if redesigned_files:
                    processed_files.extend(redesigned_files)

            if not processed_files:
                logger.log(
                    "No fue posible generar el rediseño para los archivos adjuntos proporcionados.",
                    level="ERROR",
                )
                return None

            response_data = {
                'recipient': sender,
                'subject': f"Re: {subject}",
                'body': self.response_message,
                'attachments': processed_files,
            }

            logger.log(
                f"Respuesta generada para {self.name} con {len(processed_files)} adjunto(s).",
                level="INFO",
            )

            return response_data

        except Exception as exc:
            logger.log(f"Error al procesar email en {self.name}: {exc}", level="ERROR")
            return None

    def _redesign_excel_attachment(
            self,
            attachment: Dict[str, Any],
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[List[Dict[str, Any]]]:
        """Genera archivo Excel rediseñado y archivo resumen contable."""
        filename = attachment.get('filename') or 'reporte.xlsx'
        content = attachment.get('content')

        if not content:
            logger.log(f"El adjunto '{filename}' está vacío o no pudo leerse.", level="WARNING")
            return None

        try:
            account_number = self._extract_account_number_from_b6(content, filename, logger)

            workbook_result = self._create_redesigned_workbook(
                content, filename, logger, date_range
            )

            if not workbook_result:
                return None

            workbook_bytes = workbook_result['workbook_bytes']
            data_rows = workbook_result['data_rows']

            output_name = self._build_output_filename(filename)

            attachments_list: List[Dict[str, Any]] = [
                {
                    'filename': output_name,
                    'content': workbook_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                }
            ]

            summary_bytes = self._create_summary_workbook(data_rows, account_number, logger)

            if summary_bytes:
                summary_name = self._build_summary_filename(output_name)
                attachments_list.append({
                    'filename': summary_name,
                    'content': summary_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                })
            else:
                logger.log(
                    f"No se pudo generar el archivo resumen contable para '{filename}'.",
                    level="WARNING",
                )

            return attachments_list

        except Exception as exc:
            logger.log(
                f"Error inesperado al rediseñar el archivo '{filename}': {exc}",
                level="ERROR",
            )
            return None

    def _extract_account_number_from_b6(
            self,
            file_bytes: bytes,
            original_name: str,
            logger
    ) -> str:
        """Extrae el número de cuenta de la celda B6, removiendo todas las letras."""
        try:
            from openpyxl import load_workbook
            import warnings

            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)

            sheet = workbook.active
            cell_value = sheet.cell(row=6, column=2).value

            if not cell_value:
                logger.log(
                    f"La celda B6 está vacía en '{original_name}'. Se usará cadena vacía para cuenta.",
                    level="WARNING",
                )
                return ''

            value_str = str(cell_value).strip()
            account_number = ''.join(char for char in value_str if char.isdigit())

            if account_number:
                logger.log(
                    f"Número de cuenta extraído de B6: '{account_number}'",
                    level="INFO",
                )
            else:
                logger.log(
                    f"No se pudieron extraer dígitos de B6 en '{original_name}'.",
                    level="WARNING",
                )

            return account_number

        except Exception as exc:
            logger.log(
                f"Error al extraer número de cuenta de B6 en '{original_name}': {exc}",
                level="ERROR",
            )
            return ''

    def _create_redesigned_workbook(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[Dict[str, Any]]:
        """Crea el nuevo archivo Excel con el encabezado y tabla actualizados."""
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image
        import warnings

        try:
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                source_wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        except Exception as exc:
            logger.log(
                f"No fue posible abrir el archivo '{original_name}' para rediseño: {exc}",
                level="ERROR",
            )
            return None

        source_ws = source_wb.active

        info_values = self._extract_info_fields(source_ws)
        header_row, header_map = self._find_header_row(source_ws)

        if not header_row or not header_map:
            logger.log(
                "No se encontraron encabezados válidos en el archivo fuente para aplicar el rediseño.",
                level="ERROR",
            )
            return None

        target_columns = [header_map.get(self._simplify_header(header)) for header in self.HEADERS]

        if not any(target_columns):
            logger.log(
                "No se pudieron mapear las columnas requeridas para el nuevo diseño.",
                level="ERROR",
            )
            return None

        missing_headers = [
            header
            for header, column_index in zip(self.HEADERS, target_columns)
            if header not in ("Código", "Revisar") and not column_index
        ]

        if missing_headers:
            logger.log(
                "No se encontraron todas las columnas esperadas. Las columnas faltantes son: "
                + ", ".join(missing_headers),
                level="WARNING",
            )

        data_rows: List[Dict[str, Any]] = []
        blank_streak = 0
        row_idx = header_row + 1
        max_row = source_ws.max_row

        while row_idx <= max_row:
            row_data: Dict[str, Any] = {}
            row_has_value = False

            for header, column_index in zip(self.HEADERS, target_columns):
                value = None
                if column_index:
                    value = source_ws.cell(row=row_idx, column=column_index).value
                row_data[header] = value
                if value not in (None, ''):
                    row_has_value = True

            if "Código" not in row_data or row_data["Código"] in (None, ''):
                row_data["Código"] = ""
            row_data["Revisar"] = ""

            if row_has_value:
                data_rows.append(row_data)
                blank_streak = 0
            else:
                blank_streak += 1
                if blank_streak >= 2:
                    break

            row_idx += 1

        if date_range:
            data_rows = self._filter_data_rows_by_date_range(data_rows, date_range, logger)

        if data_rows:
            self._assign_codes_by_description(data_rows, logger)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Detalle"

        self._insert_logo(worksheet, logger)
        self._populate_header_section(worksheet, info_values)
        self._populate_table(worksheet, data_rows)
        self._apply_styles(worksheet, len(data_rows))

        if data_rows:
            column_map = {header: idx for idx, header in enumerate(self.HEADERS, start=1)}
            header_row = 13
            data_start = header_row + 1
            data_end = data_start + len(data_rows) - 1
            self._highlight_rows_by_filters(
                worksheet,
                column_map,
                data_start,
                data_end,
                len(self.HEADERS),
                logger,
            )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return {
            'workbook_bytes': output.read(),
            'data_rows': data_rows
        }

    def _create_summary_workbook(
            self,
            data_rows: List[Dict[str, Any]],
            account_number: str,
            logger
    ) -> Optional[bytes]:
        """Genera el archivo resumen contable con los campos solicitados."""
        try:
            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Movimientos"

            headers = [
                "Cuenta Bancaria",
                "Tipo Documento",
                "Número",
                "Monto",
                "Fecha documento"
            ]
            worksheet.append(headers)

            rows_added = 0

            for row_data in data_rows:
                tipo_documento = row_data.get("Código", "")
                numero = row_data.get("Ref.", "")
                fecha_documento = row_data.get("Fecha")

                debit_value = row_data.get("Débitos (DR)")
                credit_value = row_data.get("Créditos (CR)")

                debit_amount = self._to_number(debit_value)
                credit_amount = self._to_number(credit_value)

                monto = None
                if debit_amount > 0:
                    monto = debit_amount
                elif credit_amount > 0:
                    monto = credit_amount

                if monto is None or monto == 0:
                    continue

                summary_row = [
                    account_number or '',
                    tipo_documento if tipo_documento else '',
                    numero if numero not in (None, '') else '',
                    monto,
                    fecha_documento if fecha_documento not in (None, '') else ''
                ]

                worksheet.append(summary_row)
                rows_added += 1

            for column_index in (4,):
                for cell in worksheet.iter_cols(
                        min_col=column_index,
                        max_col=column_index,
                        min_row=2,
                        max_row=worksheet.max_row
                ):
                    for numeric_cell in cell:
                        if isinstance(numeric_cell.value, (int, float)):
                            numeric_cell.number_format = '#,##0.00'

            for cell in worksheet.iter_cols(
                    min_col=5,
                    max_col=5,
                    min_row=2,
                    max_row=worksheet.max_row
            ):
                for date_cell in cell:
                    if isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'dd/mm/yyyy'

            logger.log(
                f"Se generó el archivo resumen contable con {rows_added} fila(s).",
                level="INFO",
            )

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.read()

        except Exception as exc:
            logger.log(
                f"Error al generar el archivo resumen contable: {exc}",
                level="ERROR",
            )
            return None

    def _build_summary_filename(self, formatted_name: str) -> str:
        """Construye el nombre del archivo resumen basado en el archivo formateado."""
        base, extension = os.path.splitext(formatted_name)
        marker = '_caso4_'
        if marker in base:
            prefix, suffix = base.split(marker, 1)
            return f"{prefix}_contable_{suffix}{extension}"
        return f"{base}_contable{extension}"

    def _insert_logo(self, worksheet, logger) -> None:
        """Inserta el logo de Davivienda en la celda A1."""
        from openpyxl.drawing.image import Image as OpenpyxlImage

        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            logo_path = os.path.join(current_dir, 'davivienda.png')

            logger.log(
                f"Buscando imagen en: {logo_path}",
                level="INFO",
            )

            if not os.path.exists(logo_path):
                logger.log(
                    f"ADVERTENCIA: No se encontró el archivo 'davivienda.png' en {current_dir}. "
                    f"Por favor, coloca el archivo en la misma carpeta que case4.py",
                    level="WARNING",
                )
                return

            img = OpenpyxlImage(logo_path)

            original_width = img.width
            original_height = img.height
            logger.log(
                f"Dimensiones originales de la imagen: {original_width}x{original_height}",
                level="INFO",
            )

            img.width = 150
            img.height = 50

            worksheet.row_dimensions[1].height = 50
            worksheet.column_dimensions['A'].width = 25

            img.anchor = 'A1'
            worksheet.add_image(img)

            logger.log(
                "Logo de Davivienda insertado correctamente en la celda A1.",
                level="INFO",
            )

        except Exception as exc:
            logger.log(
                f"Error al insertar el logo de Davivienda: {exc}",
                level="ERROR",
            )
            import traceback
            logger.log(
                f"Detalles del error: {traceback.format_exc()}",
                level="ERROR",
            )

    def _populate_header_section(self, worksheet, info_values: Dict[str, Any]) -> None:
        start_row = 5
        for offset, (label, _) in enumerate(self.INFO_FIELDS):
            row_index = start_row + offset
            worksheet.cell(row=row_index, column=1, value=label)
            worksheet.cell(row=row_index, column=2, value=info_values.get(label) or '')

    def _populate_table(self, worksheet, data_rows: List[Dict[str, Any]]) -> None:
        header_row = 13
        for col_idx, header in enumerate(self.HEADERS, start=1):
            worksheet.cell(row=header_row, column=col_idx, value=header)

        data_start = header_row + 1
        for row_offset, row_data in enumerate(data_rows):
            for col_idx, header in enumerate(self.HEADERS, start=1):
                worksheet.cell(row=data_start + row_offset, column=col_idx, value=row_data.get(header))

    def _apply_styles(self, worksheet, data_length: int) -> None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        header_row = 13
        start_row = header_row + 1
        end_row = start_row + max(data_length - 1, 0)

        title_font = Font(bold=True)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type='solid', fgColor='004C97')
        thin_border = Border(
            left=Side(border_style='thin', color='B0B0B0'),
            right=Side(border_style='thin', color='B0B0B0'),
            top=Side(border_style='thin', color='B0B0B0'),
            bottom=Side(border_style='thin', color='B0B0B0'),
        )

        for offset in range(len(self.INFO_FIELDS)):
            cell = worksheet.cell(row=5 + offset, column=1)
            cell.font = title_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

        for col_idx in range(1, len(self.HEADERS) + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        numeric_headers = {"Débitos (DR)", "Créditos (CR)", "Saldo Contable"}
        date_column = None

        for idx, header in enumerate(self.HEADERS, start=1):
            if header == "Fecha":
                date_column = idx
                break

        for row_idx in range(start_row, end_row + 1):
            for col_idx, header in enumerate(self.HEADERS, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if header in numeric_headers:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif date_column and col_idx == date_column:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        worksheet.freeze_panes = "A14"

        for col_idx in range(1, len(self.HEADERS) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                if cell.value is None:
                    continue
                text = str(cell.value)
                if len(text) > max_length:
                    max_length = len(text)
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 40)

    def _highlight_rows_by_filters(
            self,
            worksheet,
            column_map: Dict[str, int],
            start_row: int,
            end_row: int,
            total_columns: int,
            logger,
    ) -> None:
        """Resalta filas cuya descripción coincida con los filtros configurados para el Caso 4."""
        filters = self.config_manager.get_case4_filters()
        if not filters:
            return

        normalized_filters = [
            self._normalize_text(filter_text)
            for filter_text in filters
            if self._normalize_text(filter_text)
        ]

        if not normalized_filters:
            return

        description_column = column_map.get('Descripción')
        review_column = column_map.get('Revisar')

        if not description_column:
            logger.log(
                "No se encontró la columna de descripción para aplicar los filtros del Caso 4.",
                level="WARNING",
            )
            return

        from openpyxl.styles import Alignment, PatternFill

        highlight_fill = PatternFill(fill_type='solid', fgColor='FFF3B0')
        highlighted_rows = 0

        for row_idx in range(start_row, end_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=description_column).value
            if cell_value in (None, ''):
                continue

            normalized_value = self._normalize_text(str(cell_value))
            if not normalized_value:
                continue

            if any(filter_text in normalized_value for filter_text in normalized_filters):
                for col_idx in range(1, total_columns + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = highlight_fill

                if review_column:
                    review_cell = worksheet.cell(row=row_idx, column=review_column)
                    review_cell.value = 'Revisar'
                    review_cell.alignment = Alignment(horizontal='center', vertical='center')

                highlighted_rows += 1

        if highlighted_rows:
            logger.log(
                f"Se resaltaron {highlighted_rows} fila(s) según los filtros configurados para el Caso 4.",
                level="INFO",
            )

    def _extract_info_fields(self, worksheet) -> Dict[str, Any]:
        info: Dict[str, Any] = {label: '' for label, _ in self.INFO_FIELDS}

        max_row = min(worksheet.max_row, 60)
        max_col = min(worksheet.max_column, 20)

        lookup = {pattern: label for label, pattern in self.INFO_FIELDS}

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if not isinstance(cell_value, str):
                    continue

                simplified = self._simplify_header(cell_value)
                if simplified in lookup:
                    label = lookup[simplified]
                    value_cell = worksheet.cell(row=row, column=col + 1)
                    info[label] = value_cell.value if value_cell else ''

        return info

    def _extract_date_range(self, subject: str) -> Optional[Tuple[datetime, datetime]]:
        """Extrae un rango de fechas en formato dd/mm/yyyy del asunto del correo."""
        if not subject:
            return None

        matches = re.findall(r"(\d{2}/\d{2}/\d{4})", subject)
        if len(matches) < 2:
            return None

        try:
            start = datetime.strptime(matches[0], "%d/%m/%Y")
            end = datetime.strptime(matches[1], "%d/%m/%Y")
            return start, end
        except ValueError:
            return None

    def _parse_date_from_value(self, value: Any) -> Optional[datetime]:
        """Convierte el valor de una celda en un objeto datetime si es posible."""
        if isinstance(value, datetime):
            return value

        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())

        if isinstance(value, (int, float)) and value > 0:
            try:
                base_date = datetime(1899, 12, 30)
                converted = base_date + timedelta(days=float(value))
                if 1900 <= converted.year <= 9999:
                    return converted
            except Exception:
                pass

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None

            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue

        return None

    def _filter_data_rows_by_date_range(
            self,
            data_rows: List[Dict[str, Any]],
            date_range: Tuple[datetime, datetime],
            logger,
    ) -> List[Dict[str, Any]]:
        """Devuelve únicamente las filas cuyo valor en 'Fecha' está dentro del rango."""
        if not data_rows:
            return data_rows

        start, end = date_range
        if start > end:
            start, end = end, start

        filtered_rows: List[Dict[str, Any]] = []
        excluded_rows = 0
        unparsable_rows = 0

        for row in data_rows:
            raw_value = row.get("Fecha")
            parsed_date = self._parse_date_from_value(raw_value)

            if parsed_date is None:
                unparsable_rows += 1
                continue

            if start.date() <= parsed_date.date() <= end.date():
                filtered_rows.append(row)
            else:
                excluded_rows += 1

        if excluded_rows:
            logger.log(
                f"Se omitieron {excluded_rows} fila(s) fuera del rango de fechas solicitado.",
                level="INFO",
            )

        if unparsable_rows:
            logger.log(
                (
                    "Se omitieron {unparsable_rows} fila(s) adicionales porque no fue posible "
                    "interpretar la fecha en la columna correspondiente."
                ).format(unparsable_rows=unparsable_rows),
                level="WARNING",
            )

        if not filtered_rows:
            formatted_start = start.strftime("%d/%m/%Y")
            formatted_end = end.strftime("%d/%m/%Y")
            logger.log(
                (
                    "No se encontraron movimientos dentro del rango de fechas "
                    f"{formatted_start} - {formatted_end}."
                ),
                level="WARNING",
            )

        return filtered_rows

    def _assign_codes_by_description(
            self,
            data_rows: List[Dict[str, Any]],
            logger,
    ) -> None:
        """Asigna códigos a las filas basándose en reglas de descripción."""
        if not data_rows:
            return

        codification_rules = self._get_codification_rules()
        assigned_count = 0

        for row_data in data_rows:
            code = self._determine_codification(row_data, codification_rules)
            if code:
                row_data['Código'] = code
                assigned_count += 1
            else:
                row_data['Código'] = row_data.get('Código', '') or ''

        if assigned_count:
            logger.log(
                f"Se asignaron códigos automáticamente a {assigned_count} fila(s) según las reglas configuradas.",
                level="INFO",
            )

    def _get_codification_rules(self) -> Dict[str, List[Tuple[str, str]]]:
        """Obtiene y prepara las reglas de codificación configuradas para el caso."""
        fetchers = {
            'case4': self.config_manager.get_case4_codification_rules,
            'case5': self.config_manager.get_case5_codification_rules,
        }
        getter = fetchers.get(getattr(self, 'config_case_key', 'case4'), self.config_manager.get_case4_codification_rules)
        raw_rules = getter()
        prepared: Dict[str, List[Tuple[str, str]]] = {'debit': [], 'credit': []}

        for key in ('debit', 'credit'):
            entries = raw_rules.get(key, [])
            if not isinstance(entries, list):
                continue
            for item in entries:
                if not isinstance(item, dict):
                    continue
                search_text = item.get('search_text', '')
                code = item.get('code', '')
                if not isinstance(search_text, str) or not isinstance(code, str):
                    continue
                normalized_search = self._normalize_text(search_text)
                if normalized_search and code.strip():
                    prepared[key].append((normalized_search, code.strip()))

        return prepared

    def _determine_codification(
            self,
            row_data: Dict[str, Any],
            codification_rules: Dict[str, List[Tuple[str, str]]],
    ) -> str:
        """Determina el código a asignar a la fila según los filtros configurados."""
        description = row_data.get("Descripción")
        if not isinstance(description, str):
            return ''

        normalized_description = self._normalize_text(description)
        if not normalized_description:
            return ''

        credit_amount = self._to_number(row_data.get("Créditos (CR)"))
        debit_amount = self._to_number(row_data.get("Débitos (DR)"))

        if credit_amount > 0:
            code = self._match_codification(normalized_description, codification_rules.get('credit', []))
            if code:
                return code

        if debit_amount > 0:
            code = self._match_codification(normalized_description, codification_rules.get('debit', []))
            if code:
                return code

        return ''

    def _match_codification(
            self,
            normalized_description: str,
            rules: List[Tuple[str, str]],
    ) -> str:
        """Devuelve el código correspondiente si alguna regla coincide con la descripción."""
        for search_text, code in rules:
            if search_text and code and search_text in normalized_description:
                return code
        return ''

    def _to_number(self, value: Any) -> float:
        """Convierte un valor a número flotante cuando es posible."""
        if value is None:
            return 0.0

        if isinstance(value, (int, float)):
            return float(value)

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return 0.0
            cleaned = cleaned.replace(' ', '')
            if ',' in cleaned and '.' in cleaned:
                if cleaned.rfind(',') > cleaned.rfind('.'):
                    cleaned = cleaned.replace('.', '')
                    cleaned = cleaned.replace(',', '.')
                else:
                    cleaned = cleaned.replace(',', '')
            elif ',' in cleaned:
                cleaned = cleaned.replace(',', '.')
            else:
                cleaned = cleaned.replace(',', '')

            try:
                return float(cleaned)
            except ValueError:
                return 0.0

        return 0.0

    def _find_header_row(self, worksheet) -> Tuple[Optional[int], Dict[str, int]]:
        target_headers = {self._simplify_header(header) for header in self.HEADERS}
        best_row: Optional[int] = None
        best_matches = 0
        header_map: Dict[str, int] = {}

        max_row = min(worksheet.max_row, 80)

        for row_idx in range(1, max_row + 1):
            current_map: Dict[str, int] = {}
            matches = 0
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if not isinstance(cell_value, str):
                    continue
                simplified = self._simplify_header(cell_value)
                if not simplified:
                    continue
                current_map[simplified] = col_idx
                if simplified in target_headers:
                    matches += 1

            if matches > best_matches:
                best_matches = matches
                best_row = row_idx
                header_map = current_map

            if matches == len(target_headers):
                break

        if not best_row or best_matches == 0:
            return None, {}

        return best_row, header_map

    def _simplify_header(self, text: Any) -> str:
        if not isinstance(text, str):
            return ''
        normalized = self._normalize_text(text)
        return re.sub(r'[^a-z0-9]+', '', normalized)

    def _build_output_filename(self, original_name: str) -> str:
        """Construye el nombre del archivo de salida rediseñado."""
        base, _ = os.path.splitext(original_name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base}_caso4_{timestamp}.xlsx"