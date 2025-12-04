# Archivo: case2.py
# Ubicación: raíz del proyecto
# Descripción: Caso 2 - Recibe archivos Excel del estado de cuenta, mejora el formato y filtra por rango de fechas

import io
import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from case1 import Case as BaseCase
from config_manager import ConfigManager


class DateRangeNotFoundError(Exception):
    """Raised when the requested date range has no matching rows in the workbook."""


class Case(BaseCase):
    """Case 2 - Enhances Excel attachments and filters rows by a date range."""

    def __init__(self):
        super().__init__()
        self.name = "Caso 2"
        self.description = (
            "Recibe archivos Excel del estado de cuenta, mejora el formato, elimina la "
            "columna de balance y filtra los movimientos por el rango de fechas "
            "incluido en el asunto del correo."
        )
        self.config_manager = ConfigManager()
        self.config_case_key = 'case2'
        self.invalid_range_message_template = (
            "Hola,\n\n"
            "No se encontraron movimientos entre el {start} y el {end} en el/los archivo(s) "
            "recibido(s). Por favor verifica que el rango de fechas sea correcto y envía "
            "nuevamente archivo(s) con un rango válido.\n\n"
            "Saludos cordiales."
        )

    def get_search_keywords(self) -> List[str]:
        """Obtains the search keywords for case 2 from the configuration."""
        try:
            from config_manager import ConfigManager

            config_manager = ConfigManager()
            config = config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso2', '').strip()

            if keyword:
                return [keyword]
            return []
        except Exception as exc:
            print(f"Error al cargar palabras clave para caso2: {exc}")
            return []

    def process_email(self, email_data: Dict[str, Any], logger) -> Optional[Dict[str, Any]]:
        """Processes the email, enhances Excel attachments and applies the date filter."""
        try:
            sender = email_data.get('sender', '')
            subject = email_data.get('subject', '')
            attachments = email_data.get('attachments', [])

            logger.log(
                f"Procesando {self.name} para email de {sender} con asunto: {subject}",
                level="INFO",
            )

            excel_attachments = [
                attachment
                for attachment in attachments
                if self._is_excel_file(attachment.get('filename'))
            ]

            if not excel_attachments:
                logger.log(
                    "No se encontraron adjuntos de Excel para procesar en el correo recibido.",
                    level="WARNING",
                )
                return None

            date_range = self._extract_date_range(subject)
            if date_range:
                start, end = date_range
                logger.log(
                    (
                        "Se aplicará un filtrado de fechas desde "
                        f"{start.strftime('%d/%m/%Y')} hasta {end.strftime('%d/%m/%Y')}"
                    ),
                    level="INFO",
                )
            else:
                logger.log(
                    "No se encontró un rango de fechas válido en el asunto. Se conservarán "
                    "todos los movimientos.",
                    level="WARNING",
                )

            processed_files: List[Dict[str, Any]] = []
            files_with_invalid_range = 0

            for attachment in excel_attachments:
                try:
                    processed = self._enhance_excel_attachment(attachment, logger, date_range)
                    if processed:
                        processed_files.append(processed)
                except DateRangeNotFoundError as date_error:
                    logger.log(
                        f"El archivo '{attachment.get('filename')}' no contiene datos en el rango solicitado.",
                        level="WARNING",
                    )
                    files_with_invalid_range += 1
                    continue

            if files_with_invalid_range > 0 and not processed_files and date_range:
                start, end = date_range
                logger.log(
                    (
                        f"Ninguno de los {files_with_invalid_range} archivo(s) contenía movimientos "
                        "en el rango de fechas solicitado. Se enviará una respuesta solicitando un rango válido."
                    ),
                    level="ERROR",
                )
                return self._build_invalid_range_response(sender, subject, start, end)

            if not processed_files:
                logger.log(
                    "No fue posible mejorar el formato de los archivos adjuntos proporcionados.",
                    level="ERROR",
                )
                return None

            if files_with_invalid_range > 0:
                logger.log(
                    (
                        f"Se omitieron {files_with_invalid_range} archivo(s) por no contener datos "
                        f"en el rango solicitado. Se procesaron {len(processed_files)} archivo(s) exitosamente."
                    ),
                    level="INFO",
                )

            response_body = self.response_message

            response_data = {
                'recipient': sender,
                'subject': f"Re: {subject}",
                'body': response_body,
                'attachments': processed_files,
            }

            logger.log(
                f"Respuesta generada para {self.name} con {len(processed_files)} adjunto(s) mejorado(s).",
                level="INFO",
            )

            return response_data

        except Exception as exc:
            logger.log(f"Error al procesar email en {self.name}: {exc}", level="ERROR")
            return None

    def set_response_message(self, message: str) -> None:
        """Overrides setter to keep interface explicit in subclass."""
        self.response_message = message

    def _build_invalid_range_response(
            self,
            sender: str,
            subject: str,
            start: datetime,
            end: datetime,
    ) -> Dict[str, Any]:
        """Builds the response payload when the requested date range is invalid."""
        formatted_start = start.strftime("%d/%m/%Y")
        formatted_end = end.strftime("%d/%m/%Y")
        body = self.invalid_range_message_template.format(
            start=formatted_start,
            end=formatted_end,
        )
        return {
            'recipient': sender,
            'subject': f"Re: {subject}",
            'body': body,
            'attachments': [],
        }

    def _enhance_excel_attachment(
            self,
            attachment: Dict[str, Any],
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[Dict[str, Any]]:
        """Generates a new Excel file with enhanced formatting and date filtering."""
        filename = attachment.get('filename') or 'reporte.xls'
        content = attachment.get('content')

        if not content:
            logger.log(f"El adjunto '{filename}' está vacío o no pudo leerse.", level="WARNING")
            return None

        try:
            product_name = self._extract_product_name(content, filename, logger)
            workbook_bytes = self._create_formatted_workbook(content, filename, logger, date_range)
            if not workbook_bytes:
                return None

            output_name = self._build_output_filename(filename, product_name)
            return {
                'filename': output_name,
                'content': workbook_bytes,
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            }
        except DateRangeNotFoundError as exc:
            raise
        except ImportError as dependency_error:
            logger.log(
                f"Dependencia faltante para procesar '{filename}': {dependency_error}",
                level="ERROR",
            )
            return None
        except Exception as exc:
            logger.log(
                f"Error inesperado al mejorar el formato del archivo '{filename}': {exc}",
                level="ERROR",
            )
            return None

    def _create_formatted_workbook(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[bytes]:
        """Creates a new XLSX file applying formatting, removing balance columns and filtering dates."""
        from openpyxl import Workbook

        data_rows, max_cols = self._read_excel_matrix(file_bytes, original_name, logger)

        if not data_rows:
            logger.log(
                f"El archivo '{original_name}' no contiene datos para procesar.", level="WARNING"
            )
            return None

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Detalle"

        for row_idx, row in enumerate(data_rows, start=1):
            for col_idx in range(1, max_cols + 1):
                value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        self._insert_logo(worksheet, logger)
        self._remove_empty_columns(worksheet)
        self._remove_balance_columns(worksheet, logger)
        self._remove_debit_column(worksheet, logger)
        self._remove_summary_section(worksheet, logger)
        self._filter_rows_by_date_range(worksheet, date_range, logger)
        self._remove_zero_credit_rows(worksheet, logger)

        self._process_duplicate_references(worksheet, logger)

        adjusted_max_cols = worksheet.max_column

        self._apply_styles(worksheet, adjusted_max_cols, logger)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _process_duplicate_references(self, worksheet, logger) -> None:
        """Procesa referencias duplicadas y aplica las transformaciones de código y descripción."""
        header_row = 14
        if worksheet.max_row < header_row:
            return

        header_map = self._extract_header_map(worksheet, worksheet.max_column)
        reference_column = header_map.get('referencia')
        code_column = header_map.get('codigo')
        description_column = header_map.get('descripcion')

        if not all([reference_column, code_column, description_column]):
            logger.log(
                "No se encontraron todas las columnas necesarias (Referencia, Código, Descripción) "
                "para procesar referencias duplicadas.",
                level="WARNING",
            )
            return

        data_start_row = 16 if worksheet.max_row >= 16 else None
        if not data_start_row:
            return

        data_end_row = worksheet.max_row

        reference_groups: Dict[str, List[int]] = {}

        for row_idx in range(data_start_row, data_end_row + 1):
            reference_value = worksheet.cell(row=row_idx, column=reference_column).value

            if reference_value in (None, ''):
                continue

            reference_str = str(reference_value).strip()

            if not reference_str:
                continue

            if reference_str not in reference_groups:
                reference_groups[reference_str] = []

            reference_groups[reference_str].append(row_idx)

        duplicates_processed = 0

        for reference, row_indices in reference_groups.items():
            if len(row_indices) != 2:
                continue

            row1, row2 = row_indices

            code1_cell = worksheet.cell(row=row1, column=code_column)
            code2_cell = worksheet.cell(row=row2, column=code_column)

            code1 = str(code1_cell.value).strip().upper() if code1_cell.value else ''
            code2 = str(code2_cell.value).strip().upper() if code2_cell.value else ''

            wd_wc_row = None
            three_v_row = None
            wd_wc_code = None

            if code1 in ('WD', 'WC') and code2 == '3V':
                wd_wc_row = row1
                three_v_row = row2
                wd_wc_code = code1
            elif code2 in ('WD', 'WC') and code1 == '3V':
                wd_wc_row = row2
                three_v_row = row1
                wd_wc_code = code2

            if not wd_wc_row or not three_v_row:
                continue

            wd_wc_code_cell = worksheet.cell(row=wd_wc_row, column=code_column)
            three_v_code_cell = worksheet.cell(row=three_v_row, column=code_column)

            if wd_wc_code == 'WD':
                wd_wc_code_cell.value = 'T/D'
            elif wd_wc_code == 'WC':
                wd_wc_code_cell.value = 'T/C'

            three_v_code_cell.value = 'O/D'

            wd_wc_description_cell = worksheet.cell(row=wd_wc_row, column=description_column)
            three_v_description_cell = worksheet.cell(row=three_v_row, column=description_column)

            wd_wc_description = str(wd_wc_description_cell.value).strip() if wd_wc_description_cell.value else ''

            new_three_v_description = f"Comisión bancaria {wd_wc_description}"
            three_v_description_cell.value = new_three_v_description

            duplicates_processed += 1

            logger.log(
                f"Referencia duplicada procesada: '{reference}' - "
                f"Código {wd_wc_code} → {'T/D' if wd_wc_code == 'WD' else 'T/C'}, "
                f"Código 3V → O/D, "
                f"Descripción 3V actualizada con 'Comisión bancaria'",
                level="INFO",
            )

        if duplicates_processed > 0:
            logger.log(
                f"Se procesaron {duplicates_processed} par(es) de referencias duplicadas.",
                level="INFO",
            )
        else:
            logger.log(
                "No se encontraron referencias duplicadas con los códigos WD/WC y 3V para procesar.",
                level="INFO",
            )

    def _remove_balance_columns(self, worksheet, logger) -> None:
        """Removes columns whose headers correspond to balance information."""
        header_row = 14
        if worksheet.max_row < header_row:
            return

        columns_to_remove: List[int] = []
        for col_idx in range(worksheet.max_column, 0, -1):
            header_value = worksheet.cell(row=header_row, column=col_idx).value
            if isinstance(header_value, str):
                normalized = self._normalize_text(header_value)
                if normalized.startswith('balance'):
                    columns_to_remove.append(col_idx)

        for column in columns_to_remove:
            worksheet.delete_cols(column)
            logger.log(
                "Se eliminó la columna de balance según las reglas del Caso 2.",
                level="INFO",
            )

    def _remove_debit_column(self, worksheet, logger) -> None:
        """Removes the 'Débitos' column starting from the detail header row."""
        header_row = 14
        if worksheet.max_row < header_row:
            return

        debit_column: Optional[int] = None
        for col_idx in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=header_row, column=col_idx).value
            if isinstance(header_value, str):
                normalized = self._normalize_text(header_value)
                if normalized == 'debitos':
                    debit_column = col_idx
                    break

        if not debit_column:
            return

        max_column = worksheet.max_column
        for row_idx in range(header_row, worksheet.max_row + 1):
            for col_idx in range(debit_column, max_column):
                source_col = col_idx + 1
                target_cell = worksheet.cell(row=row_idx, column=col_idx)
                if source_col <= max_column:
                    source_cell = worksheet.cell(row=row_idx, column=source_col)
                    target_cell.value = source_cell.value
                else:
                    target_cell.value = None
            last_cell = worksheet.cell(row=row_idx, column=max_column)
            last_cell.value = None

        logger.log(
            "Se eliminó la columna de 'Débitos' a partir de la fila 14 según las reglas del Caso 2.",
            level="INFO",
        )

    def _filter_rows_by_date_range(
            self,
            worksheet,
            date_range: Optional[Tuple[datetime, datetime]],
            logger,
    ) -> None:
        """Filters worksheet rows to keep only those within the provided date range."""
        if not date_range:
            return

        start, end = date_range
        if start > end:
            start, end = end, start

        header_map = self._extract_header_map(worksheet, worksheet.max_column)
        date_column = self._locate_date_column(header_map)

        if not date_column:
            logger.log(
                "No se encontró una columna de fecha para aplicar el filtrado.",
                level="WARNING",
            )
            return

        summary_row = self._find_row_with_text(worksheet, 'Cuadro de Resumen')
        data_end_row = summary_row - 2 if summary_row else worksheet.max_row
        data_start_row = 16 if worksheet.max_row >= 16 else None

        if not data_start_row or data_start_row > data_end_row:
            return

        rows_to_delete: List[int] = []
        rows_in_range = 0
        for row_idx in range(data_start_row, data_end_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=date_column).value
            date_value = self._parse_date_from_value(cell_value)
            if date_value is None:
                continue

            if not (start.date() <= date_value.date() <= end.date()):
                rows_to_delete.append(row_idx)
            else:
                rows_in_range += 1

        for row_idx in reversed(rows_to_delete):
            worksheet.delete_rows(row_idx)

        if rows_to_delete:
            logger.log(
                f"Se eliminaron {len(rows_to_delete)} fila(s) fuera del rango de fechas.",
                level="INFO",
            )

        if rows_in_range == 0:
            formatted_start = start.strftime("%d/%m/%Y")
            formatted_end = end.strftime("%d/%m/%Y")
            raise DateRangeNotFoundError(
                (
                    "No se encontraron movimientos dentro del rango de fechas "
                    f"{formatted_start} - {formatted_end}."
                )
            )

    def _remove_summary_section(self, worksheet, logger) -> None:
        """Deletes the 'Cuadro de Resumen' section completely from the worksheet."""
        summary_row = self._find_row_with_text(worksheet, 'Cuadro de Resumen')
        if not summary_row:
            return

        rows_to_remove = worksheet.max_row - summary_row + 1
        worksheet.delete_rows(summary_row, rows_to_remove)
        logger.log(
            "Se eliminó la sección 'Cuadro de Resumen' del archivo procesado.",
            level="INFO",
        )

    def _highlight_rows_by_filters(
            self,
            worksheet,
            header_map: Dict[str, int],
            start_row: Optional[int],
            end_row: Optional[int],
            max_cols: int,
            logger,
    ) -> None:
        """Highlights rows whose description matches the configured filters for case 2."""
        filters = self.config_manager.get_case2_filters()
        if not filters:
            return

        normalized_filters = [
            self._normalize_text(filter_text)
            for filter_text in filters
            if self._normalize_text(filter_text)
        ]
        if not normalized_filters:
            return

        if not start_row or not end_row or start_row > end_row:
            return

        description_column = header_map.get('descripcion')

        if not description_column:
            logger.log(
                "No se encontró una columna de descripción para aplicar los filtros del Caso 2.",
                level="WARNING",
            )
            return

        from openpyxl.styles import Alignment, PatternFill

        highlight_fill = PatternFill(fill_type='solid', fgColor='FFF3B0')
        highlighted_rows = 0
        review_column = header_map.get('revisar')

        for row_idx in range(start_row, end_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=description_column).value
            if cell_value in (None, ''):
                continue

            normalized_value = self._normalize_text(str(cell_value))
            if not normalized_value:
                continue

            if any(filter_text in normalized_value for filter_text in normalized_filters):
                for col_idx in range(1, max_cols + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = highlight_fill
                if review_column:
                    review_cell = worksheet.cell(row=row_idx, column=review_column)
                    review_cell.value = 'Revisar'
                    review_cell.alignment = Alignment(horizontal='center', vertical='center')
                highlighted_rows += 1

        if highlighted_rows:
            logger.log(
                (
                    "Se resaltaron "
                    f"{highlighted_rows} fila(s) que coinciden con los filtros configurados del Caso 2."
                ),
                level="INFO",
            )

    def _locate_date_column(self, header_map: Dict[str, int]) -> Optional[int]:
        """Finds the column index that represents dates."""
        for header, index in header_map.items():
            if 'fecha' in header:
                return index
        return None

    def _extract_date_range(self, subject: str) -> Optional[Tuple[datetime, datetime]]:
        """Extracts a date range from the email subject."""
        if not subject:
            return None

        matches = re.findall(r"(\d{2}/\d{2}/\d{4})", subject)
        if len(matches) < 2:
            return None

        try:
            start = datetime.strptime(matches[0], "%d/%m/%Y")
            end = datetime.strptime(matches[1], "%d/%m/%Y")
            return start, end
        except ValueError:
            return None

    def _parse_date_from_value(self, value: Any) -> Optional[datetime]:
        """Attempts to parse a cell value into a datetime instance."""
        if isinstance(value, datetime):
            return value

        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())

        if isinstance(value, (int, float)) and value > 0:
            try:
                base_date = datetime(1899, 12, 30)
                return base_date + timedelta(days=float(value))
            except Exception:
                pass

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None

            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue

            numeric_candidate = cleaned.replace(',', '.')
            if re.fullmatch(r"-?\d+(\.\d+)?", numeric_candidate):
                try:
                    serial = float(numeric_candidate)
                    if serial > 0:
                        base_date = datetime(1899, 12, 30)
                        return base_date + timedelta(days=serial)
                except ValueError:
                    pass
        return None