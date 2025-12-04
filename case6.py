# Archivo: case6.py
# Ubicación: raíz del proyecto
# Descripción: Caso 6 - Transforma archivos Excel buscando CP/CB en columna Revisar

from __future__ import annotations

import io
import os
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from config_manager import ConfigManager


class MissingRequiredRowsError(Exception):
    """Excepción lanzada cuando no se encuentran filas CP/CB requeridas en el archivo."""
    pass


class InvalidFileFormatError(Exception):
    """Excepción lanzada cuando el archivo no tiene el formato esperado."""
    pass


class Case:
    """Caso 6 - Genera plantillas estándar buscando CP/CB en columna Revisar."""

    OUTPUT_HEADERS_CP = [
        "Proveedor",
        "Número",
        "Tipo Documento",
        "Fecha Documento",
        "Fecha Rige",
        "Aplicacion",
        "Monto",
        "Subtotal",
        "Descuento",
        "Impuesto1",
        "Impuesto2",
        "Rubro1",
        "Rubro2",
        "Condición De Pago",
        "Moneda",
        "Cuenta Bancaria",
        "Subtipo Documento",
        "Fecha Vence",
        "Codigo_impuesto",
        "Tipo Asiento",
        "Paquete",
        "Actividad Comercial",
    ]

    OUTPUT_HEADERS_CB = [
        "Cuenta Bancaria",
        "tipo Documento",
        "Numero",
        "Subtipo Documento",
        "Fecha",
        "Fecha Contable",
        "Concepto",
        "Monto",
        "Confirmado/entregado",
        "tipo Asiento",
        "Paquete",
        "Cod_impuesto",
    ]

    def __init__(self) -> None:
        self.name = "Caso 6"
        self.description = (
            "Recibe archivos Excel, busca filas con 'CP' o 'CB' en la columna Revisar "
            "y genera plantillas estándar correspondientes."
        )
        self.response_message = (
            "Hola,\n\nSe adjunta(n) el/los archivo(s) con la información transformada según la "
            "plantilla solicitada. Quedo atento a cualquier comentario.\n\nSaludos cordiales."
        )
        self.missing_rows_message = (
            "Hola,\n\nNo se pudieron procesar los archivos adjuntos porque no contienen las filas "
            "necesarias marcadas con 'CP' o 'CB' en la columna 'Revisar'.\n\n"
            "Por favor, envía un archivo que contenga las filas requeridas con las marcas 'CP' y/o 'CB' "
            "en la columna 'Revisar'.\n\nSaludos cordiales."
        )
        self.invalid_format_message = (
            "Hola,\n\nNo se pudieron procesar los archivos adjuntos porque no tienen el formato esperado. "
            "El archivo debe contener una columna llamada 'Revisar' en la fila de encabezados.\n\n"
            "Por favor, verifica que el archivo enviado sea el correcto y que contenga todas las columnas "
            "necesarias, incluyendo la columna 'Revisar'.\n\nSaludos cordiales."
        )
        self.config_manager = ConfigManager()

    def get_name(self) -> str:
        """Obtiene el nombre del caso"""
        return self.name

    def get_description(self) -> str:
        """Obtiene la descripción del caso"""
        return self.description

    def get_search_keywords(self) -> List[str]:
        """Obtiene la palabra clave configurada para el caso 6"""
        try:
            config = self.config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso6', '').strip()
            if keyword:
                return [keyword]
            return []
        except Exception as exc:
            print(f"Error al cargar palabras clave para caso6: {exc}")
            return []

    def process_email(self, email_data: Dict[str, Any], logger) -> Optional[Dict[str, Any]]:
        """Procesa el correo y genera la respuesta con los templates"""
        try:
            sender = email_data.get('sender', '')
            subject = email_data.get('subject', '')
            attachments = email_data.get('attachments', [])

            logger.log(
                f"Procesando {self.name} para email de {sender} con asunto: {subject}",
                level="INFO",
            )

            excel_attachments = [
                attachment
                for attachment in attachments
                if self._is_excel_file(attachment.get('filename'))
            ]

            if not excel_attachments:
                logger.log(
                    "No se encontraron adjuntos de Excel para procesar en el correo recibido.",
                    level="WARNING",
                )
                return None

            processed_files: List[Dict[str, Any]] = []
            files_without_rows = 0
            files_with_invalid_format = 0

            for attachment in excel_attachments:
                try:
                    files = self._create_template_workbooks(attachment, logger)
                    if files:
                        processed_files.extend(files)
                except MissingRequiredRowsError:
                    logger.log(
                        f"El archivo '{attachment.get('filename')}' no contiene filas CP/CB necesarias.",
                        level="WARNING",
                    )
                    files_without_rows += 1
                    continue
                except InvalidFileFormatError:
                    logger.log(
                        f"El archivo '{attachment.get('filename')}' no tiene el formato esperado.",
                        level="WARNING",
                    )
                    files_with_invalid_format += 1
                    continue

            # Prioridad 1: Si todos los archivos tienen formato inválido
            if files_with_invalid_format > 0 and not processed_files and files_without_rows == 0:
                logger.log(
                    f"Ninguno de los {files_with_invalid_format} archivo(s) tiene el formato esperado. "
                    "Se enviará una respuesta solicitando archivos con el formato correcto.",
                    level="ERROR",
                )
                return self._build_invalid_format_response(sender, subject)

            # Prioridad 2: Si todos los archivos no tienen filas CP/CB
            if files_without_rows > 0 and not processed_files and files_with_invalid_format == 0:
                logger.log(
                    f"Ninguno de los {files_without_rows} archivo(s) contiene las filas CP/CB necesarias. "
                    "Se enviará una respuesta solicitando archivos válidos.",
                    level="ERROR",
                )
                return self._build_missing_rows_response(sender, subject)

            # Prioridad 3: Si hay archivos con ambos problemas
            if (files_with_invalid_format > 0 or files_without_rows > 0) and not processed_files:
                logger.log(
                    f"Los archivos tienen problemas de formato o no contienen filas CP/CB. "
                    f"Formato inválido: {files_with_invalid_format}, Sin filas: {files_without_rows}. "
                    "Se enviará una respuesta de formato inválido.",
                    level="ERROR",
                )
                return self._build_invalid_format_response(sender, subject)

            if not processed_files:
                logger.log(
                    "No fue posible generar los archivos de plantilla requeridos para los adjuntos.",
                    level="ERROR",
                )
                return None

            if files_without_rows > 0 or files_with_invalid_format > 0:
                logger.log(
                    f"Se omitieron {files_without_rows + files_with_invalid_format} archivo(s) "
                    f"(Formato inválido: {files_with_invalid_format}, Sin filas: {files_without_rows}). "
                    f"Se procesaron {len(processed_files)} archivo(s) exitosamente.",
                    level="INFO",
                )

            response_data = {
                'recipient': sender,
                'subject': f"Re: {subject}",
                'body': self.response_message,
                'attachments': processed_files,
            }

            logger.log(
                f"Respuesta generada para {self.name} con {len(processed_files)} adjunto(s).",
                level="INFO",
            )

            return response_data
        except Exception as exc:
            logger.log(f"Error al procesar email en {self.name}: {exc}", level="ERROR")
            return None

    def get_response_message(self) -> str:
        """Obtiene el mensaje de respuesta"""
        return self.response_message

    def set_response_message(self, message: str) -> None:
        """Establece un nuevo mensaje de respuesta"""
        self.response_message = message

    def _build_missing_rows_response(
            self,
            sender: str,
            subject: str,
    ) -> Dict[str, Any]:
        """Construye la respuesta cuando los archivos no contienen filas CP/CB."""
        return {
            'recipient': sender,
            'subject': f"Re: {subject}",
            'body': self.missing_rows_message,
            'attachments': [],
        }

    def _build_invalid_format_response(
            self,
            sender: str,
            subject: str,
    ) -> Dict[str, Any]:
        """Construye la respuesta cuando los archivos no tienen el formato esperado."""
        return {
            'recipient': sender,
            'subject': f"Re: {subject}",
            'body': self.invalid_format_message,
            'attachments': [],
        }

    def _is_excel_file(self, filename: Optional[str]) -> bool:
        """Valida si el nombre de archivo corresponde a un Excel soportado"""
        if not filename:
            return False
        extension = os.path.splitext(filename)[1].lower()
        return extension in {'.xls', '.xlsx'}

    def _create_template_workbooks(self, attachment: Dict[str, Any], logger) -> List[Dict[str, Any]]:
        """Genera los archivos de plantilla desde el adjunto"""
        filename = attachment.get('filename') or 'reporte.xlsx'
        content = attachment.get('content')

        if not content:
            logger.log(
                f"El adjunto '{filename}' está vacío o no pudo leerse.",
                level="WARNING",
            )
            return []

        try:
            # Extraer información general del archivo
            file_info = self._extract_file_info(content, filename, logger)

            if not file_info:
                logger.log(
                    f"No se pudo extraer información del archivo '{filename}'.",
                    level="WARNING",
                )
                return []

            cuenta_bancaria = file_info.get('cuenta_bancaria', '')
            moneda = file_info.get('moneda', '')

            # Extraer filas CP/CB
            extraction_result = self._extract_rows_by_type(content, filename, logger)

            if not extraction_result:
                logger.log(
                    f"No se pudieron extraer datos del archivo '{filename}'.",
                    level="WARNING",
                )
                return []

            cp_rows = extraction_result['cp_rows']
            cb_rows = extraction_result['cb_rows']

            if not cp_rows and not cb_rows:
                logger.log(
                    f"No se encontraron filas con 'CP' o 'CB' en la columna Revisar del archivo '{filename}'.",
                    level="WARNING",
                )
                raise MissingRequiredRowsError(
                    f"El archivo '{filename}' no contiene filas CP/CB necesarias."
                )

            result_files: List[Dict[str, Any]] = []

            if cp_rows:
                logger.log(
                    f"Se encontraron {len(cp_rows)} fila(s) con 'CP' en la columna Revisar.",
                    level="INFO",
                )
                cp_workbook_bytes = self._build_cp_workbook(
                    cp_rows,
                    cuenta_bancaria,
                    moneda,
                    logger
                )
                cp_output_name = self._build_output_filename(filename, 'CP')
                result_files.append({
                    'filename': cp_output_name,
                    'content': cp_workbook_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                })

            if cb_rows:
                logger.log(
                    f"Se encontraron {len(cb_rows)} fila(s) con 'CB' en la columna Revisar.",
                    level="INFO",
                )
                cb_workbook_bytes = self._build_cb_workbook(
                    cb_rows,
                    cuenta_bancaria,
                    logger
                )
                cb_output_name = self._build_output_filename(filename, 'CB')
                result_files.append({
                    'filename': cb_output_name,
                    'content': cb_workbook_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                })

            return result_files

        except (MissingRequiredRowsError, InvalidFileFormatError):
            raise
        except ImportError as dependency_error:
            logger.log(
                f"Dependencia faltante para procesar '{filename}': {dependency_error}",
                level="ERROR",
            )
            return []
        except Exception as exc:
            logger.log(
                f"Error inesperado al transformar el archivo '{filename}': {exc}",
                level="ERROR",
            )
            return []

    def _extract_file_info(
            self,
            file_bytes: bytes,
            original_name: str,
            logger
    ) -> Optional[Dict[str, Any]]:
        """Extrae información general del archivo como cuenta bancaria y moneda"""
        extension = os.path.splitext(original_name)[1].lower()

        if extension == '.xls':
            return self._extract_info_from_xls(file_bytes, original_name, logger)
        else:
            return self._extract_info_from_xlsx(file_bytes, original_name, logger)

    def _extract_info_from_xls(
            self,
            file_bytes: bytes,
            original_name: str,
            logger
    ) -> Optional[Dict[str, Any]]:
        """Extrae información general de un archivo .xls"""
        import xlrd

        try:
            workbook = xlrd.open_workbook(file_contents=file_bytes)
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            logger.log(
                f"Error al abrir el archivo .xls '{original_name}': {exc}",
                level="ERROR",
            )
            return None

        # Extraer cuenta bancaria de fila 6, columna B (índice 1)
        cuenta_bancaria = ''
        try:
            cell_value = sheet.cell_value(5, 1)  # Fila 6 es índice 5
            if cell_value:
                # Remover todas las letras, solo mantener números
                cuenta_bancaria = ''.join(c for c in str(cell_value) if c.isdigit())
                logger.log(
                    f"Cuenta bancaria extraída (fila 6, col B): '{cuenta_bancaria}'",
                    level="INFO",
                )
        except Exception as exc:
            logger.log(
                f"Error al extraer cuenta bancaria de fila 6, columna B: {exc}",
                level="WARNING",
            )

        # Extraer moneda de fila 7, columna B (índice 1)
        moneda = ''
        try:
            cell_value = sheet.cell_value(6, 1)  # Fila 7 es índice 6
            if cell_value:
                moneda = str(cell_value).strip()
                logger.log(
                    f"Moneda extraída (fila 7, col B): '{moneda}'",
                    level="INFO",
                )
        except Exception as exc:
            logger.log(
                f"Error al extraer moneda de fila 7, columna B: {exc}",
                level="WARNING",
            )

        return {
            'cuenta_bancaria': cuenta_bancaria,
            'moneda': moneda
        }

    def _extract_info_from_xlsx(
            self,
            file_bytes: bytes,
            original_name: str,
            logger
    ) -> Optional[Dict[str, Any]]:
        """Extrae información general de un archivo .xlsx"""
        from openpyxl import load_workbook
        import warnings

        try:
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = workbook.active
        except Exception as exc:
            logger.log(
                f"Error al abrir el archivo .xlsx '{original_name}': {exc}",
                level="ERROR",
            )
            return None

        # Extraer cuenta bancaria de fila 6, columna B
        cuenta_bancaria = ''
        try:
            cell_value = sheet.cell(row=6, column=2).value
            if cell_value:
                # Remover todas las letras, solo mantener números
                cuenta_bancaria = ''.join(c for c in str(cell_value) if c.isdigit())
                logger.log(
                    f"Cuenta bancaria extraída (fila 6, col B): '{cuenta_bancaria}'",
                    level="INFO",
                )
        except Exception as exc:
            logger.log(
                f"Error al extraer cuenta bancaria de fila 6, columna B: {exc}",
                level="WARNING",
            )

        # Extraer moneda de fila 7, columna B
        moneda = ''
        try:
            cell_value = sheet.cell(row=7, column=2).value
            if cell_value:
                moneda = str(cell_value).strip()
                logger.log(
                    f"Moneda extraída (fila 7, col B): '{moneda}'",
                    level="INFO",
                )
        except Exception as exc:
            logger.log(
                f"Error al extraer moneda de fila 7, columna B: {exc}",
                level="WARNING",
            )

        return {
            'cuenta_bancaria': cuenta_bancaria,
            'moneda': moneda
        }

    def _extract_rows_by_type(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
    ) -> Optional[Dict[str, Any]]:
        """Extrae las filas donde la columna Revisar contiene 'CP' o 'CB'"""
        extension = os.path.splitext(original_name)[1].lower()

        if extension == '.xls':
            return self._extract_from_xls(file_bytes, original_name, logger)
        else:
            return self._extract_from_xlsx(file_bytes, original_name, logger)

    def _extract_from_xls(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
    ) -> Optional[Dict[str, Any]]:
        """Extrae datos de un archivo .xls usando xlrd"""
        import xlrd

        try:
            workbook = xlrd.open_workbook(file_contents=file_bytes)
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            logger.log(
                f"Error al abrir el archivo .xls '{original_name}': {exc}",
                level="ERROR",
            )
            return None

        header_row = 12
        header_map = self._build_header_map_xls(sheet, header_row)

        logger.log(
            f"Encabezados detectados en fila {header_row + 1}: {list(header_map.keys())}",
            level="INFO",
        )

        return self._extract_data_rows_xls(sheet, header_row, header_map, logger, workbook)

    def _extract_from_xlsx(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
    ) -> Optional[Dict[str, Any]]:
        """Extrae datos de un archivo .xlsx usando openpyxl"""
        from openpyxl import load_workbook
        import warnings

        try:
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = workbook.active
        except Exception as exc:
            logger.log(
                f"Error al abrir el archivo .xlsx '{original_name}': {exc}",
                level="ERROR",
            )
            return None

        header_row = 13
        header_map = self._build_header_map_xlsx(sheet, header_row)

        logger.log(
            f"Encabezados detectados en fila {header_row}: {list(header_map.keys())}",
            level="INFO",
        )

        return self._extract_data_rows_xlsx(sheet, header_row, header_map, logger)

    def _build_header_map_xls(self, sheet, header_row: int) -> Dict[str, int]:
        """Construye un mapa de encabezados normalizados para archivos .xls"""
        header_map: Dict[str, int] = {}
        for col_idx in range(sheet.ncols):
            value = sheet.cell_value(header_row, col_idx)
            if value:
                normalized = self._normalize_text(value)
                if normalized:
                    header_map[normalized] = col_idx
        return header_map

    def _build_header_map_xlsx(self, sheet, header_row: int) -> Dict[str, int]:
        """Construye un mapa de encabezados normalizados para archivos .xlsx"""
        header_map: Dict[str, int] = {}
        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row=header_row, column=col_idx).value
            if value is not None:
                normalized = self._normalize_text(value)
                if normalized:
                    header_map[normalized] = col_idx
        return header_map

    def _extract_data_rows_xls(
            self,
            sheet,
            header_row: int,
            header_map: Dict[str, int],
            logger,
            workbook
    ) -> Optional[Dict[str, Any]]:
        """Extrae las filas de datos con 'CP' o 'CB' en la columna Revisar para archivos .xls"""
        import xlrd

        review_column = header_map.get('revisar')
        fecha_column = header_map.get('fecha')
        descripcion_column = header_map.get('descripcion')
        ref_column = header_map.get('ref')
        debito_column = header_map.get('debitosdr')
        credito_column = header_map.get('creditoscr')

        if review_column is None:
            logger.log(
                f"No se encontró la columna 'Revisar' en el archivo. Columnas disponibles: {list(header_map.keys())}",
                level="ERROR",
            )
            raise InvalidFileFormatError("Columna 'Revisar' no encontrada en el archivo")

        data_start = header_row + 1
        max_row = sheet.nrows

        cp_rows: List[Dict[str, Any]] = []
        cb_rows: List[Dict[str, Any]] = []

        for row_idx in range(data_start, max_row):
            review_value = sheet.cell_value(row_idx, review_column) if review_column < sheet.ncols else None

            if review_value is None:
                continue

            review_str = str(review_value).strip().upper()

            if review_str not in ('CP', 'CB'):
                continue

            fecha_value = sheet.cell_value(row_idx, fecha_column) if fecha_column is not None else None
            parsed_date = self._parse_date_value_xls(fecha_value, workbook)

            descripcion_value = ''
            if descripcion_column is not None:
                desc_val = sheet.cell_value(row_idx, descripcion_column)
                descripcion_value = str(desc_val).strip() if desc_val else ''

            ref_value = ''
            if ref_column is not None:
                ref_val = sheet.cell_value(row_idx, ref_column)
                ref_value = str(ref_val).strip() if ref_val else ''

            debito_value = sheet.cell_value(row_idx, debito_column) if debito_column is not None else None
            credito_value = sheet.cell_value(row_idx, credito_column) if credito_column is not None else None

            debito_amount = self._parse_decimal(debito_value)
            credito_amount = self._parse_decimal(credito_value)

            row_data = {
                'fecha': parsed_date,
                'descripcion': descripcion_value,
                'referencia': ref_value,
                'debito': debito_amount if debito_amount is not None else 0,
                'credito': credito_amount if credito_amount is not None else 0,
            }

            if review_str == 'CP':
                cp_rows.append(row_data)
            elif review_str == 'CB':
                cb_rows.append(row_data)

        return {
            'cp_rows': cp_rows,
            'cb_rows': cb_rows,
        }

    def _extract_data_rows_xlsx(
            self,
            sheet,
            header_row: int,
            header_map: Dict[str, int],
            logger
    ) -> Optional[Dict[str, Any]]:
        """Extrae las filas de datos con 'CP' o 'CB' en la columna Revisar para archivos .xlsx"""
        review_column = header_map.get('revisar')
        fecha_column = header_map.get('fecha')
        descripcion_column = header_map.get('descripcion')
        ref_column = header_map.get('ref')
        debito_column = header_map.get('debitosdr')
        credito_column = header_map.get('creditoscr')

        if review_column is None:
            logger.log(
                f"No se encontró la columna 'Revisar' en el archivo. Columnas disponibles: {list(header_map.keys())}",
                level="ERROR",
            )
            raise InvalidFileFormatError("Columna 'Revisar' no encontrada en el archivo")

        data_start = header_row + 1
        max_row = sheet.max_row

        cp_rows: List[Dict[str, Any]] = []
        cb_rows: List[Dict[str, Any]] = []

        for row_idx in range(data_start, max_row + 1):
            review_value = sheet.cell(row=row_idx, column=review_column).value if review_column else None

            if review_value is None:
                continue

            review_str = str(review_value).strip().upper()

            if review_str not in ('CP', 'CB'):
                continue

            fecha_value = sheet.cell(row=row_idx, column=fecha_column).value if fecha_column else None
            parsed_date = self._parse_date_value(fecha_value)

            descripcion_value = ''
            if descripcion_column:
                desc_cell_value = sheet.cell(row=row_idx, column=descripcion_column).value
                descripcion_value = str(desc_cell_value).strip() if desc_cell_value is not None else ''

            ref_value = ''
            if ref_column:
                ref_cell_value = sheet.cell(row=row_idx, column=ref_column).value
                ref_value = str(ref_cell_value).strip() if ref_cell_value is not None else ''

            debito_value = sheet.cell(row=row_idx, column=debito_column).value if debito_column else None
            credito_value = sheet.cell(row=row_idx, column=credito_column).value if credito_column else None

            debito_amount = self._parse_decimal(debito_value)
            credito_amount = self._parse_decimal(credito_value)

            row_data = {
                'fecha': parsed_date,
                'descripcion': descripcion_value,
                'referencia': ref_value,
                'debito': debito_amount if debito_amount is not None else 0,
                'credito': credito_amount if credito_amount is not None else 0,
            }

            if review_str == 'CP':
                cp_rows.append(row_data)
            elif review_str == 'CB':
                cb_rows.append(row_data)

        return {
            'cp_rows': cp_rows,
            'cb_rows': cb_rows,
        }

    def _build_cp_workbook(
            self,
            cp_rows: List[Dict[str, Any]],
            cuenta_bancaria: str,
            moneda: str,
            logger
    ) -> bytes:
        """Construye el archivo de salida CP con las filas filtradas"""
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        sheet.append(self.OUTPUT_HEADERS_CP)

        for row_data in cp_rows:
            fecha = row_data.get('fecha')
            descripcion = row_data.get('descripcion', '')
            referencia = row_data.get('referencia', '')
            debito = row_data.get('debito', 0)
            credito = row_data.get('credito', 0)

            # Determinar el monto: tomar el valor que no sea 0
            monto = 0
            if debito and debito != 0:
                monto = debito
            elif credito and credito != 0:
                monto = credito

            row = [''] * len(self.OUTPUT_HEADERS_CP)
            row[0] = ''  # Proveedor
            row[1] = referencia  # Número
            row[2] = 'TEF'  # Tipo Documento
            row[3] = fecha  # Fecha Documento
            row[4] = fecha  # Fecha Rige
            row[5] = descripcion  # Aplicacion
            row[6] = monto  # Monto (toma valor no-cero de débito o crédito)
            row[7] = monto  # Subtotal (mismo valor que Monto)
            row[8] = 0  # Descuento
            row[9] = 0  # Impuesto1
            row[10] = 0  # Impuesto2
            row[11] = 0  # Rubro1
            row[12] = 0  # Rubro2
            row[13] = 0  # Condición De Pago
            row[14] = moneda  # Moneda (extraída de fila 7, columna B)
            row[15] = cuenta_bancaria  # Cuenta Bancaria (extraída de fila 6, columna B, solo números)
            row[16] = 0  # Subtipo Documento
            row[17] = fecha  # Fecha Vence
            row[19] = 'CP'  # Tipo Asiento
            row[20] = 'CP'  # Paquete
            row[21] = 523906  # Actividad Comercial

            sheet.append(row)

        # Aplicar formato de fecha a las columnas correspondientes
        for column_index in (4, 5, 18):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for date_cell in column_cells:
                    if isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'dd/mm/yyyy'

        # Aplicar formato numérico a las columnas de monto
        for column_index in (7, 8):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for numeric_cell in column_cells:
                    if isinstance(numeric_cell.value, (int, float)):
                        numeric_cell.number_format = '#,##0.00'

        logger.log(
            f"Se generó el archivo CP con {len(cp_rows)} fila(s).",
            level="INFO",
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _build_cb_workbook(
            self,
            cb_rows: List[Dict[str, Any]],
            cuenta_bancaria: str,
            logger
    ) -> bytes:
        """Construye el archivo de salida CB con las filas filtradas"""
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        sheet.append(self.OUTPUT_HEADERS_CB)

        for row_data in cb_rows:
            fecha = row_data.get('fecha')
            descripcion = row_data.get('descripcion', '')
            referencia = row_data.get('referencia', '')
            debito = row_data.get('debito', 0)
            credito = row_data.get('credito', 0)

            # Determinar el monto: tomar el valor que no sea 0
            monto = 0
            if credito and credito != 0:
                monto = credito
            elif debito and debito != 0:
                monto = debito

            row = [''] * len(self.OUTPUT_HEADERS_CB)
            row[0] = cuenta_bancaria  # Cuenta Bancaria (extraída de fila 6, columna B, solo números)
            row[1] = ''  # tipo Documento
            row[2] = referencia  # Numero
            row[3] = ''  # Subtipo Documento
            row[4] = fecha  # Fecha
            row[5] = fecha  # Fecha Contable
            row[6] = descripcion  # Concepto
            row[7] = monto  # Monto (toma valor no-cero)
            row[8] = ''  # Confirmado/entregado
            row[9] = 'CB'  # tipo Asiento
            row[10] = 'CB'  # Paquete
            row[11] = 'ND'  # Cod_impuesto

            sheet.append(row)

        # Aplicar formato de fecha a las columnas correspondientes
        for column_index in (5, 6):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for date_cell in column_cells:
                    if isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'dd/mm/yyyy'

        # Aplicar formato numérico a la columna de monto
        for column_index in (8,):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for numeric_cell in column_cells:
                    if isinstance(numeric_cell.value, (int, float)):
                        numeric_cell.number_format = '#,##0.00'

        logger.log(
            f"Se generó el archivo CB con {len(cb_rows)} fila(s).",
            level="INFO",
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _build_output_filename(self, original_name: str, file_type: str) -> str:
        """Construye el nombre del archivo de salida"""
        base, _ = os.path.splitext(original_name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base}_caso6_{file_type}_{timestamp}.xlsx"

    def _normalize_text(self, text: Any) -> str:
        """Normaliza texto eliminando acentos, espacios y caracteres especiales"""
        import unicodedata

        if not isinstance(text, str):
            return ''
        normalized = unicodedata.normalize('NFKD', text)
        normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r'[^\w\s]', '', normalized)
        # Eliminar todos los espacios para que "Débitos (DR)" se convierta en "debitosdr"
        return normalized.lower().strip().replace(' ', '')

    def _parse_date_value_xls(self, value: Any, workbook) -> Optional[datetime]:
        """Intenta convertir diferentes formatos de fecha a datetime para archivos .xls"""
        import xlrd

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue
            return None

        if isinstance(value, float):
            try:
                return datetime(*xlrd.xldate_as_tuple(value, workbook.datemode))
            except Exception:
                return None

        return None

    def _parse_date_value(self, value: Any) -> Optional[datetime]:
        """Intenta convertir diferentes formatos de fecha a datetime"""
        if isinstance(value, datetime):
            return value

        if isinstance(value, (int, float)):
            try:
                base_date = datetime(1899, 12, 30)
                converted = base_date + timedelta(days=float(value))
                if 1900 <= converted.year <= 9999:
                    return converted
            except Exception:
                pass

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue

        return None

    def _parse_decimal(self, value: Any) -> Optional[float]:
        """Convierte cadenas con separadores en valores numéricos"""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if not text or text in {'-', '--'}:
                return None
            text = text.replace('\xa0', '').replace(' ', '')
            text = re.sub(r'[^0-9,.-]', '', text)
            if not text:
                return None

            if text.count(',') > 1 and '.' not in text:
                last_comma = text.rfind(',')
                text = text[:last_comma].replace(',', '') + '.' + text[last_comma + 1:]
            elif text.count('.') > 1 and ',' not in text:
                last_dot = text.rfind('.')
                text = text[:last_dot].replace('.', '') + '.' + text[last_dot + 1:]
            elif text.count(',') == 1 and '.' not in text:
                text = text.replace(',', '.')
            elif text.count('.') == 1 and text.count(',') == 1:
                if text.rfind('.') < text.rfind(','):
                    text = text.replace('.', '')
                    text = text.replace(',', '.')
                else:
                    text = text.replace(',', '')

            try:
                return float(text)
            except ValueError:
                return None
        return None