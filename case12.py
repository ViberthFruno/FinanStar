# Archivo: case12.py
# Ubicación: raíz del proyecto
# Descripción: Caso 12 - Plantillas CP/CB desde archivos del Caso 10 usando columna Código

from __future__ import annotations

import io
import os
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from config_manager import ConfigManager


class MissingRequiredRowsError(Exception):
    """Excepción lanzada cuando no se encuentran filas CP/CB requeridas en el archivo."""
    pass


class Case:
    """Caso 12 - Genera plantillas CP/CB desde archivos del Caso 10 usando la columna Código."""

    OUTPUT_HEADERS_CP = [
        "Proveedor",
        "Número",
        "Tipo Documento",
        "Fecha Documento",
        "Fecha Rige",
        "Aplicacion",
        "Monto",
        "Subtotal",
        "Descuento",
        "Impuesto1",
        "Impuesto2",
        "Rubro1",
        "Rubro2",
        "Condición De Pago",
        "Moneda",
        "Cuenta Bancaria",
        "Subtipo Documento",
        "Fecha Vence",
        "Codigo_impuesto",
        "Tipo Asiento",
        "Paquete",
        "Actividad Comercial",
    ]

    OUTPUT_HEADERS_CB = [
        "Cuenta Bancaria",
        "tipo Documento",
        "Numero",
        "Subtipo Documento",
        "Fecha",
        "Fecha Contable",
        "Concepto",
        "Monto",
        "Confirmado/entregado",
        "tipo Asiento",
        "Paquete",
        "Cod_impuesto",
    ]

    def __init__(self) -> None:
        self.name = "Caso 12"
        self.description = (
            "Recibe archivos Excel generados por el Caso 10, identifica la cuenta bancaria "
            "y transforma su contenido en plantillas estándar (CP/CB) utilizando la columna "
            "'Código' para clasificar los movimientos según la configuración de cada cuenta."
        )
        self.response_message = (
            "Hola,\n\nSe adjunta(n) el/los archivo(s) con la información transformada según la "
            "plantilla solicitada. Quedo atento a cualquier comentario.\n\nSaludos cordiales."
        )
        self.missing_rows_message = (
            "Hola,\n\nNo se pudieron procesar los archivos adjuntos porque no contienen las filas "
            "necesarias marcadas con 'CP' o 'CB' en la columna 'Código' (o en 'Revisar').\n\n"
            "Por favor, envía un archivo generado por el Caso 10 y que "
            "contenga las filas requeridas con las marcas 'CP' y/o 'CB' en la columna 'Código'.\n\n"
            "Saludos cordiales."
        )
        self.config_manager = ConfigManager()
        self.case_key = 'case12'

    def get_name(self) -> str:
        """Obtiene el nombre del caso"""
        return self.name

    def get_description(self) -> str:
        """Obtiene la descripción del caso"""
        return self.description

    def get_search_keywords(self) -> List[str]:
        """Obtiene la palabra clave configurada para el caso 12"""
        try:
            config = self.config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso12', '').strip()
            if keyword:
                return [keyword]
            return []
        except Exception as exc:
            print(f"Error al cargar palabras clave para caso12: {exc}")
            return []

    def process_email(self, email_data: Dict[str, Any], logger) -> Optional[Dict[str, Any]]:
        """Procesa el correo y genera la respuesta con los templates"""
        try:
            sender = email_data.get('sender', '')
            subject = email_data.get('subject', '')
            attachments = email_data.get('attachments', [])

            logger.log(
                f"Procesando {self.name} para email de {sender} con asunto: {subject}",
                level="INFO",
            )

            excel_attachments = [
                attachment
                for attachment in attachments
                if self._is_excel_file(attachment.get('filename'))
            ]

            if not excel_attachments:
                logger.log(
                    "No se encontraron adjuntos de Excel para procesar en el correo recibido.",
                    level="WARNING",
                )
                return None

            processed_files: List[Dict[str, Any]] = []
            files_without_rows = 0

            for attachment in excel_attachments:
                try:
                    files = self._create_template_workbooks(attachment, logger)
                    if files:
                        processed_files.extend(files)
                except MissingRequiredRowsError:
                    logger.log(
                        f"El archivo '{attachment.get('filename')}' no contiene filas CP/CB necesarias.",
                        level="WARNING",
                    )
                    files_without_rows += 1
                    continue

            if files_without_rows > 0 and not processed_files:
                logger.log(
                    f"Ninguno de los {files_without_rows} archivo(s) contiene las filas CP/CB necesarias. "
                    "Se enviará una respuesta solicitando archivos válidos.",
                    level="ERROR",
                )
                return self._build_missing_rows_response(sender, subject)

            if not processed_files:
                logger.log(
                    "No fue posible generar los archivos de plantilla requeridos para los adjuntos.",
                    level="ERROR",
                )
                return None

            if files_without_rows > 0:
                logger.log(
                    f"Se omitieron {files_without_rows} archivo(s) por no contener filas CP/CB. "
                    f"Se procesaron {len(processed_files)} archivo(s) exitosamente.",
                    level="INFO",
                )

            response_data = {
                'recipient': sender,
                'subject': f"Re: {subject}",
                'body': self.response_message,
                'attachments': processed_files,
            }

            logger.log(
                f"Respuesta generada para {self.name} con {len(processed_files)} adjunto(s).",
                level="INFO",
            )

            return response_data
        except Exception as exc:
            logger.log(f"Error al procesar email en {self.name}: {exc}", level="ERROR")
            return None

    def get_response_message(self) -> str:
        """Obtiene el mensaje de respuesta"""
        return self.response_message

    def set_response_message(self, message: str) -> None:
        """Establece un nuevo mensaje de respuesta"""
        self.response_message = message

    def _build_missing_rows_response(
        self,
        sender: str,
        subject: str,
    ) -> Dict[str, Any]:
        """Construye la respuesta cuando los archivos no contienen filas CP/CB."""
        return {
            'recipient': sender,
            'subject': f"Re: {subject}",
            'body': self.missing_rows_message,
            'attachments': [],
        }

    def _is_excel_file(self, filename: Optional[str]) -> bool:
        """Valida si el nombre de archivo corresponde a un Excel soportado"""
        if not filename:
            return False
        extension = os.path.splitext(filename)[1].lower()
        return extension in {'.xls', '.xlsx'}

    def _create_template_workbooks(self, attachment: Dict[str, Any], logger) -> List[Dict[str, Any]]:
        """Genera los archivos de plantilla desde el adjunto"""
        filename = attachment.get('filename') or 'reporte.xlsx'
        content = attachment.get('content')

        if not content:
            logger.log(
                f"El adjunto '{filename}' está vacío o no pudo leerse.",
                level="WARNING",
            )
            return []

        try:
            account_code = self._extract_account_code(content, filename, logger)

            if not account_code:
                logger.log(
                    f"No se pudo extraer el código de cuenta del archivo '{filename}'. "
                    "Se omitirá el procesamiento de este archivo.",
                    level="WARNING",
                )
                return []

            account_name = self.config_manager.find_account_by_code(account_code, case_key=self.case_key)

            if not account_name:
                logger.log(
                    f"El código '{account_code}' no está asociado a ninguna cuenta configurada. "
                    f"Archivo '{filename}' omitido.",
                    level="WARNING",
                )
                return []

            logger.log(
                f"Archivo identificado para cuenta: '{account_name}' (código: {account_code})",
                level="INFO",
            )

            account_config = self.config_manager.get_case_account_config(self.case_key, account_name)

            if not account_config:
                logger.log(
                    f"No se pudo cargar la configuración de la cuenta '{account_name}'.",
                    level="ERROR",
                )
                return []

            extraction_result = self._extract_rows_by_type(
                content,
                filename,
                logger,
                account_config
            )

            if not extraction_result:
                logger.log(
                    f"No se pudieron extraer datos del archivo '{filename}'.",
                    level="WARNING",
                )
                return []

            cp_rows = extraction_result['cp_rows']
            cb_rows = extraction_result['cb_rows']
            currency_value = extraction_result['currency']
            account_number = extraction_result['account_number']

            if not cp_rows and not cb_rows:
                logger.log(
                    f"No se encontraron filas con 'CP' o 'CB' en las columnas 'Código' o 'Revisar' del archivo '{filename}'.",
                    level="WARNING",
                )
                raise MissingRequiredRowsError(
                    f"El archivo '{filename}' no contiene filas CP/CB necesarias."
                )

            result_files: List[Dict[str, Any]] = []

            if cp_rows:
                logger.log(
                    f"Se encontraron {len(cp_rows)} fila(s) clasificadas como 'CP' usando la columna 'Código'.",
                    level="INFO",
                )
                cp_workbook_bytes = self._build_cp_workbook(
                    cp_rows,
                    currency_value,
                    account_number,
                    account_config,
                    logger
                )
                cp_output_name = self._build_output_filename(filename, 'CP', account_name)
                result_files.append({
                    'filename': cp_output_name,
                    'content': cp_workbook_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                })

            if cb_rows:
                logger.log(
                    f"Se encontraron {len(cb_rows)} fila(s) clasificadas como 'CB' usando la columna 'Código'.",
                    level="INFO",
                )
                cb_workbook_bytes = self._build_cb_workbook(
                    cb_rows,
                    account_number,
                    account_config,
                    logger
                )
                cb_output_name = self._build_output_filename(filename, 'CB', account_name)
                result_files.append({
                    'filename': cb_output_name,
                    'content': cb_workbook_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                })

            return result_files

        except MissingRequiredRowsError:
            raise
        except ImportError as dependency_error:
            logger.log(
                f"Dependencia faltante para procesar '{filename}': {dependency_error}",
                level="ERROR",
            )
            return []
        except Exception as exc:
            logger.log(
                f"Error inesperado al transformar el archivo '{filename}': {exc}",
                level="ERROR",
            )
            return []

    def _extract_account_code(self, file_bytes: bytes, original_name: str, logger) -> Optional[str]:
        """Extrae el código de cuenta de la celda B4"""
        extension = os.path.splitext(original_name)[1].lower()

        try:
            if extension == '.xls':
                import xlrd
                workbook = xlrd.open_workbook(file_contents=file_bytes)
                sheet = workbook.sheet_by_index(0)

                if sheet.nrows >= 4 and sheet.ncols >= 2:
                    product_value = sheet.cell_value(3, 1)
                else:
                    return None
            else:
                from openpyxl import load_workbook
                workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
                sheet = workbook.active
                product_cell = sheet.cell(row=4, column=2)
                product_value = product_cell.value

            if not product_value:
                return None

            account_code = self._parse_account_code_from_product(product_value)

            if account_code:
                logger.log(
                    f"Código de cuenta extraído de B4: '{account_code}'",
                    level="INFO",
                )

            return account_code

        except Exception as exc:
            logger.log(
                f"Error al extraer código de cuenta de '{original_name}': {exc}",
                level="WARNING",
            )
            return None

    def _parse_account_code_from_product(self, product_value: Any) -> Optional[str]:
        """Extrae el código de cuenta del valor de producto en B4"""
        if not product_value:
            return None

        text = str(product_value).strip()
        if not text:
            return None

        pattern = r'CR\d{20}'
        match = re.search(pattern, text)

        if match:
            return match.group(0)

        return None

    def _extract_rows_by_type(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            account_config: Dict[str, Any],
    ) -> Optional[Dict[str, Any]]:
        """Extrae las filas donde las columnas 'Código' o 'Revisar' contienen 'CP' o 'CB'"""
        extension = os.path.splitext(original_name)[1].lower()

        if extension == '.xls':
            return self._extract_from_xls(file_bytes, original_name, logger, account_config)
        else:
            return self._extract_from_xlsx(file_bytes, original_name, logger, account_config)

    def _extract_from_xls(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            account_config: Dict[str, Any],
    ) -> Optional[Dict[str, Any]]:
        """Extrae datos de un archivo .xls usando xlrd"""
        import xlrd

        try:
            workbook = xlrd.open_workbook(file_contents=file_bytes)
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            logger.log(
                f"Error al abrir el archivo .xls '{original_name}': {exc}",
                level="ERROR",
            )
            return None

        currency_value = self._extract_currency_from_cell_xls(sheet, 6, 4, logger)
        product_name = self._extract_product_from_cell_xls(sheet, 6, 1, logger)
        account_number = self._extract_account_number(product_name)

        if account_number:
            logger.log(
                f"Número de cuenta bancaria extraído: '{account_number}'",
                level="INFO",
            )
        else:
            logger.log(
                "No se pudo extraer el número de cuenta bancaria del producto.",
                level="WARNING",
            )

        header_row = self._detect_header_row_xls(sheet)
        if header_row is None:
            logger.log(
                f"No se pudo identificar la fila de encabezados en '{original_name}'.",
                level="WARNING",
            )
            return None

        header_map = self._build_header_map_xls(sheet, header_row)

        logger.log(
            f"Encabezados detectados en fila {header_row + 1}: {list(header_map.keys())}",
            level="INFO",
        )

        return self._extract_data_rows(
            sheet, header_row, header_map, logger,
            account_number, currency_value, workbook, is_xls=True
        )

    def _extract_from_xlsx(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            account_config: Dict[str, Any],
    ) -> Optional[Dict[str, Any]]:
        """Extrae datos de un archivo .xlsx usando openpyxl"""
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = workbook.active
        except Exception as exc:
            logger.log(
                f"Error al abrir el archivo .xlsx '{original_name}': {exc}",
                level="ERROR",
            )
            return None

        currency_value = self._extract_currency_from_e7(sheet, logger)
        product_name = self._extract_product_from_b7(sheet, logger)
        account_number = self._extract_account_number(product_name)

        if account_number:
            logger.log(
                f"Número de cuenta bancaria extraído: '{account_number}'",
                level="INFO",
            )
        else:
            logger.log(
                "No se pudo extraer el número de cuenta bancaria del producto.",
                level="WARNING",
            )

        header_row = self._detect_header_row(sheet)
        if not header_row:
            logger.log(
                f"No se pudo identificar la fila de encabezados en '{original_name}'.",
                level="WARNING",
            )
            return None

        header_map = self._build_header_map(sheet, header_row)

        logger.log(
            f"Encabezados detectados en fila {header_row}: {list(header_map.keys())}",
            level="INFO",
        )

        return self._extract_data_rows(
            sheet, header_row, header_map, logger,
            account_number, currency_value, None, is_xls=False
        )

    def _extract_data_rows(
            self,
            sheet,
            header_row,
            header_map: Dict[str, int],
            logger,
            account_number: str,
            currency_value: str,
            workbook_xls=None,
            is_xls: bool = False
    ) -> Optional[Dict[str, Any]]:
        """Extrae las filas de datos con 'CP' o 'CB' utilizando las columnas 'Código' o 'Revisar'"""
        date_column = self._locate_date_column(header_map)
        review_column = header_map.get('revisar')
        description_column = header_map.get('descripcion')
        debit_column = header_map.get('debitos')
        credit_column = header_map.get('creditos')
        reference_column = header_map.get('referencia')
        code_column = header_map.get('codigo')

        if not date_column:
            logger.log(
                "No se encontró una columna de fecha.",
                level="WARNING",
            )
            return None

        if not review_column and not code_column:
            logger.log(
                "No se encontraron las columnas 'Código' ni 'Revisar' en el archivo.",
                level="WARNING",
            )
            return None

        if is_xls:
            summary_row = self._find_row_with_text_xls(sheet, 'Cuadro de Resumen')
            max_row = sheet.nrows
            data_start = header_row + 1
        else:
            summary_row = self._find_row_with_text(sheet, 'Cuadro de Resumen')
            max_row = sheet.max_row
            data_start = header_row + 1

        data_end = (
                       summary_row - 2 if summary_row and summary_row > data_start else max_row) - 1 if is_xls else summary_row - 2 if summary_row and summary_row > data_start else max_row

        cp_rows: List[Dict[str, Any]] = []
        cb_rows: List[Dict[str, Any]] = []

        row_range = range(data_start, data_end + 1)

        for row_idx in row_range:
            if is_xls:
                review_value = sheet.cell_value(row_idx, review_column - 1) if review_column else None
                code_value_raw = sheet.cell_value(row_idx, code_column - 1) if code_column else None
            else:
                review_value = sheet.cell(row=row_idx, column=review_column).value if review_column else None
                code_value_raw = sheet.cell(row=row_idx, column=code_column).value if code_column else None

            marker = ''
            if code_value_raw not in (None, ''):
                marker = str(code_value_raw).strip().upper()

            if not marker and review_value not in (None, ''):
                marker = str(review_value).strip().upper()

            if marker not in ('CP', 'CB'):
                continue

            if is_xls:
                date_value = sheet.cell_value(row_idx, date_column - 1)
                parsed_date = self._parse_date_value_xls(date_value, workbook_xls)

                description_value = ''
                if description_column:
                    desc_val = sheet.cell_value(row_idx, description_column - 1)
                    description_value = str(desc_val).strip() if desc_val else ''

                debit_value = sheet.cell_value(row_idx, debit_column - 1) if debit_column else None
                credit_value = sheet.cell_value(row_idx, credit_column - 1) if credit_column else None

                reference_value = ''
                if reference_column:
                    ref_val = sheet.cell_value(row_idx, reference_column - 1)
                    reference_value = str(ref_val).strip() if ref_val else ''

                code_value = ''
                if code_column:
                    code_val = sheet.cell_value(row_idx, code_column - 1)
                    code_value = str(code_val).strip() if code_val not in (None, '') else ''
            else:
                date_value = sheet.cell(row=row_idx, column=date_column).value
                parsed_date = self._parse_date_value(date_value)

                description_value = ''
                if description_column:
                    desc_cell_value = sheet.cell(row=row_idx, column=description_column).value
                    description_value = str(desc_cell_value).strip() if desc_cell_value is not None else ''

                debit_value = sheet.cell(row=row_idx, column=debit_column).value if debit_column else None
                credit_value = sheet.cell(row=row_idx, column=credit_column).value if credit_column else None

                reference_value = ''
                if reference_column:
                    ref_cell_value = sheet.cell(row=row_idx, column=reference_column).value
                    reference_value = str(ref_cell_value).strip() if ref_cell_value is not None else ''

                code_value = ''
                if code_column:
                    code_cell_value = sheet.cell(row=row_idx, column=code_column).value
                    code_value = str(code_cell_value).strip() if code_cell_value not in (None, '') else ''

            debit_amount = self._parse_decimal(debit_value)
            credit_amount = self._parse_decimal(credit_value)

            if not parsed_date:
                continue

            row_data = {
                'fecha': parsed_date,
                'descripcion': description_value,
                'debito': debit_amount if debit_amount is not None else 0,
                'credito': credit_amount if credit_amount is not None else 0,
                'referencia': reference_value,
                'codigo': code_value
            }

            if marker == 'CP':
                cp_rows.append(row_data)
            elif marker == 'CB':
                cb_rows.append(row_data)

        return {
            'cp_rows': cp_rows,
            'cb_rows': cb_rows,
            'currency': currency_value,
            'account_number': account_number or ''
        }

    def _extract_currency_from_cell_xls(self, sheet, row: int, col: int, logger) -> str:
        """Extrae el valor de moneda de una celda en formato .xls"""
        try:
            if row >= sheet.nrows or col >= sheet.ncols:
                logger.log(
                    f"La celda ({row}, {col}) está fuera de rango.",
                    level="WARNING",
                )
                return ''

            currency_value = sheet.cell_value(row, col)

            if not currency_value:
                logger.log(
                    f"La celda ({row}, {col}) está vacía. Se usará valor por defecto para Moneda.",
                    level="WARNING",
                )
                return ''

            currency_str = str(currency_value).strip()
            logger.log(
                f"Valor de moneda extraído: '{currency_str}'",
                level="INFO",
            )
            return currency_str

        except Exception as exc:
            logger.log(
                f"Error al extraer el valor de moneda: {exc}",
                level="WARNING",
            )
            return ''

    def _extract_product_from_cell_xls(self, sheet, row: int, col: int, logger) -> Optional[str]:
        """Extrae el valor de producto de una celda en formato .xls"""
        try:
            if row >= sheet.nrows or col >= sheet.ncols:
                logger.log(
                    f"La celda ({row}, {col}) está fuera de rango.",
                    level="WARNING",
                )
                return None

            product_value = sheet.cell_value(row, col)

            if not product_value:
                logger.log(
                    f"La celda ({row}, {col}) está vacía. No se podrá extraer el número de cuenta.",
                    level="WARNING",
                )
                return None

            product_str = str(product_value).strip()

            if not product_str:
                return None

            logger.log(
                f"Valor de producto extraído: '{product_str}'",
                level="INFO",
            )
            return product_str

        except Exception as exc:
            logger.log(
                f"Error al extraer el valor de producto: {exc}",
                level="WARNING",
            )
            return None

    def _detect_header_row_xls(self, sheet) -> Optional[int]:
        """Detecta la fila de encabezados en una hoja .xls (fila 7)"""
        # Primero verificar en fila 7 (índice 6)
        if sheet.nrows >= 7:
            normalized_values = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(6, col_idx)
                if isinstance(cell_value, str):
                    normalized_values.append(self._normalize_text(cell_value))

            if normalized_values:
                has_fecha = any('fecha' in value for value in normalized_values)
                has_balance = any('balance' in value for value in normalized_values)
                has_creditos = any('credito' in value for value in normalized_values)
                has_codigo = any('codigo' in value for value in normalized_values)
                has_revisar = any('revisar' in value for value in normalized_values)

                if has_fecha or (has_balance and has_creditos) or has_codigo or has_revisar:
                    return 6

        # Si no se encuentra en fila 7, buscar en las primeras filas
        max_rows = min(sheet.nrows, 50)
        for row_idx in range(max_rows):
            normalized_values = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                if isinstance(cell_value, str):
                    normalized_values.append(self._normalize_text(cell_value))

            if not normalized_values:
                continue

            if any(value.startswith('fecha') for value in normalized_values):
                return row_idx

        return None

    def _build_header_map_xls(self, sheet, header_row: int) -> Dict[str, int]:
        """Construye un mapa de encabezados normalizados para archivos .xls"""
        header_map: Dict[str, int] = {}
        for col_idx in range(sheet.ncols):
            value = sheet.cell_value(header_row, col_idx)
            if value:
                normalized = self._normalize_text(value)
                if normalized:
                    header_map[normalized] = col_idx + 1
        return header_map

    def _find_row_with_text_xls(self, sheet, text: str) -> Optional[int]:
        """Busca la fila que contiene el texto indicado en un archivo .xls"""
        if not text:
            return None
        target = self._normalize_text(text)
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                value = sheet.cell_value(row_idx, col_idx)
                if isinstance(value, str) and self._normalize_text(value) == target:
                    return row_idx
        return None

    def _parse_date_value_xls(self, value: Any, workbook) -> Optional[datetime]:
        """Intenta convertir diferentes formatos de fecha a datetime para archivos .xls"""
        import xlrd

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue
            return None

        if isinstance(value, float):
            try:
                return datetime(*xlrd.xldate_as_tuple(value, workbook.datemode))
            except Exception:
                return None

        return None

    def _parse_decimal(self, value: Any) -> Optional[float]:
        """Convierte cadenas con separadores en valores numéricos"""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if not text or text in {'-', '--'}:
                return None
            text = text.replace('\xa0', '').replace(' ', '')
            text = re.sub(r'[^0-9,.-]', '', text)
            if not text:
                return None

            if text.count(',') > 1 and '.' not in text:
                last_comma = text.rfind(',')
                text = text[:last_comma].replace(',', '') + '.' + text[last_comma + 1:]
            elif text.count('.') > 1 and ',' not in text:
                last_dot = text.rfind('.')
                text = text[:last_dot].replace('.', '') + '.' + text[last_dot + 1:]
            elif text.count(',') == 1 and '.' not in text:
                text = text.replace(',', '.')
            elif text.count('.') == 1 and text.count(',') == 1:
                if text.rfind('.') < text.rfind(','):
                    text = text.replace('.', '')
                    text = text.replace(',', '.')
                else:
                    text = text.replace(',', '')

            try:
                return float(text)
            except ValueError:
                return None
        return None

    def _extract_currency_from_e7(self, sheet, logger) -> str:
        """Extrae el valor de la celda E7 que contiene la moneda"""
        try:
            currency_cell = sheet.cell(row=7, column=5)
            currency_value = currency_cell.value

            if currency_value is None:
                logger.log(
                    "La celda E7 está vacía. Se usará valor por defecto para Moneda.",
                    level="WARNING",
                )
                return ''

            currency_str = str(currency_value).strip()
            logger.log(
                f"Valor de moneda extraído de E7: '{currency_str}'",
                level="INFO",
            )
            return currency_str

        except Exception as exc:
            logger.log(
                f"Error al extraer el valor de moneda de la celda E7: {exc}",
                level="WARNING",
            )
            return ''

    def _extract_product_from_b7(self, sheet, logger) -> Optional[str]:
        """Extrae el valor de la celda B7 que contiene el nombre del producto"""
        try:
            product_cell = sheet.cell(row=7, column=2)
            product_value = product_cell.value

            if product_value is None:
                logger.log(
                    "La celda B7 está vacía. No se podrá extraer el número de cuenta.",
                    level="WARNING",
                )
                return None

            if isinstance(product_value, str):
                product_str = product_value.strip()
            else:
                product_str = str(product_value).strip()

            if not product_str:
                return None

            logger.log(
                f"Valor de producto extraído de B7: '{product_str}'",
                level="INFO",
            )
            return product_str

        except Exception as exc:
            logger.log(
                f"Error al extraer el valor de producto de la celda B7: {exc}",
                level="WARNING",
            )
            return None

    def _extract_account_number(self, product_value: Optional[str]) -> Optional[str]:
        """Obtiene el número de cuenta bancaria a partir del valor de producto"""
        if not product_value:
            return None

        text = str(product_value).strip()
        if not text:
            return None

        if ':' in text:
            text = text.split(':', 1)[1].strip()

        tokens = text.split()
        if len(tokens) > 1:
            text = tokens[-1]

        condensed = ''.join(ch for ch in text if not ch.isspace())
        if len(condensed) <= 13:
            return None

        account = condensed[12:-1]
        account = account.strip()
        return account or None

    def _find_provider_code(self, description: str, account_config: Dict[str, Any], logger) -> str:
        """Busca el código de proveedor en la configuración de la cuenta"""
        if not description:
            return ''

        providers_list = account_config.get('providers', [])

        if not providers_list:
            return ''

        description_lower = description.lower()

        for provider in providers_list:
            search_text = provider.get('search_text', '')
            provider_code = provider.get('provider_code', '')

            if not search_text or not provider_code:
                continue

            search_text_lower = search_text.lower()

            if search_text_lower in description_lower:
                logger.log(
                    f"Proveedor encontrado: '{provider_code}' para descripción que contiene '{search_text}'",
                    level="INFO",
                )
                return provider_code

        return ''

    def _find_subtype_value(self, document_type: str, description: str, account_config: Dict[str, Any], logger) -> str:
        """Busca el valor de subtipo en la configuración de la cuenta"""
        if not document_type or not description:
            return ''

        subtypes_list = account_config.get('subtypes', [])

        if not subtypes_list:
            return ''

        document_type_upper = document_type.strip().upper()
        description_lower = description.lower()

        for subtype_rule in subtypes_list:
            rule_doc_type = subtype_rule.get('document_type', '').strip().upper()
            rule_search_text = subtype_rule.get('search_text', '').strip().lower()
            subtype_value = subtype_rule.get('subtype_value', '').strip()

            if not rule_doc_type or not rule_search_text or not subtype_value:
                continue

            if rule_doc_type == document_type_upper and rule_search_text in description_lower:
                logger.log(
                    f"Subtipo encontrado: '{subtype_value}' para tipo '{document_type}' con texto '{rule_search_text}' en descripción",
                    level="INFO",
                )
                return subtype_value

        return ''

    def _get_amount(self, debit: float, credit: float) -> float:
        """Obtiene el monto correcto: el que sea mayor a 0"""
        if debit > 0:
            return debit
        elif credit > 0:
            return credit
        return 0

    def _build_cp_workbook(
            self,
            cp_rows: List[Dict[str, Any]],
            currency_value: str,
            account_number: str,
            account_config: Dict[str, Any],
            logger
    ) -> bytes:
        """Construye el archivo de salida CP con las filas filtradas"""
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        sheet.append(self.OUTPUT_HEADERS_CP)

        for row_data in cp_rows:
            fecha = row_data.get('fecha')
            descripcion = row_data.get('descripcion', '')
            debito = row_data.get('debito', 0)
            referencia = row_data.get('referencia', '')

            provider_code = self._find_provider_code(descripcion, account_config, logger)

            row = [''] * len(self.OUTPUT_HEADERS_CP)
            row[0] = provider_code
            row[1] = referencia
            row[2] = 'TEF'
            row[3] = fecha
            row[4] = fecha
            row[5] = descripcion
            row[6] = debito
            row[7] = debito
            row[8] = 0
            row[9] = 0
            row[10] = 0
            row[11] = 0
            row[12] = 0
            row[13] = 0
            row[14] = currency_value
            row[15] = account_number
            row[16] = 0
            row[17] = fecha
            row[19] = 'CP'
            row[20] = 'CP'
            row[21] = 523906

            sheet.append(row)

        for column_index in (4, 5, 18):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for date_cell in column_cells:
                    if isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'dd/mm/yyyy'

        for column_index in (7, 8):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for numeric_cell in column_cells:
                    if isinstance(numeric_cell.value, (int, float)):
                        numeric_cell.number_format = '#,##0.00'

        logger.log(
            f"Se generó el archivo CP con {len(cp_rows)} fila(s), moneda '{currency_value}' y cuenta '{account_number}'.",
            level="INFO",
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _build_cb_workbook(
            self,
            cb_rows: List[Dict[str, Any]],
            account_number: str,
            account_config: Dict[str, Any],
            logger
    ) -> bytes:
        """Construye el archivo de salida CB con las filas filtradas"""
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        sheet.append(self.OUTPUT_HEADERS_CB)

        for row_data in cb_rows:
            fecha = row_data.get('fecha')
            descripcion = row_data.get('descripcion', '')
            debito = row_data.get('debito', 0)
            credito = row_data.get('credito', 0)
            referencia = row_data.get('referencia', '')
            codigo = row_data.get('codigo', '')

            monto = self._get_amount(debito, credito)
            subtype_value = self._find_subtype_value(codigo, descripcion, account_config, logger)

            row = [''] * len(self.OUTPUT_HEADERS_CB)
            row[0] = account_number
            row[1] = codigo
            row[2] = referencia
            row[3] = subtype_value
            row[4] = fecha
            row[5] = fecha
            row[6] = descripcion
            row[7] = monto
            row[8] = ''
            row[9] = 'CB'
            row[10] = 'CB'
            row[11] = 'ND'

            sheet.append(row)

        for column_index in (5, 6):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for date_cell in column_cells:
                    if isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'dd/mm/yyyy'

        for column_index in (8,):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for numeric_cell in column_cells:
                    if isinstance(numeric_cell.value, (int, float)):
                        numeric_cell.number_format = '#,##0.00'

        logger.log(
            f"Se generó el archivo CB con {len(cb_rows)} fila(s) y cuenta '{account_number}'.",
            level="INFO",
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _build_output_filename(self, original_name: str, file_type: str, account_name: str) -> str:
        """Construye el nombre del archivo de salida incluyendo el nombre de la cuenta"""
        base, _ = os.path.splitext(original_name)
        safe_account_name = re.sub(r'[^\w\s-]', '', account_name).strip().replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base}_{safe_account_name}_{file_type}_{timestamp}.xlsx"

    def _detect_header_row(self, sheet) -> Optional[int]:
        """Detecta la fila de encabezados en la hoja (fila 7)"""
        # Primero verificar en fila 7
        if sheet.max_row >= 7:
            normalized_values = [
                self._normalize_text(cell.value)
                for cell in sheet[7]
                if isinstance(cell.value, str)
            ]
            if normalized_values:
                has_fecha = any('fecha' in value for value in normalized_values)
                has_balance = any('balance' in value for value in normalized_values)
                has_creditos = any('credito' in value for value in normalized_values)
                has_codigo = any('codigo' in value for value in normalized_values)
                has_revisar = any('revisar' in value for value in normalized_values)

                if has_fecha or (has_balance and has_creditos) or has_codigo or has_revisar:
                    return 7

        # Si no se encuentra en fila 7, buscar en las primeras filas
        max_rows = min(sheet.max_row, 50)
        for row_idx in range(1, max_rows + 1):
            normalized_values = [
                self._normalize_text(cell.value)
                for cell in sheet[row_idx]
                if isinstance(cell.value, str)
            ]
            if not normalized_values:
                continue
            if any(value.startswith('fecha') for value in normalized_values):
                return row_idx
        return None

    def _build_header_map(self, sheet, header_row: int) -> Dict[str, int]:
        """Construye un mapa de encabezados normalizados"""
        header_map: Dict[str, int] = {}
        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row=header_row, column=col_idx).value
            if value is not None:
                original = str(value)
                normalized = self._normalize_text(value)
                if normalized:
                    header_map[normalized] = col_idx
        return header_map

    def _locate_date_column(self, header_map: Dict[str, int]) -> Optional[int]:
        """Localiza la columna de fecha en el mapa de encabezados"""
        if 'fecha' in header_map:
            return header_map['fecha']
        if 'fecha documento' in header_map:
            return header_map['fecha documento']
        for key, index in header_map.items():
            if key.startswith('fecha'):
                return index
        return None

    def _find_row_with_text(self, sheet, text: str) -> Optional[int]:
        """Busca la fila que contiene el texto indicado"""
        if not text:
            return None
        target = self._normalize_text(text)
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(value, str) and self._normalize_text(value) == target:
                    return row_idx
        return None

    def _normalize_text(self, value: Any) -> str:
        """Normaliza texto eliminando acentos, espacios y caracteres especiales"""
        import unicodedata
        import re

        if not isinstance(value, str):
            return ''
        normalized = unicodedata.normalize('NFKD', value)
        normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r'[^\w\s]', '', normalized)
        return normalized.lower().strip()

    def _parse_date_value(self, value: Any) -> Optional[datetime]:
        """Intenta convertir diferentes formatos de fecha a datetime"""
        if isinstance(value, datetime):
            return value

        if isinstance(value, (int, float)):
            try:
                base_date = datetime(1899, 12, 30)
                converted = base_date + timedelta(days=float(value))
                if 1900 <= converted.year <= 9999:
                    return converted
            except Exception:
                pass

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue
        return None