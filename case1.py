# Archivo: case1.py
# Ubicación: raíz del proyecto
# Descripción: Caso 1 - Procesa adjuntos Excel y devuelve una versión con formato mejorado

import io
import os
import re
import unicodedata
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from config_manager import ConfigManager


class Case:
    def __init__(self):
        """Inicializa el caso 1"""
        self.name = "Caso 1"
        self.description = (
            "Recibe archivos Excel del estado de cuenta, mejora el formato y lo reenvía al remitente"
        )
        self.search_keywords = []
        self.response_message = (
            "Hola,\n\nSe adjunta el archivo con el formato mejorado, manteniendo toda la información"
            " original. Quedo atento a cualquier comentario.\n\nSaludos cordiales."
        )
        self.config_manager = ConfigManager()
        self.config_case_key = 'case1'

    def get_name(self):
        """Obtiene el nombre del caso"""
        return self.name

    def get_description(self):
        """Obtiene la descripción del caso"""
        return self.description

    def get_search_keywords(self):
        """Obtiene las palabras clave de búsqueda desde la configuración"""
        try:
            from config_manager import ConfigManager

            config_manager = ConfigManager()
            config = config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso1', '').strip()

            if keyword:
                return [keyword]
            else:
                return []
        except Exception as e:
            print(f"Error al cargar palabras clave para caso1: {str(e)}")
            return []

    def process_email(self, email_data, logger):
        """Procesa el email y genera la respuesta automática"""
        try:
            sender = email_data.get('sender', '')
            subject = email_data.get('subject', '')
            attachments = email_data.get('attachments', [])

            logger.log(
                f"Procesando {self.name} para email de {sender} con asunto: {subject}",
                level="INFO",
            )

            excel_attachments = [att for att in attachments if self._is_excel_file(att.get('filename'))]

            if not excel_attachments:
                logger.log(
                    "No se encontraron adjuntos de Excel para procesar en el correo recibido.",
                    level="WARNING",
                )
                return None

            date_range = self._extract_date_range(subject)
            if date_range:
                start, end = date_range
                logger.log(
                    (
                        "Se aplicará un filtrado de fechas desde "
                        f"{start.strftime('%d/%m/%Y')} hasta {end.strftime('%d/%m/%Y')}"
                    ),
                    level="INFO",
                )
            else:
                logger.log(
                    "No se encontró un rango de fechas válido en el asunto. Se conservarán "
                    "todos los movimientos.",
                    level="WARNING",
                )

            processed_files = []
            for attachment in excel_attachments:
                processed = self._enhance_excel_attachment(attachment, logger, date_range)
                if isinstance(processed, list):
                    processed_files.extend(item for item in processed if item)
                elif processed:
                    processed_files.append(processed)

            if not processed_files:
                logger.log(
                    "No fue posible mejorar el formato de los archivos adjuntos proporcionados.",
                    level="ERROR",
                )
                return None

            response_body = self.response_message

            response_data = {
                'recipient': sender,
                'subject': f"Re: {subject}",
                'body': response_body,
                'attachments': processed_files,
            }

            logger.log(
                f"Respuesta generada para {self.name} con {len(processed_files)} adjunto(s) mejorado(s).",
                level="INFO",
            )

            return response_data

        except Exception as e:
            logger.log(f"Error al procesar email en caso1: {str(e)}", level="ERROR")
            return None

    def get_response_message(self):
        """Obtiene el mensaje de respuesta"""
        return self.response_message

    def set_response_message(self, message):
        """Establece un nuevo mensaje de respuesta"""
        self.response_message = message

    def _is_excel_file(self, filename: Optional[str]) -> bool:
        """Valida si el nombre de archivo corresponde a un Excel soportado"""
        if not filename:
            return False
        extension = os.path.splitext(filename)[1].lower()
        return extension in {'.xls', '.xlsx'}

    def _enhance_excel_attachment(
            self,
            attachment: Dict[str, Any],
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[List[Dict[str, Any]]]:
        """Genera nuevos archivos Excel con formato mejorado y resumen contable."""
        filename = attachment.get('filename') or 'reporte.xls'
        content = attachment.get('content')

        if not content:
            logger.log(f"El adjunto '{filename}' está vacío o no pudo leerse.", level="WARNING")
            return None

        try:
            product_name = self._extract_product_name(content, filename, logger)
            account_number = self._extract_account_number(product_name)
            workbook_bytes = self._create_formatted_workbook(
                content,
                filename,
                logger,
                date_range=date_range,
            )
            if not workbook_bytes:
                return None

            output_name = self._build_output_filename(filename, product_name)
            summary_bytes = self._create_summary_workbook(
                workbook_bytes,
                logger,
                account_number=account_number,
            )
            attachments: List[Dict[str, Any]] = [
                {
                    'filename': output_name,
                    'content': workbook_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                }
            ]

            if summary_bytes:
                summary_name = self._build_summary_filename(output_name)
                attachments.append(
                    {
                        'filename': summary_name,
                        'content': summary_bytes,
                        'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    }
                )
            else:
                logger.log(
                    (
                        "No se pudo generar el archivo resumen contable para "
                        f"'{filename}'. Se enviará únicamente el archivo formateado."
                    ),
                    level="WARNING",
                )

            return attachments
        except ImportError as dependency_error:
            logger.log(
                f"Dependencia faltante para procesar '{filename}': {dependency_error}",
                level="ERROR",
            )
            return None
        except Exception as exc:
            logger.log(
                f"Error inesperado al mejorar el formato del archivo '{filename}': {exc}",
                level="ERROR",
            )
            return None

    def _create_formatted_workbook(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[bytes]:
        """Crea un nuevo archivo XLSX con formato optimizado a partir del contenido original"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        data_rows, max_cols = self._read_excel_matrix(file_bytes, original_name, logger)

        if not data_rows:
            logger.log(
                f"El archivo '{original_name}' no contiene datos para procesar.", level="WARNING"
            )
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Detalle"

        for row_idx, row in enumerate(data_rows, start=1):
            for col_idx in range(1, max_cols + 1):
                value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                ws.cell(row=row_idx, column=col_idx, value=value)

        self._insert_logo(ws, logger)
        self._remove_empty_columns(ws)
        self._filter_rows_by_date_range(ws, date_range, logger)
        self._remove_zero_credit_rows(ws, logger)

        self._process_duplicate_references(ws, logger)

        adjusted_max_cols = ws.max_column

        self._apply_styles(ws, adjusted_max_cols, logger)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _process_duplicate_references(self, worksheet, logger) -> None:
        """Procesa referencias duplicadas y aplica las transformaciones de código y descripción."""
        header_row = 14
        if worksheet.max_row < header_row:
            return

        header_map = self._extract_header_map(worksheet, worksheet.max_column)
        reference_column = header_map.get('referencia')
        code_column = header_map.get('codigo')
        description_column = header_map.get('descripcion')

        if not all([reference_column, code_column, description_column]):
            logger.log(
                "No se encontraron todas las columnas necesarias (Referencia, Código, Descripción) "
                "para procesar referencias duplicadas.",
                level="WARNING",
            )
            return

        data_start_row = 16 if worksheet.max_row >= 16 else None
        if not data_start_row:
            return

        summary_row = self._find_row_with_text(worksheet, 'Cuadro de Resumen')
        data_end_row = summary_row - 2 if summary_row else worksheet.max_row

        reference_groups: Dict[str, List[int]] = {}

        for row_idx in range(data_start_row, data_end_row + 1):
            reference_value = worksheet.cell(row=row_idx, column=reference_column).value

            if reference_value in (None, ''):
                continue

            reference_str = str(reference_value).strip()

            if not reference_str:
                continue

            if reference_str not in reference_groups:
                reference_groups[reference_str] = []

            reference_groups[reference_str].append(row_idx)

        duplicates_processed = 0

        for reference, row_indices in reference_groups.items():
            if len(row_indices) != 2:
                continue

            row1, row2 = row_indices

            code1_cell = worksheet.cell(row=row1, column=code_column)
            code2_cell = worksheet.cell(row=row2, column=code_column)

            code1 = str(code1_cell.value).strip().upper() if code1_cell.value else ''
            code2 = str(code2_cell.value).strip().upper() if code2_cell.value else ''

            wd_wc_row = None
            three_v_row = None
            wd_wc_code = None

            if code1 in ('WD', 'WC') and code2 == '3V':
                wd_wc_row = row1
                three_v_row = row2
                wd_wc_code = code1
            elif code2 in ('WD', 'WC') and code1 == '3V':
                wd_wc_row = row2
                three_v_row = row1
                wd_wc_code = code2

            if not wd_wc_row or not three_v_row:
                continue

            wd_wc_code_cell = worksheet.cell(row=wd_wc_row, column=code_column)
            three_v_code_cell = worksheet.cell(row=three_v_row, column=code_column)

            if wd_wc_code == 'WD':
                wd_wc_code_cell.value = 'T/D'
            elif wd_wc_code == 'WC':
                wd_wc_code_cell.value = 'T/C'

            three_v_code_cell.value = 'O/D'

            wd_wc_description_cell = worksheet.cell(row=wd_wc_row, column=description_column)
            three_v_description_cell = worksheet.cell(row=three_v_row, column=description_column)

            wd_wc_description = str(wd_wc_description_cell.value).strip() if wd_wc_description_cell.value else ''

            new_three_v_description = f"Comisión bancaria {wd_wc_description}"
            three_v_description_cell.value = new_three_v_description

            duplicates_processed += 1

            logger.log(
                f"Referencia duplicada procesada: '{reference}' - "
                f"Código {wd_wc_code} → {'T/D' if wd_wc_code == 'WD' else 'T/C'}, "
                f"Código 3V → O/D, "
                f"Descripción 3V actualizada con 'Comisión bancaria'",
                level="INFO",
            )

        if duplicates_processed > 0:
            logger.log(
                f"Se procesaron {duplicates_processed} par(es) de referencias duplicadas.",
                level="INFO",
            )
        else:
            logger.log(
                "No se encontraron referencias duplicadas con los códigos WD/WC y 3V para procesar.",
                level="INFO",
            )

    def _insert_logo(self, worksheet, logger) -> None:
        """Inserta el logo de BAC en la fila 6, columna I"""
        from openpyxl.drawing.image import Image

        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            logo_path = os.path.join(current_dir, 'bac.png')

            if not os.path.exists(logo_path):
                logger.log(
                    f"ADVERTENCIA: No se encontró el archivo 'bac.png' en {current_dir}. "
                    f"El logo no será insertado.",
                    level="WARNING",
                )
                return

            img = Image(logo_path)

            img.width = 120
            img.height = 40

            worksheet.row_dimensions[6].height = 40

            img.anchor = 'I6'
            worksheet.add_image(img)

            logger.log(
                "Logo de BAC insertado correctamente en la fila 6, columna I.",
                level="INFO",
            )

        except Exception as exc:
            logger.log(
                f"Error al insertar el logo de BAC: {exc}",
                level="WARNING",
            )

    def _read_excel_matrix(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
    ) -> Tuple[List[List[Any]], int]:
        """Lee el contenido del Excel original y lo devuelve como matriz"""
        extension = os.path.splitext(original_name)[1].lower()
        rows: List[List[Any]] = []
        max_cols = 0

        if extension == '.xls':
            import xlrd

            workbook = xlrd.open_workbook(file_contents=file_bytes)
            sheet = workbook.sheet_by_index(0)

            for row_idx in range(sheet.nrows):
                row_values: List[Any] = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = datetime(*xlrd.xldate_as_tuple(value, workbook.datemode))
                        except Exception:
                            pass
                    row_values.append(value)
                rows.append(row_values)
                max_cols = max(max_cols, len(row_values))
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                current_row = list(row)
                rows.append(current_row)
                max_cols = max(max_cols, len(current_row))

        return rows, max_cols

    def _remove_empty_columns(self, worksheet) -> None:
        """Elimina columnas completamente vacías para evitar espacios innecesarios"""
        max_row = worksheet.max_row
        for col_idx in range(worksheet.max_column, 0, -1):
            has_content = False
            for row_idx in range(1, max_row + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                if isinstance(value, str):
                    if value.strip():
                        has_content = True
                        break
                elif value not in (None, ''):
                    has_content = True
                    break
            if not has_content:
                worksheet.delete_cols(col_idx)

    def _remove_zero_credit_rows(self, worksheet, logger) -> None:
        """Elimina filas sin movimiento (débito y crédito vacíos o en cero)."""
        if worksheet.max_row < 16:
            return

        header_map = self._extract_header_map(worksheet, worksheet.max_column)
        debit_column = header_map.get('debitos')
        credit_column = header_map.get('creditos')

        if not debit_column and not credit_column:
            return

        summary_row = self._find_row_with_text(worksheet, 'Cuadro de Resumen')
        data_end_row = summary_row - 2 if summary_row else worksheet.max_row
        data_start_row = 16

        if data_start_row > data_end_row:
            return

        rows_to_delete: List[int] = []

        def is_zero(amount: Optional[float]) -> bool:
            return amount is None or abs(amount) < 1e-9

        for row_idx in range(data_start_row, data_end_row + 1):
            debit_amount: Optional[float] = None
            credit_amount: Optional[float] = None

            if debit_column:
                debit_value = worksheet.cell(row=row_idx, column=debit_column).value
                debit_amount = self._parse_decimal(debit_value)

            if credit_column:
                credit_value = worksheet.cell(row=row_idx, column=credit_column).value
                credit_amount = self._parse_decimal(credit_value)

            if is_zero(debit_amount) and is_zero(credit_amount):
                rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            worksheet.delete_rows(row_idx)

        if rows_to_delete:
            logger.log(
                (
                    "Se eliminaron "
                    f"{len(rows_to_delete)} fila(s) sin movimiento (débito y crédito en cero)."
                ),
                level="INFO",
            )

    def _apply_styles(self, worksheet, max_cols: int, logger) -> None:
        """Aplica estilos mejorados respetando la estructura proporcionada"""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )

        header_fill = PatternFill('solid', fgColor='C00000')
        header_secondary_fill = PatternFill('solid', fgColor='F8CBAD')
        zebra_fill_a = PatternFill('solid', fgColor='FFFFFF')
        zebra_fill_b = PatternFill('solid', fgColor='F8FBFF')
        section_fill = PatternFill('solid', fgColor='EFF3F9')
        summary_title_fill = PatternFill('solid', fgColor='F8CBAD')

        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = worksheet.cell(row=16, column=1)

        self._ensure_review_column(worksheet)
        max_cols = worksheet.max_column

        for col_idx in range(1, max_cols + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                if cell.value is None:
                    continue
                value_length = len(str(cell.value))
                if value_length > max_length:
                    max_length = value_length
            worksheet.column_dimensions[column_letter].width = max(12, min(max_length + 2, 45))

        if worksheet.max_row >= 2:
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)
            title_cell = worksheet.cell(row=2, column=1)
            title_cell.value = "DETALLE DE MOVIMIENTOS DEL PERÍODO"
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(bold=True, size=16, color='FFFFFF')
            title_cell.fill = header_fill
            worksheet.row_dimensions[2].height = 28

        for row_idx in range(5, min(14, worksheet.max_row + 1)):
            for col_idx in range(1, max_cols + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if row_idx == 5:
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.fill = section_fill
                    cell.alignment = Alignment(vertical='center')
                    if col_idx % 2 == 1 and cell.value not in (None, ''):
                        cell.font = Font(bold=True)
                cell.border = thin_border

        if worksheet.max_row >= 14:
            for col_idx in range(1, max_cols + 1):
                cell = worksheet.cell(row=14, column=col_idx)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = header_fill
                cell.border = thin_border
            worksheet.row_dimensions[14].height = 22

        if worksheet.max_row >= 15:
            for col_idx in range(1, max_cols + 1):
                cell = worksheet.cell(row=15, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = header_secondary_fill
                cell.border = thin_border

        header_map = self._extract_header_map(worksheet, max_cols)
        date_column = header_map.get('fecha')
        debit_column = header_map.get('debitos')
        credit_column = header_map.get('creditos')
        balance_column = header_map.get('balance')

        summary_row = self._find_row_with_text(worksheet, 'Cuadro de Resumen')
        data_end_row = summary_row - 2 if summary_row else worksheet.max_row
        data_start_row = 16 if worksheet.max_row >= 16 else None
        data_last_row = None
        if data_start_row:
            data_last_row = max(data_start_row, data_end_row)

        if data_start_row and debit_column and data_last_row:
            self._update_codes_for_positive_debits(
                worksheet,
                header_map,
                data_start_row,
                data_last_row,
                logger,
            )

        if data_start_row and credit_column and data_last_row:
            self._update_codes_for_non_negative_credits(
                worksheet,
                header_map,
                data_start_row,
                data_last_row,
                logger,
            )

        if data_start_row and data_last_row:
            self._resolve_duplicate_reference_codes(
                worksheet,
                header_map,
                data_start_row,
                data_last_row,
                logger,
            )

        if data_start_row and data_last_row:
            self._override_codes_by_description(
                worksheet,
                header_map,
                data_start_row,
                data_last_row,
                logger,
            )

        if data_start_row and data_last_row:
            zebra_toggle = True
            for row_idx in range(data_start_row, data_last_row + 1):
                current_fill = zebra_fill_a if zebra_toggle else zebra_fill_b
                zebra_toggle = not zebra_toggle
                for col_idx in range(1, max_cols + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if cell.value not in (None, ''):
                        if col_idx == date_column:
                            self._apply_date_format(cell)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif col_idx in {debit_column, credit_column, balance_column}:
                            self._apply_currency_format(cell)
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(vertical='center')
                    else:
                        cell.alignment = Alignment(vertical='center')
                    cell.fill = current_fill

        self._highlight_rows_by_filters(
            worksheet,
            header_map,
            data_start_row,
            data_end_row,
            max_cols,
            logger,
        )

        if summary_row:
            worksheet.merge_cells(
                start_row=summary_row,
                start_column=1,
                end_row=summary_row,
                end_column=max_cols,
            )
            summary_title = worksheet.cell(row=summary_row, column=1)
            summary_title.font = Font(bold=True, size=14, color='FFFFFF')
            summary_title.fill = summary_title_fill
            summary_title.alignment = Alignment(horizontal='center', vertical='center')

            summary_header_row = summary_row + 1
            if summary_header_row <= worksheet.max_row:
                for col_idx in range(1, max_cols + 1):
                    cell = worksheet.cell(row=summary_header_row, column=col_idx)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = header_fill
                    cell.border = thin_border

            summary_data_start = summary_row + 2
            if summary_data_start <= worksheet.max_row:
                zebra_toggle = True
                for row_idx in range(summary_data_start, worksheet.max_row + 1):
                    row_offset = row_idx - summary_data_start
                    highlight_row = row_offset < 1
                    current_fill = header_fill if highlight_row else (
                        zebra_fill_a if zebra_toggle else zebra_fill_b
                    )
                    zebra_toggle = not zebra_toggle
                    for col_idx in range(1, max_cols + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = current_fill
                        cell.border = thin_border
                        if cell.value not in (None, ''):
                            normalized_header = self._normalize_text(
                                worksheet.cell(row=summary_header_row, column=col_idx).value
                                if summary_header_row <= worksheet.max_row
                                else ''
                            )
                            if normalized_header in {'debitos', 'creditos', 'montos', 'monto'}:
                                self._apply_currency_format(cell)
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            else:
                                cell.alignment = Alignment(vertical='center')
                        if highlight_row:
                            cell.font = Font(bold=True, color='FFFFFF')

    def _update_codes_for_positive_debits(
            self,
            worksheet,
            header_map: Dict[str, int],
            start_row: int,
            end_row: int,
            logger,
    ) -> None:
        """Actualiza los códigos únicamente cuando el débito es mayor a cero."""
        code_column = header_map.get('codigo')
        debit_column = header_map.get('debitos')
        credit_column = header_map.get('creditos')

        if not code_column or not debit_column:
            return

        replacement_map = self.config_manager.get_positive_debit_code_map(
            self._get_config_case_key()
        )
        if not replacement_map:
            return

        updates = 0
        for row_idx in range(start_row, end_row + 1):
            debit_value = worksheet.cell(row=row_idx, column=debit_column).value
            debit_amount = self._parse_decimal(debit_value)

            credit_amount: Optional[float] = None
            if credit_column:
                credit_value = worksheet.cell(row=row_idx, column=credit_column).value
                credit_amount = self._parse_decimal(credit_value)

            if not self._is_positive(debit_amount):
                continue

            if self._is_positive(credit_amount):
                continue

            code_cell = worksheet.cell(row=row_idx, column=code_column)
            if code_cell.value in (None, ''):
                continue

            current_code = str(code_cell.value).strip().upper()
            new_code = replacement_map.get(current_code)

            if new_code and code_cell.value != new_code:
                code_cell.value = new_code
                updates += 1

        if updates:
            logger.log(
                f"Se actualizaron {updates} código(s) por reglas de débitos positivos.",
                level='INFO',
            )

    def _update_codes_for_non_negative_credits(
            self,
            worksheet,
            header_map: Dict[str, int],
            start_row: int,
            end_row: int,
            logger,
    ) -> None:
        """Actualiza los códigos únicamente cuando el crédito es mayor a cero."""
        code_column = header_map.get('codigo')
        credit_column = header_map.get('creditos')
        debit_column = header_map.get('debitos')

        if not code_column or not credit_column:
            return

        replacement_map = self.config_manager.get_non_negative_credit_code_map(
            self._get_config_case_key()
        )
        if not replacement_map:
            return

        updates = 0
        for row_idx in range(start_row, end_row + 1):
            credit_value = worksheet.cell(row=row_idx, column=credit_column).value
            credit_amount = self._parse_decimal(credit_value)

            debit_amount: Optional[float] = None
            if debit_column:
                debit_value = worksheet.cell(row=row_idx, column=debit_column).value
                debit_amount = self._parse_decimal(debit_value)

            if not self._is_positive(credit_amount):
                continue

            if self._is_positive(debit_amount):
                continue

            code_cell = worksheet.cell(row=row_idx, column=code_column)
            if code_cell.value in (None, ''):
                continue

            current_code = str(code_cell.value).strip().upper()
            new_code = replacement_map.get(current_code)

            if new_code and code_cell.value != new_code:
                code_cell.value = new_code
                updates += 1

        if updates:
            logger.log(
                f"Se actualizaron {updates} código(s) por reglas de créditos positivos.",
                level='INFO',
            )

    def _override_codes_by_description(
            self,
            worksheet,
            header_map: Dict[str, int],
            start_row: int,
            end_row: int,
            logger,
    ) -> None:
        """Sobrescribe códigos cuando la descripción coincide con valores especiales."""
        code_column = header_map.get('codigo')
        description_column = header_map.get('descripcion')

        if not code_column or not description_column:
            return

        rules = self.config_manager.get_description_override_rules(
            self._get_config_case_key()
        )
        normalized_rules = [
            (
                self._normalize_text(rule.get('search_text', '')),
                str(rule.get('code', '')).strip().upper(),
            )
            for rule in rules
            if self._normalize_text(rule.get('search_text', ''))
            and str(rule.get('code', '')).strip()
        ]

        if not normalized_rules:
            return

        overrides = 0
        for row_idx in range(start_row, end_row + 1):
            description_value = worksheet.cell(row=row_idx, column=description_column).value
            if description_value in (None, ''):
                continue

            normalized_description = self._normalize_text(description_value)
            if not normalized_description:
                continue

            for search_text, new_code in normalized_rules:
                if search_text not in normalized_description:
                    continue

                if not new_code:
                    continue

                code_cell = worksheet.cell(row=row_idx, column=code_column)
                current_code = str(code_cell.value).strip().upper() if code_cell.value else ''

                if current_code == new_code:
                    break

                code_cell.value = new_code
                overrides += 1
                break

        if overrides:
            logger.log(
                (
                    "Se actualizaron "
                    f"{overrides} código(s) según las reglas configuradas por coincidencia de descripción."
                ),
                level='INFO',
            )

    def _resolve_duplicate_reference_codes(
            self,
            worksheet,
            header_map: Dict[str, int],
            start_row: int,
            end_row: int,
            logger,
    ) -> None:
        """Actualiza códigos según el monto del débito en referencias duplicadas con código PP."""
        code_column = header_map.get('codigo')
        reference_column = header_map.get('referencia')
        debit_column = header_map.get('debitos')

        if not code_column or not reference_column or not debit_column:
            return

        grouped_rows: Dict[str, List[Tuple[int, Optional[float]]]] = {}

        for row_idx in range(start_row, end_row + 1):
            code_cell = worksheet.cell(row=row_idx, column=code_column)
            code_value = code_cell.value
            if code_value in (None, ''):
                continue

            if str(code_value).strip().upper() != 'PP':
                continue

            reference_value = worksheet.cell(row=row_idx, column=reference_column).value
            if reference_value in (None, ''):
                continue

            debit_value = worksheet.cell(row=row_idx, column=debit_column).value
            debit_amount = self._parse_decimal(debit_value)

            reference_key = str(reference_value).strip()
            grouped_rows.setdefault(reference_key, []).append((row_idx, debit_amount))

        updates = 0

        for reference_key, entries in grouped_rows.items():
            valid_entries = [entry for entry in entries if entry[1] is not None]

            if len(valid_entries) < 2:
                if len(valid_entries) == 1:
                    row_idx, debit_amount = valid_entries[0]
                    if self._is_positive(debit_amount):
                        code_cell = worksheet.cell(row=row_idx, column=code_column)
                        if str(code_cell.value).strip().upper() != 'T/D':
                            code_cell.value = 'T/D'
                            updates += 1
                continue

            sorted_entries = sorted(valid_entries, key=lambda item: (item[1], item[0]))

            lowest_entry = sorted_entries[0]
            highest_entry = sorted_entries[-1]

            if highest_entry[0] == lowest_entry[0]:
                continue

            highest_cell = worksheet.cell(row=highest_entry[0], column=code_column)
            lowest_cell = worksheet.cell(row=lowest_entry[0], column=code_column)

            if str(highest_cell.value).strip().upper() != 'T/D':
                highest_cell.value = 'T/D'
                updates += 1

            if str(lowest_cell.value).strip().upper() != 'O/D':
                lowest_cell.value = 'O/D'
                updates += 1

        if updates:
            logger.log(
                (
                    "Se ajustaron "
                    f"{updates} código(s) por referencias duplicadas con montos de débito mayor y menor."
                ),
                level='INFO',
            )

    def _apply_currency_format(self, cell) -> None:
        """Intenta convertir el valor a número y aplicar formato monetario"""
        number = self._parse_decimal(cell.value)
        if number is not None:
            cell.value = number
            cell.number_format = '#,##0.00'

    def _apply_date_format(self, cell) -> None:
        """Intenta convertir el valor a una fecha con formato dd/mm/yyyy"""
        value = cell.value
        if isinstance(value, datetime):
            cell.number_format = 'dd/mm/yyyy'
            return

        if isinstance(value, (int, float)) and value > 40000:
            try:
                base_date = datetime(1899, 12, 30)
                converted = base_date + timedelta(days=float(value))
                cell.value = converted
                cell.number_format = 'dd/mm/yyyy'
                return
            except Exception:
                pass

        if isinstance(value, str):
            cleaned = value.strip()
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
                try:
                    parsed = datetime.strptime(cleaned, fmt)
                    cell.value = parsed
                    cell.number_format = 'dd/mm/yyyy'
                    return
                except ValueError:
                    continue

    @staticmethod
    def _is_positive(amount: Optional[float]) -> bool:
        """Determina si un monto es estrictamente mayor a cero con tolerancia numérica."""
        return amount is not None and amount > 1e-9

    def _parse_decimal(self, value: Any) -> Optional[float]:
        """Convierte cadenas con separadores en valores numéricos"""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if not text or text in {'-', '--'}:
                return None
            text = text.replace('\xa0', '').replace(' ', '')
            text = re.sub(r'[^0-9,.-]', '', text)
            if not text:
                return None

            if text.count(',') > 1 and '.' not in text:
                last_comma = text.rfind(',')
                text = text[:last_comma].replace(',', '') + '.' + text[last_comma + 1:]
            elif text.count('.') > 1 and ',' not in text:
                last_dot = text.rfind('.')
                text = text[:last_dot].replace('.', '') + '.' + text[last_dot + 1:]
            elif text.count(',') == 1 and '.' not in text:
                text = text.replace(',', '.')
            elif text.count('.') == 1 and text.count(',') == 1:
                if text.rfind('.') < text.rfind(','):
                    text = text.replace('.', '')
                    text = text.replace(',', '.')
                else:
                    text = text.replace(',', '')

            try:
                return float(text)
            except ValueError:
                return None
        return None

    def _extract_header_map(self, worksheet, max_cols: int) -> Dict[str, int]:
        """Genera un mapa de encabezados normalizados y su columna"""
        header_row = 14
        headers: Dict[str, int] = {}
        if worksheet.max_row < header_row:
            return headers

        for col_idx in range(1, max_cols + 1):
            cell_value = worksheet.cell(row=header_row, column=col_idx).value
            if isinstance(cell_value, str):
                normalized = self._normalize_text(cell_value)
                if normalized:
                    headers[normalized] = col_idx
        return headers

    def _highlight_rows_by_filters(
            self,
            worksheet,
            header_map: Dict[str, int],
            start_row: Optional[int],
            end_row: Optional[int],
            max_cols: int,
            logger,
    ) -> None:
        """Resalta las filas cuyas descripciones coincidan con filtros del Caso 1."""
        if not start_row or not end_row or start_row > end_row:
            return

        filters = self.config_manager.get_case1_filters()
        if not filters:
            return

        normalized_filters = [
            self._normalize_text(filter_text)
            for filter_text in filters
            if self._normalize_text(filter_text)
        ]
        if not normalized_filters:
            return

        description_column = header_map.get('descripcion')
        if not description_column:
            logger.log(
                "No se encontró una columna de descripción para aplicar los filtros del Caso 1.",
                level="WARNING",
            )
            return

        from openpyxl.styles import Alignment, PatternFill

        highlight_fill = PatternFill(fill_type='solid', fgColor='FFF3B0')
        highlighted_rows = 0
        review_column = header_map.get('revisar')

        for row_idx in range(start_row, end_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=description_column).value
            if cell_value in (None, ''):
                continue

            normalized_value = self._normalize_text(str(cell_value))
            if not normalized_value:
                continue

            if any(filter_text in normalized_value for filter_text in normalized_filters):
                for col_idx in range(1, max_cols + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = highlight_fill
                if review_column:
                    review_cell = worksheet.cell(row=row_idx, column=review_column)
                    review_cell.value = 'Revisar'
                    review_cell.alignment = Alignment(horizontal='center', vertical='center')
                highlighted_rows += 1

        if highlighted_rows:
            logger.log(
                (
                    "Se resaltaron "
                    f"{highlighted_rows} fila(s) que coinciden con los filtros configurados del Caso 1."
                ),
                level="INFO",
            )

    def _ensure_review_column(self, worksheet) -> Optional[int]:
        """Asegura la creación de la columna 'Revisar' contigua a 'Créditos'."""
        header_row = 14
        if worksheet.max_row < header_row:
            return None

        header_map = self._extract_header_map(worksheet, worksheet.max_column)
        review_column = header_map.get('revisar')
        if review_column:
            return review_column

        credit_column = header_map.get('creditos')
        if not credit_column:
            return None

        insert_position = credit_column + 1
        worksheet.insert_cols(insert_position)
        worksheet.cell(row=header_row, column=insert_position, value='Revisar')
        return insert_position

    def _normalize_text(self, text: Any) -> str:
        """Normaliza texto eliminando acentos y espacios"""
        if not isinstance(text, str):
            return ''
        normalized = unicodedata.normalize('NFKD', text)
        normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
        return normalized.lower().strip()

    def _get_config_case_key(self) -> str:
        """Devuelve la clave del caso para acceder a la configuración dinámica."""
        return getattr(self, 'config_case_key', 'case1')

    def _find_row_with_text(self, worksheet, text: str) -> Optional[int]:
        """Busca la fila que contiene el texto indicado"""
        if not text:
            return None
        normalized_target = self._normalize_text(text)
        for row_idx in range(1, worksheet.max_row + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if isinstance(cell_value, str) and self._normalize_text(cell_value) == normalized_target:
                    return row_idx
        return None

    def _extract_date_range(self, subject: str) -> Optional[Tuple[datetime, datetime]]:
        """Obtiene el rango de fechas (dd/mm/yyyy) presente en el asunto del correo."""
        if not subject:
            return None

        matches = re.findall(r"(\d{2}/\d{2}/\d{4})", subject)
        if len(matches) < 2:
            return None

        try:
            start = datetime.strptime(matches[0], "%d/%m/%Y")
            end = datetime.strptime(matches[1], "%d/%m/%Y")
            return start, end
        except ValueError:
            return None

    def _locate_date_column(self, header_map: Dict[str, int]) -> Optional[int]:
        """Identifica la columna que corresponde a fechas dentro del detalle."""
        for header, index in header_map.items():
            if 'fecha' in header:
                return index
        return None

    def _filter_rows_by_date_range(
            self,
            worksheet,
            date_range: Optional[Tuple[datetime, datetime]],
            logger,
    ) -> None:
        """Elimina filas cuyo valor de fecha esté fuera del rango solicitado."""
        if not date_range or worksheet.max_row < 16:
            return

        start, end = date_range
        if start > end:
            start, end = end, start

        header_map = self._extract_header_map(worksheet, worksheet.max_column)
        date_column = self._locate_date_column(header_map)

        if not date_column:
            logger.log(
                "No se encontró una columna de fecha para aplicar el filtrado.",
                level="WARNING",
            )
            return

        summary_row = self._find_row_with_text(worksheet, 'Cuadro de Resumen')
        data_end_row = summary_row - 2 if summary_row else worksheet.max_row
        data_start_row = 16

        if data_start_row > data_end_row:
            return

        rows_to_delete: List[int] = []
        rows_in_range = 0
        for row_idx in range(data_start_row, data_end_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=date_column).value
            date_value = self._parse_date_value(cell_value)
            if date_value is None:
                continue

            if not (start.date() <= date_value.date() <= end.date()):
                rows_to_delete.append(row_idx)
            else:
                rows_in_range += 1

        for row_idx in reversed(rows_to_delete):
            worksheet.delete_rows(row_idx)

        if rows_to_delete:
            logger.log(
                f"Se eliminaron {len(rows_to_delete)} fila(s) fuera del rango de fechas.",
                level="INFO",
            )

        if rows_in_range == 0:
            formatted_start = start.strftime('%d/%m/%Y')
            formatted_end = end.strftime('%d/%m/%Y')
            logger.log(
                (
                    "No se encontraron movimientos dentro del rango de fechas "
                    f"{formatted_start} - {formatted_end}."
                ),
                level="WARNING",
            )

    def _build_output_filename(self, original_name: str, product_name: Optional[str] = None) -> str:
        """Construye el nombre del archivo procesado"""
        base, _ = os.path.splitext(original_name)
        if product_name:
            sanitized = self._sanitize_filename_component(product_name)
            if sanitized:
                base = sanitized
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base}_formateado_{timestamp}.xlsx"

    def _build_summary_filename(self, formatted_name: str) -> str:
        """Construye el nombre del archivo resumen basado en el archivo formateado."""
        base, extension = os.path.splitext(formatted_name)
        marker = '_formateado_'
        if marker in base:
            prefix, suffix = base.split(marker, 1)
            return f"{prefix}_contable_{suffix}{extension}"
        return f"{base}_contable{extension}"

    def _create_summary_workbook(
            self,
            formatted_bytes: bytes,
            logger,
            account_number: Optional[str] = None,
    ) -> Optional[bytes]:
        """Genera el archivo resumen contable con las columnas solicitadas."""
        try:
            from openpyxl import Workbook, load_workbook

            workbook = load_workbook(io.BytesIO(formatted_bytes), data_only=True)
            sheet = workbook.active

            header_map = self._extract_header_map(sheet, sheet.max_column)
            code_column = header_map.get('codigo')
            reference_column = header_map.get('referencia')
            debit_column = header_map.get('debitos')
            credit_column = header_map.get('creditos')
            date_column = header_map.get('fecha')

            if not all([code_column, reference_column, date_column]) or (
                    not debit_column and not credit_column
            ):
                logger.log(
                    (
                        "No se pudo generar el resumen contable porque faltan columnas "
                        "requeridas (Código, Referencia, Fecha, Débitos/Créditos)."
                    ),
                    level="WARNING",
                )
                return None

            summary_wb = Workbook()
            summary_ws = summary_wb.active
            summary_ws.title = "Movimientos"

            headers = [
                "Cuenta Bancaria",
                "Tipo Documento",
                "Número",
                "Monto",
                "Fecha documento",
            ]
            summary_ws.append(headers)

            summary_row = self._find_row_with_text(sheet, 'Cuadro de Resumen')
            data_end_row = summary_row - 2 if summary_row else sheet.max_row
            data_start_row = 16 if sheet.max_row >= 16 else 1

            for row_idx in range(data_start_row, data_end_row + 1):
                code_value = sheet.cell(row=row_idx, column=code_column).value
                reference_value = sheet.cell(row=row_idx, column=reference_column).value
                date_value = sheet.cell(row=row_idx, column=date_column).value

                debit_amount = (
                    self._parse_decimal(sheet.cell(row=row_idx, column=debit_column).value)
                    if debit_column
                    else None
                )
                credit_amount = (
                    self._parse_decimal(sheet.cell(row=row_idx, column=credit_column).value)
                    if credit_column
                    else None
                )

                amount = None
                if self._is_positive(credit_amount):
                    amount = credit_amount
                elif self._is_positive(debit_amount):
                    amount = debit_amount
                elif credit_amount is not None and abs(credit_amount) > 1e-9:
                    amount = credit_amount
                elif debit_amount is not None and abs(debit_amount) > 1e-9:
                    amount = debit_amount

                if amount is None:
                    continue

                parsed_date = self._parse_date_value(date_value)

                summary_row_values = [
                    account_number or '',
                    str(code_value).strip() if code_value not in (None, '') else '',
                    reference_value if reference_value not in (None, '') else '',
                    amount,
                    parsed_date if parsed_date else date_value,
                ]
                summary_ws.append(summary_row_values)

            for column in summary_ws.iter_cols(min_row=2, max_col=4, max_row=summary_ws.max_row):
                for cell in column:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

            date_column_idx = headers.index("Fecha documento") + 1
            for cell in summary_ws.iter_cols(
                    min_col=date_column_idx,
                    max_col=date_column_idx,
                    min_row=2,
                    max_row=summary_ws.max_row,
            ):
                for date_cell in cell:
                    if isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'dd/mm/yyyy'

            output = io.BytesIO()
            summary_wb.save(output)
            output.seek(0)
            return output.read()
        except ImportError as dependency_error:
            logger.log(
                f"Dependencia faltante para crear el resumen contable: {dependency_error}",
                level="ERROR",
            )
            return None
        except Exception as exc:
            logger.log(
                f"Error inesperado al generar el resumen contable: {exc}",
                level="ERROR",
            )
            return None

    def _sanitize_filename_component(self, text: str) -> str:
        """Sanitiza texto para ser utilizado como parte de un nombre de archivo"""
        if not isinstance(text, str):
            return ''
        sanitized = text.strip()
        sanitized = unicodedata.normalize('NFKD', sanitized)
        sanitized = ''.join(ch for ch in sanitized if not unicodedata.combining(ch))
        sanitized = re.sub(r"[^A-Za-z0-9._ -]+", "_", sanitized)
        sanitized = re.sub(r"_+", "_", sanitized)
        sanitized = sanitized.strip(' _')
        return sanitized[:100]

    def _extract_product_name(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
    ) -> Optional[str]:
        """Extrae el nombre del producto desde la fila 7, columna B del archivo"""
        try:
            extension = os.path.splitext(original_name)[1].lower()
            if extension == '.xls':
                import xlrd

                workbook = xlrd.open_workbook(file_contents=file_bytes)
                sheet = workbook.sheet_by_index(0)
                if sheet.nrows >= 7 and sheet.ncols >= 2:
                    value = sheet.cell_value(6, 1)
                else:
                    return None
            else:
                from openpyxl import load_workbook

                workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
                sheet = workbook.active
                cell = sheet.cell(row=7, column=2)
                value = cell.value

            if value is None:
                return None

            if isinstance(value, str):
                cleaned = value.strip()
            else:
                cleaned = str(value).strip()

            return cleaned or None
        except Exception as exc:
            logger.log(
                f"No se pudo extraer el nombre del producto del archivo '{original_name}': {exc}",
                level="WARNING",
            )
            return None

    def _extract_account_number(self, product_value: Optional[str]) -> Optional[str]:
        """Obtiene el número de cuenta bancaria a partir del valor de producto."""
        if not product_value:
            return None

        text = str(product_value).strip()
        if not text:
            return None

        if ':' in text:
            text = text.split(':', 1)[1].strip()

        tokens = text.split()
        if len(tokens) > 1:
            text = tokens[-1]

        condensed = ''.join(ch for ch in text if not ch.isspace())
        if len(condensed) <= 13:
            return None

        account = condensed[12:-1]
        account = account.strip()
        return account or None

    def _parse_date_value(self, value: Any) -> Optional[datetime]:
        """Intenta convertir diferentes formatos de fecha a un objeto datetime."""
        if isinstance(value, datetime):
            return value

        if isinstance(value, (int, float)):
            try:
                base_date = datetime(1899, 12, 30)
                converted = base_date + timedelta(days=float(value))
                if 1900 <= converted.year <= 9999:
                    return converted
            except Exception:
                pass

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue

        return None