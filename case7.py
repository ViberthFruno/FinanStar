# Archivo: case7.py
# Ubicación: raíz del proyecto
# Descripción: Caso 7 - Genera un reporte verde de transacciones con totales de débitos y créditos

import io
import os
import re
import unicodedata
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple
from zipfile import BadZipFile

from config_manager import ConfigManager


class Case:
    """Caso 7 - Rediseña el estado de cuenta en un formato verde con totales."""

    HEADERS: Tuple[str, ...] = (
        "Fecha",
        "Documento",
        "Descripción",
        "Débitos",
        "Créditos",
        "Saldo",
    )
    OPTIONAL_HEADERS: Tuple[str, ...] = ("Código",)
    OUTPUT_HEADERS: Tuple[str, ...] = HEADERS + OPTIONAL_HEADERS + ("Revisar",)

    def __init__(self) -> None:
        self.name = "Caso 7"
        self.description = (
            "Recibe archivos Excel con movimientos bancarios, los reorganiza en un nuevo formato "
            "con encabezado verde y calcula los totales de débitos y créditos."
        )
        self.response_message = (
            "Hola,\n\nSe adjunta el archivo con el formato verde actualizado del estado de cuenta "
            "correspondiente al Caso 7. Quedo atento a cualquier comentario.\n\nSaludos cordiales."
        )
        self.corrupted_file_message = (
            "Hola,\n\n"
            "Lamentablemente, el archivo Excel que nos envió está corrupto y no se puede procesar correctamente. "
            "Este problema ocurre ocasionalmente con archivos descargados directamente del sistema bancario.\n\n"
            "Para solucionarlo, por favor siga estos pasos:\n\n"
            "1. Abra el archivo Excel en su computadora\n"
            "2. Una vez abierto, vaya a 'Archivo' > 'Guardar como'\n"
            "3. Guarde el archivo con un nuevo nombre\n"
            "4. Envíenos este nuevo archivo guardado\n\n"
            "Este proceso reparará la estructura del archivo y permitirá que podamos procesarlo sin problemas.\n\n"
            "Quedamos atentos al reenvío del archivo.\n\n"
            "Saludos cordiales."
        )
        self.config_manager = ConfigManager()
        self.config_case_key = 'case7'

    # ==================== MÉTODOS PÚBLICOS ====================

    def get_name(self) -> str:
        return self.name

    def get_description(self) -> str:
        return self.description

    def get_search_keywords(self) -> List[str]:
        try:
            config = self.config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso7', '').strip()
            if keyword:
                return [keyword]
            return []
        except Exception as exc:
            print(f"Error al cargar palabras clave para caso7: {exc}")
            return []

    def process_email(self, email_data: Dict[str, Any], logger) -> Optional[Dict[str, Any]]:
        try:
            sender = email_data.get('sender', '')
            subject = email_data.get('subject', '')
            attachments = email_data.get('attachments', [])

            logger.log(
                f"Procesando {self.name} para email de {sender} con asunto: {subject}",
                level="INFO",
            )

            # Extraer rango de fechas del asunto
            date_range = self._extract_date_range(subject)
            if date_range:
                start, end = date_range
                logger.log(
                    f"Se aplicará un filtrado de fechas desde {start.strftime('%d/%m/%Y')} "
                    f"hasta {end.strftime('%d/%m/%Y')}",
                    level="INFO",
                )
            else:
                logger.log(
                    "No se encontró un rango de fechas válido en el asunto. "
                    "Se conservarán todos los movimientos.",
                    level="INFO",
                )

            excel_attachments = [
                attachment
                for attachment in attachments
                if self._is_excel_file(attachment.get('filename'))
            ]

            if not excel_attachments:
                logger.log(
                    "No se encontraron adjuntos de Excel para procesar en el correo recibido.",
                    level="WARNING",
                )
                return None

            processed_files: List[Dict[str, Any]] = []
            corrupted_files: List[str] = []

            for attachment in excel_attachments:
                result = self._redesign_excel_attachment(attachment, logger, subject, date_range)

                if result == 'CORRUPTED':
                    filename = attachment.get('filename', 'archivo desconocido')
                    corrupted_files.append(filename)
                    logger.log(
                        f"Archivo corrupto detectado: {filename}",
                        level="WARNING",
                    )
                elif result:
                    processed_files.extend(result)

            # Si hay archivos corruptos, enviar mensaje especial
            if corrupted_files:
                logger.log(
                    f"Se detectaron {len(corrupted_files)} archivo(s) corrupto(s). "
                    f"Enviando respuesta con instrucciones de solución.",
                    level="INFO",
                )

                response_data = {
                    'recipient': sender,
                    'subject': f"Re: {subject}",
                    'body': self.corrupted_file_message,
                    'attachments': [],
                }

                return response_data

            # Si no hay archivos procesados exitosamente
            if not processed_files:
                logger.log(
                    "No fue posible generar el archivo rediseñado para los adjuntos proporcionados.",
                    level="ERROR",
                )
                return None

            # Respuesta exitosa con archivos procesados
            response_data = {
                'recipient': sender,
                'subject': f"Re: {subject}",
                'body': self.response_message,
                'attachments': processed_files,
            }

            logger.log(
                f"Respuesta generada para {self.name} con {len(processed_files)} adjunto(s).",
                level="INFO",
            )

            return response_data

        except Exception as exc:
            logger.log(f"Error al procesar email en {self.name}: {exc}", level="ERROR")
            return None

    def get_response_message(self) -> str:
        return self.response_message

    # ==================== MÉTODOS INTERNOS ====================

    def _is_excel_file(self, filename: Optional[str]) -> bool:
        if not filename:
            return False
        extension = os.path.splitext(filename)[1].lower()
        return extension in {'.xls', '.xlsx'}

    def _extract_date_range(self, subject: str) -> Optional[Tuple[datetime, datetime]]:
        """Extrae el rango de fechas (dd/mm/yyyy) presente en el asunto del correo."""
        if not subject:
            return None

        matches = re.findall(r"(\d{1,2}/\d{1,2}/\d{4})", subject)
        if len(matches) < 2:
            return None

        try:
            start = datetime.strptime(matches[0], "%d/%m/%Y")
            end = datetime.strptime(matches[1], "%d/%m/%Y")
            return start, end
        except ValueError:
            return None

    def _redesign_excel_attachment(
            self,
            attachment: Dict[str, Any],
            logger,
            subject: str,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[List[Dict[str, Any]]]:
        filename = attachment.get('filename') or 'reporte.xlsx'
        content = attachment.get('content')

        if not content:
            logger.log(f"El adjunto '{filename}' está vacío o no pudo leerse.", level="WARNING")
            return None

        try:
            workbook_result = self._create_redesigned_workbook(
                content,
                filename,
                logger,
                subject,
                date_range
            )

            if workbook_result == 'CORRUPTED':
                return 'CORRUPTED'

            if not workbook_result:
                return None

            workbook_bytes, summary_bytes = workbook_result

            output_name = self._build_output_filename(filename)
            attachments_list: List[Dict[str, Any]] = [
                {
                    'filename': output_name,
                    'content': workbook_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                }
            ]

            if summary_bytes:
                summary_name = self._build_summary_filename(output_name)
                attachments_list.append(
                    {
                        'filename': summary_name,
                        'content': summary_bytes,
                        'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    }
                )

            return attachments_list

        except Exception as exc:
            logger.log(
                f"Error inesperado al rediseñar el archivo '{filename}': {exc}",
                level="ERROR",
            )
            return None

    def _create_redesigned_workbook(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            subject: str,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[Tuple[bytes, Optional[bytes]]]:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        import warnings

        try:
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                source_wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        except BadZipFile:
            logger.log(
                f"El archivo '{original_name}' está corrupto (no es un archivo ZIP válido). "
                f"Se enviará una respuesta con instrucciones para solucionar el problema.",
                level="WARNING",
            )
            return 'CORRUPTED'
        except Exception as exc:
            error_message = str(exc).lower()
            if 'zip' in error_message or 'corrupt' in error_message or 'not a zip file' in error_message:
                logger.log(
                    f"El archivo '{original_name}' parece estar corrupto: {exc}. "
                    f"Se enviará una respuesta con instrucciones para solucionar el problema.",
                    level="WARNING",
                )
                return 'CORRUPTED'

            logger.log(
                f"No fue posible abrir el archivo '{original_name}' para rediseño: {exc}",
                level="ERROR",
            )
            return None

        source_ws = source_wb.active

        metadata = self._extract_metadata(source_ws, logger)
        if not metadata['date_range']:
            metadata['date_range'] = self._build_date_range_from_subject(subject)

        data_rows = self._extract_table_rows(source_ws, logger)

        if not data_rows:
            logger.log(
                f"El archivo '{original_name}' no contiene datos de movimientos para rediseñar.",
                level="WARNING",
            )
            return None

        # Aplicar filtrado por fechas si existe un rango
        if date_range:
            data_rows = self._filter_data_rows_by_date_range(data_rows, date_range, logger)

            if not data_rows:
                start, end = date_range
                logger.log(
                    f"No se encontraron movimientos dentro del rango de fechas "
                    f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}.",
                    level="WARNING",
                )
                return None

        if data_rows:
            self._assign_codes_by_description(data_rows, logger)
            self._apply_code_replacement_rules(data_rows, logger)

        wb = Workbook()
        ws = wb.active
        ws.title = "Transacciones"

        # Encabezados principales
        ws.cell(row=2, column=1, value=metadata['title'] or 'TRANSACCIONES POR FECHA')
        ws.cell(row=3, column=1, value=metadata['bank'] or 'Banco Promerica Costa Rica')
        ws.cell(row=4, column=1, value=metadata['report_date'] or '')
        ws.cell(row=5, column=1, value=metadata['date_range'] or '')
        ws.cell(row=6, column=1, value=metadata['account'] or '')

        header_row = 8
        output_headers = list(self.OUTPUT_HEADERS)
        column_map = {header: idx for idx, header in enumerate(output_headers, start=1)}
        total_columns = len(output_headers)

        for header, col_idx in column_map.items():
            ws.cell(row=header_row, column=col_idx, value=header)

        data_start = header_row + 1
        for row_offset, row_data in enumerate(data_rows):
            current_row = data_start + row_offset
            row_data.setdefault('Código', '')
            row_data.setdefault('Revisar', '')

            for header in output_headers:
                col_idx = column_map[header]
                cell = ws.cell(row=current_row, column=col_idx)
                value = row_data.get(header)

                if header == 'Fecha':
                    parsed_date = self._parse_date_from_value(value)
                    cell.value = parsed_date if parsed_date else value
                elif header in {'Débitos', 'Créditos', 'Saldo'}:
                    number = self._to_number(value)
                    cell.value = number if value not in (None, '') else None
                else:
                    cell.value = '' if value is None else value

        # Estilos
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=12)
        header_font = Font(bold=True, color='FFFFFF')
        regular_font = Font(size=11)
        header_fill = PatternFill(fill_type='solid', fgColor='00843D')  # Verde
        thin_border = Border(
            left=Side(border_style='thin', color='B0B0B0'),
            right=Side(border_style='thin', color='B0B0B0'),
            top=Side(border_style='thin', color='B0B0B0'),
            bottom=Side(border_style='thin', color='B0B0B0'),
        )

        ws.cell(row=2, column=1).font = title_font
        ws.cell(row=3, column=1).font = subtitle_font
        for row_idx in range(4, 7):
            cell = ws.cell(row=row_idx, column=1)
            cell.font = regular_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        numeric_columns = {
            'Débitos': column_map.get('Débitos'),
            'Créditos': column_map.get('Créditos'),
            'Saldo': column_map.get('Saldo'),
        }
        numeric_column_indices = {idx for idx in numeric_columns.values() if idx}

        date_column = column_map.get('Fecha')

        data_end = data_start + len(data_rows) - 1
        for row_idx in range(data_start, data_end + 1):
            for header in output_headers:
                col_idx = column_map[header]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx in numeric_column_indices:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif date_column and col_idx == date_column:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif header == 'Revisar':
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        if data_rows:
            self._highlight_rows_by_filters(
                ws,
                column_map,
                data_start,
                data_end,
                total_columns,
                logger,
            )

        ws.freeze_panes = 'A9'

        total_columns = len(output_headers)

        for col_idx in range(1, total_columns + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value is None:
                    continue
                cell_value = cell.value
                if isinstance(cell_value, (int, float)):
                    text = f"{cell_value:,.2f}" if col_idx in numeric_column_indices else str(cell_value)
                elif isinstance(cell_value, datetime):
                    text = cell_value.strftime('%d/%m/%Y')
                else:
                    text = str(cell_value)
                if len(text) > max_length:
                    max_length = len(text)
            ws.column_dimensions[column_letter].width = min(max_length + 4, 45)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        workbook_bytes = output.read()

        summary_bytes = self._create_summary_workbook(data_rows, metadata, logger)

        return workbook_bytes, summary_bytes

    def _filter_data_rows_by_date_range(
            self,
            data_rows: List[Dict[str, Any]],
            date_range: Tuple[datetime, datetime],
            logger,
    ) -> List[Dict[str, Any]]:
        """Filtra las filas de datos según el rango de fechas proporcionado."""
        if not date_range:
            return data_rows

        start, end = date_range
        if start > end:
            start, end = end, start

        filtered_rows: List[Dict[str, Any]] = []
        rows_filtered_out = 0

        for row_data in data_rows:
            date_value = row_data.get('Fecha')
            parsed_date = self._parse_date_from_value(date_value)

            if parsed_date is None:
                # Si no se puede parsear la fecha, incluir la fila
                filtered_rows.append(row_data)
                continue

            # Verificar si la fecha está dentro del rango
            if start.date() <= parsed_date.date() <= end.date():
                filtered_rows.append(row_data)
            else:
                rows_filtered_out += 1

        if rows_filtered_out > 0:
            logger.log(
                f"Se filtraron {rows_filtered_out} fila(s) fuera del rango de fechas "
                f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}.",
                level="INFO",
            )

        return filtered_rows

    def _assign_codes_by_description(
            self,
            data_rows: List[Dict[str, Any]],
            logger,
    ) -> None:
        """Asigna códigos basados en reglas configuradas para la descripción."""
        if not data_rows:
            return

        codification_rules = self._get_codification_rules()
        assigned_count = 0

        for row_data in data_rows:
            code = self._determine_codification(row_data, codification_rules)
            if code:
                row_data['Código'] = code
                assigned_count += 1
            else:
                row_data['Código'] = row_data.get('Código', '') or ''

        if assigned_count:
            logger.log(
                f"Se asignaron códigos automáticamente a {assigned_count} fila(s) según las reglas configuradas.",
                level="INFO",
            )

    def _get_codification_rules(self) -> Dict[str, List[Tuple[str, str]]]:
        """Obtiene y prepara las reglas de codificación para el Caso 7."""
        raw_rules = self.config_manager.get_case7_codification_rules()
        prepared: Dict[str, List[Tuple[str, str]]] = {'debit': [], 'credit': []}

        for key in ('debit', 'credit'):
            entries = raw_rules.get(key, [])
            if not isinstance(entries, list):
                continue
            for item in entries:
                if not isinstance(item, dict):
                    continue
                search_text = item.get('search_text', '')
                code = item.get('code', '')
                if not isinstance(search_text, str) or not isinstance(code, str):
                    continue
                normalized_search = self._normalize_text(search_text)
                if normalized_search and code.strip():
                    prepared[key].append((normalized_search, code.strip()))

        return prepared

    def _determine_codification(
            self,
            row_data: Dict[str, Any],
            codification_rules: Dict[str, List[Tuple[str, str]]],
    ) -> str:
        """Determina el código a asignar a la fila según las reglas disponibles."""
        description = row_data.get('Descripción')
        if not isinstance(description, str):
            return ''

        normalized_description = self._normalize_text(description)
        if not normalized_description:
            return ''

        credit_amount = self._to_number(row_data.get('Créditos'))
        debit_amount = self._to_number(row_data.get('Débitos'))

        if credit_amount > 0:
            code = self._match_codification(normalized_description, codification_rules.get('credit', []))
            if code:
                return code

        if debit_amount > 0:
            code = self._match_codification(normalized_description, codification_rules.get('debit', []))
            if code:
                return code

        return ''

    def _match_codification(
            self,
            normalized_description: str,
            rules: List[Tuple[str, str]],
    ) -> str:
        """Devuelve el código correspondiente si alguna regla coincide con la descripción."""
        for search_text, code in rules:
            if search_text and code and search_text in normalized_description:
                return code
        return ''

    def _apply_code_replacement_rules(
            self,
            data_rows: List[Dict[str, Any]],
            logger,
    ) -> None:
        """Aplica las reglas configurables de códigos sobre las filas de datos."""
        if not data_rows:
            return

        self._update_codes_for_positive_debits(data_rows, logger)
        self._update_codes_for_non_negative_credits(data_rows, logger)
        self._override_codes_by_description(data_rows, logger)

    def _update_codes_for_positive_debits(
            self,
            data_rows: List[Dict[str, Any]],
            logger,
    ) -> None:
        if not data_rows or 'Código' not in data_rows[0] or 'Débitos' not in data_rows[0]:
            return

        replacement_map = self.config_manager.get_positive_debit_code_map(self.config_case_key)
        if not replacement_map:
            return

        updates = 0
        for row_data in data_rows:
            debit_amount = self._to_number(row_data.get('Débitos'))
            credit_amount = self._to_number(row_data.get('Créditos')) if 'Créditos' in row_data else 0.0

            if debit_amount <= 1e-9:
                continue

            if credit_amount > 1e-9:
                continue

            current_code = str(row_data.get('Código') or '').strip().upper()
            if not current_code:
                continue

            new_code = replacement_map.get(current_code)
            if new_code and current_code != new_code:
                row_data['Código'] = new_code
                updates += 1

        if updates:
            logger.log(
                f"Se actualizaron {updates} código(s) por reglas de débitos positivos.",
                level="INFO",
            )

    def _update_codes_for_non_negative_credits(
            self,
            data_rows: List[Dict[str, Any]],
            logger,
    ) -> None:
        if not data_rows or 'Código' not in data_rows[0] or 'Créditos' not in data_rows[0]:
            return

        replacement_map = self.config_manager.get_non_negative_credit_code_map(self.config_case_key)
        if not replacement_map:
            return

        updates = 0
        for row_data in data_rows:
            credit_amount = self._to_number(row_data.get('Créditos'))
            debit_amount = self._to_number(row_data.get('Débitos')) if 'Débitos' in row_data else 0.0

            if credit_amount <= 1e-9:
                continue

            if debit_amount > 1e-9:
                continue

            current_code = str(row_data.get('Código') or '').strip().upper()
            if not current_code:
                continue

            new_code = replacement_map.get(current_code)
            if new_code and current_code != new_code:
                row_data['Código'] = new_code
                updates += 1

        if updates:
            logger.log(
                f"Se actualizaron {updates} código(s) por reglas de créditos positivos.",
                level="INFO",
            )

    def _override_codes_by_description(
            self,
            data_rows: List[Dict[str, Any]],
            logger,
    ) -> None:
        if not data_rows or 'Código' not in data_rows[0] or 'Descripción' not in data_rows[0]:
            return

        rules = self.config_manager.get_description_override_rules(self.config_case_key)
        normalized_rules = [
            (
                self._normalize_text(rule.get('search_text', '')),
                str(rule.get('code', '')).strip().upper(),
            )
            for rule in rules
            if self._normalize_text(rule.get('search_text', ''))
            and str(rule.get('code', '')).strip()
        ]

        if not normalized_rules:
            return

        overrides = 0
        for row_data in data_rows:
            description_value = row_data.get('Descripción')
            if description_value in (None, ''):
                continue

            normalized_description = self._normalize_text(str(description_value))
            if not normalized_description:
                continue

            for search_text, new_code in normalized_rules:
                if search_text not in normalized_description:
                    continue

                current_code = str(row_data.get('Código') or '').strip().upper()
                if current_code == new_code:
                    break

                row_data['Código'] = new_code
                overrides += 1
                break

        if overrides:
            logger.log(
                (
                    "Se actualizaron "
                    f"{overrides} código(s) según las reglas configuradas por coincidencia de descripción."
                ),
                level="INFO",
            )

    def _highlight_rows_by_filters(
            self,
            worksheet,
            column_map: Dict[str, int],
            start_row: int,
            end_row: int,
            total_columns: int,
            logger,
    ) -> None:
        """Resalta filas que coincidan con los filtros configurados para el Caso 7."""
        filters = self.config_manager.get_case7_filters()
        if not filters:
            return

        normalized_filters = [
            self._normalize_text(filter_text)
            for filter_text in filters
            if self._normalize_text(filter_text)
        ]

        if not normalized_filters:
            return

        description_column = column_map.get('Descripción')
        review_column = column_map.get('Revisar')

        if not description_column:
            logger.log(
                "No se encontró una columna de descripción para aplicar filtros del Caso 7.",
                level="WARNING",
            )
            return

        from openpyxl.styles import Alignment, PatternFill

        highlight_fill = PatternFill(fill_type='solid', fgColor='FFF3B0')
        highlighted_rows = 0

        for row_idx in range(start_row, end_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=description_column).value
            if cell_value in (None, ''):
                continue

            normalized_value = self._normalize_text(str(cell_value))
            if not normalized_value:
                continue

            if any(filter_text in normalized_value for filter_text in normalized_filters):
                for col_idx in range(1, total_columns + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = highlight_fill
                if review_column:
                    review_cell = worksheet.cell(row=row_idx, column=review_column)
                    review_cell.value = 'Revisar'
                    review_cell.alignment = Alignment(horizontal='center', vertical='center')
                highlighted_rows += 1

        if highlighted_rows:
            logger.log(
                (
                    "Se resaltaron "
                    f"{highlighted_rows} fila(s) que coinciden con los filtros configurados del Caso 7."
                ),
                level="INFO",
            )

    def _extract_metadata(self, worksheet, logger) -> Dict[str, str]:
        metadata = {
            'title': '',
            'bank': '',
            'report_date': '',
            'date_range': '',
            'account': '',
        }

        max_row = min(worksheet.max_row, 40)
        max_col = min(worksheet.max_column, 12)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                value = worksheet.cell(row=row, column=col).value
                if not isinstance(value, str):
                    continue

                normalized = self._normalize_text(value)
                adjacent = worksheet.cell(row=row, column=col + 1).value if col + 1 <= max_col else None

                if not metadata['title'] and 'transacciones' in normalized and 'fecha' in normalized:
                    metadata['title'] = self._build_metadata_text(value, adjacent)
                    logger.log(f"Título encontrado en fila {row}, columna {col}: {metadata['title']}", level="INFO")
                elif not metadata['bank'] and 'banco' in normalized:
                    metadata['bank'] = self._build_metadata_text(value, adjacent)
                    logger.log(f"Banco encontrado en fila {row}, columna {col}: {metadata['bank']}", level="INFO")
                elif not metadata['report_date'] and 'fechadelreporte' in normalized:
                    metadata['report_date'] = self._build_metadata_text(value, adjacent)
                    logger.log(f"Fecha de reporte encontrada en fila {row}, columna {col}: {metadata['report_date']}",
                               level="INFO")
                elif not metadata['date_range'] and 'rangodefechas' in normalized:
                    metadata['date_range'] = self._build_metadata_text(value, adjacent)
                    logger.log(f"Rango de fechas encontrado en fila {row}, columna {col}: {metadata['date_range']}",
                               level="INFO")
                elif not metadata['account'] and 'numerodecuenta' in normalized:
                    metadata['account'] = self._build_metadata_text(value, adjacent)
                    logger.log(f"Número de cuenta encontrado en fila {row}, columna {col}: {metadata['account']}",
                               level="INFO")

        return metadata

    def _build_metadata_text(self, label: str, value: Any) -> str:
        label_text = label.strip()
        if value in (None, ''):
            return label_text

        if isinstance(value, datetime):
            value_text = value.strftime('%d/%m/%Y')
        elif isinstance(value, (int, float)):
            value_text = f"{value:,.2f}"
        else:
            value_text = str(value).strip()

        separator = ':' if not label_text.endswith(':') else ''
        return f"{label_text}{separator} {value_text}".strip()

    def _build_date_range_from_subject(self, subject: str) -> str:
        if not subject:
            return ''

        matches = re.findall(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", subject)
        if len(matches) < 2:
            return ''

        parsed: List[datetime] = []
        for item in matches[:2]:
            parsed_date = self._parse_date_string(item)
            if parsed_date:
                parsed.append(parsed_date)

        if len(parsed) < 2:
            return ''

        start, end = parsed[0], parsed[1]
        if start > end:
            start, end = end, start

        return f"Rango de Fechas: {start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"

    def _extract_table_rows(self, worksheet, logger) -> List[Dict[str, Any]]:
        header_row, header_map = self._find_header_row(worksheet)
        if not header_row or not header_map:
            logger.log(
                "No se encontraron encabezados válidos en el archivo fuente para Caso 7.",
                level="ERROR",
            )
            return []

        required_columns = {
            header: header_map.get(self._simplify_header(header))
            for header in self.HEADERS
        }
        optional_columns = {
            header: header_map.get(self._simplify_header(header))
            for header in self.OPTIONAL_HEADERS
        }

        missing = [header for header, col in required_columns.items() if not col]
        if missing:
            logger.log(
                "No se localizaron todas las columnas requeridas para Caso 7: " + ', '.join(missing),
                level="WARNING",
            )

        data_rows: List[Dict[str, Any]] = []
        empty_streak = 0
        row_idx = header_row + 1

        while row_idx <= worksheet.max_row and empty_streak < 3:
            row_data: Dict[str, Any] = {}
            empty = True
            for header in self.HEADERS:
                col_idx = required_columns.get(header)
                value = worksheet.cell(row=row_idx, column=col_idx).value if col_idx else None
                if value not in (None, ''):
                    empty = False
                row_data[header] = value

            for header in self.OPTIONAL_HEADERS:
                col_idx = optional_columns.get(header)
                value = worksheet.cell(row=row_idx, column=col_idx).value if col_idx else ''
                row_data[header] = value

            row_data['Revisar'] = ''

            if empty:
                empty_streak += 1
            else:
                empty_streak = 0
                data_rows.append(row_data)

            row_idx += 1

        return data_rows

    def _find_header_row(self, worksheet) -> Tuple[Optional[int], Dict[str, int]]:
        target_headers = {
            self._simplify_header(header)
            for header in (self.HEADERS + self.OPTIONAL_HEADERS)
        }
        best_row: Optional[int] = None
        best_matches = 0
        header_map: Dict[str, int] = {}

        max_row = min(worksheet.max_row, 80)

        for row_idx in range(1, max_row + 1):
            current_map: Dict[str, int] = {}
            matches = 0
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if not isinstance(cell_value, str):
                    continue
                simplified = self._simplify_header(cell_value)
                if not simplified:
                    continue
                current_map[simplified] = col_idx
                if simplified in target_headers:
                    matches += 1

            if matches > best_matches:
                best_matches = matches
                best_row = row_idx
                header_map = current_map

            if matches == len(target_headers):
                break

        if not best_row or best_matches == 0:
            return None, {}

        return best_row, header_map

    def _simplify_header(self, text: Any) -> str:
        if not isinstance(text, str):
            return ''
        normalized = self._normalize_text(text)
        return re.sub(r'[^a-z0-9]+', '', normalized)

    def _normalize_text(self, text: Any) -> str:
        """Normaliza texto eliminando acentos, espacios y convirtiendo a minúsculas"""
        if not isinstance(text, str):
            return ''
        normalized = unicodedata.normalize('NFKD', text)
        without_accents = ''.join(c for c in normalized if not unicodedata.combining(c))
        result = without_accents.lower().replace(' ', '')
        return result

    def _parse_date_from_value(self, value: Any) -> Optional[datetime]:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        if isinstance(value, (int, float)) and value > 0:
            try:
                base_date = datetime(1899, 12, 30)
                converted = base_date + timedelta(days=float(value))
                if 1900 <= converted.year <= 9999:
                    return converted
            except Exception:
                return None
        if isinstance(value, str):
            return self._parse_date_string(value)
        return None

    def _parse_date_string(self, value: str) -> Optional[datetime]:
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.replace('.', '/').replace('-', '/').replace('\u2013', '/')
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                parsed = datetime.strptime(cleaned, fmt)
                return parsed
            except ValueError:
                continue
        return None

    def _to_number(self, value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return 0.0
            cleaned = cleaned.replace(' ', '')
            if ',' in cleaned and '.' in cleaned:
                if cleaned.rfind(',') > cleaned.rfind('.'):
                    cleaned = cleaned.replace('.', '')
                    cleaned = cleaned.replace(',', '.')
                else:
                    cleaned = cleaned.replace(',', '')
            elif ',' in cleaned:
                cleaned = cleaned.replace(',', '.')
            else:
                cleaned = cleaned.replace(',', '')
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return 0.0

    def _build_output_filename(self, original_name: str) -> str:
        base, _ = os.path.splitext(original_name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base}_caso7_{timestamp}.xlsx"

    def _build_summary_filename(self, formatted_name: str) -> str:
        base, extension = os.path.splitext(formatted_name)
        return f"{base}_resumen{extension}"

    def _create_summary_workbook(
            self,
            data_rows: List[Dict[str, Any]],
            metadata: Dict[str, Any],
            logger,
    ) -> Optional[bytes]:
        if not data_rows:
            return None

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter
        except ImportError as dependency_error:
            logger.log(
                f"Dependencia faltante para crear el resumen contable de Caso 7: {dependency_error}",
                level="ERROR",
            )
            return None
        except Exception as exc:
            logger.log(
                f"No fue posible importar dependencias para el resumen de Caso 7: {exc}",
                level="ERROR",
            )
            return None

        try:
            summary_wb = Workbook()
            summary_ws = summary_wb.active
            summary_ws.title = "Movimientos"

            headers = [
                "Cuenta Bancaria",
                "Tipo Documento",
                "Número",
                "Monto",
                "Fecha documento",
            ]
            summary_ws.append(headers)

            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            for col_idx in range(1, len(headers) + 1):
                cell = summary_ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment

            account_number = self._extract_account_number(metadata.get('account', ''))

            for row_data in data_rows:
                debit_value = row_data.get('Débitos')
                credit_value = row_data.get('Créditos')
                debit_number = self._to_number(debit_value)
                credit_number = self._to_number(credit_value)

                amount: Optional[float] = None
                if credit_number > 0:
                    amount = credit_number
                elif debit_number > 0:
                    amount = debit_number
                elif self._has_numeric_data(credit_value) and abs(credit_number) > 1e-9:
                    amount = credit_number
                elif self._has_numeric_data(debit_value) and abs(debit_number) > 1e-9:
                    amount = debit_number
                elif self._has_numeric_data(credit_value):
                    amount = credit_number
                elif self._has_numeric_data(debit_value):
                    amount = debit_number

                if amount is None:
                    continue

                date_value = row_data.get('Fecha')
                parsed_date = self._parse_date_from_value(date_value)

                summary_row = [
                    account_number,
                    str(row_data.get('Código', '') or '').strip(),
                    row_data.get('Documento', '') or '',
                    amount,
                    parsed_date if parsed_date else date_value,
                ]

                summary_ws.append(summary_row)

            if summary_ws.max_row == 1:
                logger.log(
                    "No se encontraron movimientos con montos válidos para el resumen contable del Caso 7.",
                    level="WARNING",
                )
                return None

            amount_column = headers.index('Monto') + 1
            date_column = headers.index('Fecha documento') + 1

            for row in range(2, summary_ws.max_row + 1):
                amount_cell = summary_ws.cell(row=row, column=amount_column)
                if isinstance(amount_cell.value, (int, float)):
                    amount_cell.number_format = '#,##0.00'
                    amount_cell.alignment = Alignment(horizontal='right', vertical='center')

                date_cell = summary_ws.cell(row=row, column=date_column)
                if isinstance(date_cell.value, datetime):
                    date_cell.number_format = 'dd/mm/yyyy'
                    date_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col_idx in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in summary_ws[column_letter]:
                    if cell.value in (None, ''):
                        continue
                    value = cell.value
                    if isinstance(value, datetime):
                        text = value.strftime('%d/%m/%Y')
                    else:
                        text = str(value)
                    if len(text) > max_length:
                        max_length = len(text)
                summary_ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

            output = io.BytesIO()
            summary_wb.save(output)
            output.seek(0)
            return output.read()
        except Exception as exc:
            logger.log(
                f"Error inesperado al generar el resumen contable del Caso 7: {exc}",
                level="ERROR",
            )
            return None

    def _extract_account_number(self, account_text: Any) -> str:
        if not account_text:
            return ''
        digits = re.findall(r'\d+', str(account_text))
        if not digits:
            return ''
        return max(digits, key=len)

    def _has_numeric_data(self, value: Any) -> bool:
        if value in (None, ''):
            return False
        if isinstance(value, (int, float)):
            return abs(float(value)) > 1e-9
        if isinstance(value, str):
            return bool(re.search(r'\d', value))
        return False
