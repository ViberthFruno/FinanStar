# Archivo: case10.py
# Ubicación: raíz del proyecto
# Descripción: Caso 10 - Rediseña estados de cuenta bancarios con formato celeste profesional

import io
import os
import re
import unicodedata
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple
from zipfile import BadZipFile

from config_manager import ConfigManager


class Case:
    """Caso 10 - Rediseña estados de cuenta bancarios con encabezado celeste."""

    HEADERS: Tuple[str, ...] = (
        "Fecha Contable",
        "Fecha de Registro",
        "Hora de Registro",
        "Número Documento",
        "Descripción",
        "Oficina",
        "Débitos",
        "Créditos",
    )
    OPTIONAL_HEADERS: Tuple[str, ...] = ("Código",)
    OUTPUT_HEADERS: Tuple[str, ...] = (
        "Fecha Contable",
        "Código",
        "Fecha de Registro",
        "Hora de Registro",
        "Número Documento",
        "Descripción",
        "Oficina",
        "Débitos",
        "Créditos",
        "Revisar",
    )

    def __init__(self) -> None:
        self.name = "Caso 10"
        self.description = (
            "Recibe archivos Excel con movimientos bancarios y los reorganiza en un nuevo formato "
            "con encabezado celeste y estructura mejorada."
        )
        self.response_message = (
            "Hola,\n\nSe adjunta el archivo con el formato celeste actualizado del estado de cuenta. "
            "Quedo atento a cualquier comentario.\n\nSaludos cordiales."
        )
        self.corrupted_file_message = (
            "Hola,\n\n"
            "Lamentablemente, el archivo Excel que nos envió está corrupto y no se puede procesar correctamente. "
            "Este problema ocurre ocasionalmente con archivos descargados directamente del sistema bancario.\n\n"
            "Para solucionarlo, por favor siga estos pasos:\n\n"
            "1. Abra el archivo Excel en su computadora\n"
            "2. Una vez abierto, vaya a 'Archivo' > 'Guardar como'\n"
            "3. Guarde el archivo con un nuevo nombre\n"
            "4. Envíenos este nuevo archivo guardado\n\n"
            "Este proceso reparará la estructura del archivo y permitirá que podamos procesarlo sin problemas.\n\n"
            "Quedamos atentos al reenvío del archivo.\n\n"
            "Saludos cordiales."
        )
        self.config_manager = ConfigManager()
        self.config_case_key = 'case10'

    # ==================== MÉTODOS PÚBLICOS ====================

    def get_name(self) -> str:
        return self.name

    def get_description(self) -> str:
        return self.description

    def get_search_keywords(self) -> List[str]:
        try:
            config = self.config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso10', '').strip()
            if keyword:
                return [keyword]
            return []
        except Exception as exc:
            print(f"Error al cargar palabras clave para caso10: {exc}")
            return []

    def process_email(self, email_data: Dict[str, Any], logger) -> Optional[Dict[str, Any]]:
        try:
            sender = email_data.get('sender', '')
            subject = email_data.get('subject', '')
            attachments = email_data.get('attachments', [])

            logger.log(
                f"Procesando {self.name} para email de {sender} con asunto: {subject}",
                level="INFO",
            )

            date_range = self._extract_date_range(subject)
            if date_range:
                start, end = date_range
                logger.log(
                    "Se aplicará un filtrado de fechas desde "
                    f"{start.strftime('%d/%m/%Y')} hasta {end.strftime('%d/%m/%Y')}",
                    level="INFO",
                )
            else:
                logger.log(
                    "No se encontró un rango de fechas válido en el asunto. "
                    "Se conservarán todos los movimientos.",
                    level="INFO",
                )

            excel_attachments = [
                attachment
                for attachment in attachments
                if self._is_excel_file(attachment.get('filename'))
            ]

            if not excel_attachments:
                logger.log(
                    "No se encontraron adjuntos de Excel para procesar en el correo recibido.",
                    level="WARNING",
                )
                return None

            processed_files: List[Dict[str, Any]] = []
            corrupted_files: List[str] = []

            for attachment in excel_attachments:
                result = self._redesign_excel_attachment(
                    attachment,
                    logger,
                    date_range=date_range,
                )

                if result == 'CORRUPTED':
                    filename = attachment.get('filename', 'archivo desconocido')
                    corrupted_files.append(filename)
                    logger.log(
                        f"Archivo corrupto detectado: {filename}",
                        level="WARNING",
                    )
                elif result:
                    processed_files.extend(result)

            # Si hay archivos corruptos, enviar mensaje especial
            if corrupted_files:
                logger.log(
                    f"Se detectaron {len(corrupted_files)} archivo(s) corrupto(s). "
                    f"Enviando respuesta con instrucciones de solución.",
                    level="INFO",
                )

                response_data = {
                    'recipient': sender,
                    'subject': f"Re: {subject}",
                    'body': self.corrupted_file_message,
                    'attachments': [],
                }

                return response_data

            # Si no hay archivos procesados exitosamente
            if not processed_files:
                logger.log(
                    "No fue posible generar el archivo rediseñado para los adjuntos proporcionados.",
                    level="ERROR",
                )
                return None

            # Respuesta exitosa con archivos procesados
            response_data = {
                'recipient': sender,
                'subject': f"Re: {subject}",
                'body': self.response_message,
                'attachments': processed_files,
            }

            logger.log(
                f"Respuesta generada para {self.name} con {len(processed_files)} adjunto(s).",
                level="INFO",
            )

            return response_data

        except Exception as exc:
            logger.log(f"Error al procesar email en {self.name}: {exc}", level="ERROR")
            return None

    def get_response_message(self) -> str:
        return self.response_message

    # ==================== MÉTODOS INTERNOS ====================

    def _is_excel_file(self, filename: Optional[str]) -> bool:
        if not filename:
            return False
        extension = os.path.splitext(filename)[1].lower()
        return extension in {'.xls', '.xlsx'}

    def _extract_date_range(self, subject: str) -> Optional[Tuple[datetime, datetime]]:
        """Busca un rango de fechas (dd/mm/yyyy) dentro del asunto del correo."""
        if not subject:
            return None

        matches = re.findall(r"(\d{1,2}/\d{1,2}/\d{4})", subject)
        if len(matches) < 2:
            return None

        try:
            start = datetime.strptime(matches[0], "%d/%m/%Y")
            end = datetime.strptime(matches[1], "%d/%m/%Y")
            return start, end
        except ValueError:
            return None

    def _redesign_excel_attachment(
            self,
            attachment: Dict[str, Any],
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[List[Dict[str, Any]]]:
        filename = attachment.get('filename') or 'reporte.xlsx'
        content = attachment.get('content')

        if not content:
            logger.log(f"El adjunto '{filename}' está vacío o no pudo leerse.", level="WARNING")
            return None

        try:
            workbook_bytes = self._create_redesigned_workbook(
                content,
                filename,
                logger,
                date_range=date_range,
            )

            if workbook_bytes == 'CORRUPTED':
                return 'CORRUPTED'

            if not workbook_bytes:
                return None

            primary_bytes, summary_bytes = workbook_bytes
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            attachments: List[Dict[str, Any]] = []

            output_name = self._build_output_filename(filename, timestamp=timestamp)
            attachments.append(
                {
                    'filename': output_name,
                    'content': primary_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                }
            )

            if summary_bytes:
                summary_name = self._build_output_filename(
                    filename,
                    suffix='caso10_resumen',
                    timestamp=timestamp,
                )
                attachments.append(
                    {
                        'filename': summary_name,
                        'content': summary_bytes,
                        'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    }
                )

            return attachments

        except Exception as exc:
            logger.log(
                f"Error inesperado al rediseñar el archivo '{filename}': {exc}",
                level="ERROR",
            )
            return None

    def _create_redesigned_workbook(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[Tuple[bytes, Optional[bytes]]]:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        import warnings

        try:
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                source_wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        except BadZipFile:
            logger.log(
                f"El archivo '{original_name}' está corrupto (no es un archivo ZIP válido). "
                f"Se enviará una respuesta con instrucciones para solucionar el problema.",
                level="WARNING",
            )
            return 'CORRUPTED'
        except Exception as exc:
            error_message = str(exc).lower()
            if 'zip' in error_message or 'corrupt' in error_message or 'not a zip file' in error_message:
                logger.log(
                    f"El archivo '{original_name}' parece estar corrupto: {exc}. "
                    f"Se enviará una respuesta con instrucciones para solucionar el problema.",
                    level="WARNING",
                )
                return 'CORRUPTED'

            logger.log(
                f"No fue posible abrir el archivo '{original_name}' para rediseño: {exc}",
                level="ERROR",
            )
            return None

        source_ws = source_wb.active

        # Extraer metadata del archivo original
        metadata = self._extract_metadata(source_ws, logger)

        # Extraer filas de datos
        data_rows = self._extract_data_rows(source_ws, logger)

        if not data_rows:
            logger.log(
                f"El archivo '{original_name}' no contiene datos de movimientos para rediseñar.",
                level="WARNING",
            )
            return None

        if date_range:
            data_rows = self._filter_data_rows_by_date_range(data_rows, date_range, logger)

            if not data_rows:
                start, end = date_range
                logger.log(
                    "No se encontraron movimientos dentro del rango de fechas "
                    f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}.",
                    level="WARNING",
                )
                return None

        # Asignar códigos basados en descripción (nueva lógica)
        if data_rows:
            self._assign_codes_by_description(data_rows, logger)

        # Crear nuevo workbook con diseño mejorado
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos Bancarios"

        # Encabezados principales (metadata)
        ws.cell(row=2, column=1, value="Cliente:")
        ws.cell(row=2, column=2, value=metadata.get('cliente', ''))

        ws.cell(row=4, column=1, value="Cuenta IBAN:")
        ws.cell(row=4, column=2, value=metadata.get('cuenta_iban', ''))

        ws.cell(row=4, column=3, value="Tipo de Movimiento:")
        ws.cell(row=4, column=4, value=metadata.get('tipo_movimiento', ''))

        ws.cell(row=5, column=1, value="Fecha Desde:")
        ws.cell(row=5, column=2, value=metadata.get('fecha_desde', ''))

        ws.cell(row=5, column=3, value="Fecha Hasta:")
        ws.cell(row=5, column=4, value=metadata.get('fecha_hasta', ''))

        # Fila de encabezados de datos (fila 7)
        header_row = 7
        output_headers = list(self.OUTPUT_HEADERS)
        column_map = {header: idx for idx, header in enumerate(output_headers, start=1)}

        for header, col_idx in column_map.items():
            ws.cell(row=header_row, column=col_idx, value=header)

        # Datos
        data_start = header_row + 1
        for row_offset, row_data in enumerate(data_rows):
            current_row = data_start + row_offset
            row_data.setdefault('Código', '')
            row_data.setdefault('Revisar', '')

            for header in output_headers:
                col_idx = column_map[header]
                cell = ws.cell(row=current_row, column=col_idx)
                value = row_data.get(header)

                if header in {'Fecha Contable', 'Fecha de Registro'}:
                    parsed_date = self._parse_date_from_value(value)
                    cell.value = parsed_date if parsed_date else value
                elif header in {'Débitos', 'Créditos'}:
                    number = self._to_number(value)
                    cell.value = number if value not in (None, '') else None
                else:
                    cell.value = value if value not in (None, '') else ''

        # Aplicar estilos
        self._apply_styles(ws, header_row, data_start, len(data_rows), column_map, logger)

        data_end = data_start + len(data_rows) - 1
        if data_rows:
            self._highlight_rows_by_filters(
                ws,
                column_map,
                data_start,
                data_end,
                len(output_headers),
                logger,
            )

        # Ajustar anchos de columna
        self._adjust_column_widths(ws, len(output_headers))

        # Congelar paneles
        ws.freeze_panes = 'A8'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        summary_rows = data_rows[:-4] if len(data_rows) >= 4 else []
        removed_rows = len(data_rows) - len(summary_rows)

        if removed_rows:
            logger.log(
                "Se omitieron las últimas {0} fila(s) al generar el archivo resumen del Caso 10.".format(
                    removed_rows
                ),
                level="INFO",
            )

        summary_bytes = self._create_summary_workbook(summary_rows, metadata)

        return output.read(), summary_bytes

    def _extract_metadata(self, worksheet, logger) -> Dict[str, Any]:
        """Extrae la metadata del archivo original según las posiciones especificadas."""
        metadata = {
            'cliente': '',
            'cuenta_iban': '',
            'tipo_movimiento': '',
            'fecha_desde': '',
            'fecha_hasta': '',
        }

        try:
            # Cliente: Fila 3, Columna D
            cliente_value = worksheet.cell(row=3, column=4).value
            metadata['cliente'] = str(cliente_value).strip() if cliente_value else ''
            logger.log(f"Cliente extraído: {metadata['cliente']}", level="INFO")

            # Cuenta IBAN: Fila 6, Columna A
            iban_value = worksheet.cell(row=6, column=1).value
            metadata['cuenta_iban'] = str(iban_value).strip() if iban_value else ''
            logger.log(f"Cuenta IBAN extraída: {metadata['cuenta_iban']}", level="INFO")

            # Tipo de Movimiento: Fila 6, Columna B
            tipo_value = worksheet.cell(row=6, column=2).value
            metadata['tipo_movimiento'] = str(tipo_value).strip() if tipo_value else ''
            logger.log(f"Tipo de Movimiento extraído: {metadata['tipo_movimiento']}", level="INFO")

            # Fecha Desde: Fila 6, Columna C
            fecha_desde_value = worksheet.cell(row=6, column=3).value
            if isinstance(fecha_desde_value, datetime):
                metadata['fecha_desde'] = fecha_desde_value.strftime('%d/%m/%Y')
            else:
                metadata['fecha_desde'] = str(fecha_desde_value).strip() if fecha_desde_value else ''
            logger.log(f"Fecha Desde extraída: {metadata['fecha_desde']}", level="INFO")

            # Fecha Hasta: Fila 6, Columna D
            fecha_hasta_value = worksheet.cell(row=6, column=4).value
            if isinstance(fecha_hasta_value, datetime):
                metadata['fecha_hasta'] = fecha_hasta_value.strftime('%d/%m/%Y')
            else:
                metadata['fecha_hasta'] = str(fecha_hasta_value).strip() if fecha_hasta_value else ''
            logger.log(f"Fecha Hasta extraída: {metadata['fecha_hasta']}", level="INFO")

        except Exception as exc:
            logger.log(f"Error al extraer metadata: {exc}", level="WARNING")

        return metadata

    def _extract_data_rows(self, worksheet, logger) -> List[Dict[str, Any]]:
        """Extrae las filas de datos desde la fila 9 en adelante."""
        data_rows: List[Dict[str, Any]] = []

        # Los datos comienzan en la fila 9
        data_start = 9
        empty_streak = 0

        row_idx = data_start
        while row_idx <= worksheet.max_row and empty_streak < 5:
            row_data: Dict[str, Any] = {}
            is_empty = True

            # Extraer valores según las columnas esperadas (A-H)
            for col_idx, header in enumerate(self.HEADERS, start=1):
                value = worksheet.cell(row=row_idx, column=col_idx).value

                if value not in (None, ''):
                    is_empty = False

                row_data[header] = value

            optional_col_idx = len(self.HEADERS) + 1
            codigo_value = worksheet.cell(row=row_idx, column=optional_col_idx).value
            if codigo_value not in (None, ''):
                is_empty = False
            row_data['Código'] = codigo_value

            if all(
                isinstance(row_data.get(header), str)
                and self._normalize_text(row_data[header]) == self._normalize_text(header)
                for header in self.HEADERS
            ):
                row_idx += 1
                empty_streak = 0
                continue

            if is_empty:
                empty_streak += 1
            else:
                empty_streak = 0
                data_rows.append(row_data)

            row_idx += 1

        logger.log(f"Se extrajeron {len(data_rows)} filas de datos del archivo.", level="INFO")
        return data_rows

    def _create_summary_workbook(
            self,
            data_rows: List[Dict[str, Any]],
            metadata: Dict[str, Any],
    ) -> Optional[bytes]:
        """Genera un archivo auxiliar con la información requerida para el Caso 10."""
        if not data_rows:
            return None

        from openpyxl import Workbook

        account_value = str(metadata.get('cuenta_iban', '') or '')
        account_number = re.sub(r'\D', '', account_value)
        if not account_number:
            account_number = account_value

        headers = (
            'Cuenta Bancaria',
            'Tipo Documento',
            'Número',
            'Monto',
            'Fecha documento',
        )

        summary_wb = Workbook()
        summary_ws = summary_wb.active
        summary_ws.title = 'Resumen'
        summary_ws.append(headers)

        for row_data in data_rows:
            tipo_documento = row_data.get('Código', '')
            numero_documento = row_data.get('Número Documento', '')
            debito = self._to_number(row_data.get('Débitos'))
            credito = self._to_number(row_data.get('Créditos'))

            if debito > 0 and credito > 0:
                monto = max(debito, credito)
            elif debito > 0:
                monto = debito
            elif credito > 0:
                monto = credito
            else:
                monto = 0

            fecha_valor = row_data.get('Fecha Contable')
            fecha_documento = self._parse_date_from_value(fecha_valor)
            if fecha_documento:
                fecha_resultado = fecha_documento.strftime('%d/%m/%Y')
            elif fecha_valor not in (None, ''):
                if isinstance(fecha_valor, datetime):
                    fecha_resultado = fecha_valor.strftime('%d/%m/%Y')
                else:
                    fecha_resultado = str(fecha_valor)
            else:
                fecha_resultado = ''

            summary_ws.append(
                [
                    account_number,
                    tipo_documento or '',
                    numero_documento or '',
                    monto,
                    fecha_resultado,
                ]
            )

        output = io.BytesIO()
        summary_wb.save(output)
        output.seek(0)
        return output.read()

    def _filter_data_rows_by_date_range(
            self,
            data_rows: List[Dict[str, Any]],
            date_range: Tuple[datetime, datetime],
            logger,
    ) -> List[Dict[str, Any]]:
        """Filtra las filas según el rango de fechas indicado."""
        if not data_rows:
            return []

        start, end = date_range
        filtered_rows: List[Dict[str, Any]] = []
        rows_filtered_out = 0

        for row in data_rows:
            date_value = row.get('Fecha Contable') or row.get('Fecha de Registro')
            parsed_date = self._parse_date_from_value(date_value)

            if parsed_date is None:
                filtered_rows.append(row)
                continue

            if start.date() <= parsed_date.date() <= end.date():
                filtered_rows.append(row)
            else:
                rows_filtered_out += 1

        if rows_filtered_out:
            logger.log(
                "Se filtraron {0} fila(s) fuera del rango de fechas {1} - {2}.".format(
                    rows_filtered_out,
                    start.strftime('%d/%m/%Y'),
                    end.strftime('%d/%m/%Y'),
                ),
                level="INFO",
            )

        return filtered_rows

    def _assign_codes_by_description(
            self,
            data_rows: List[Dict[str, Any]],
            logger,
    ) -> None:
        """Asigna códigos a las filas basándose en reglas de descripción (nueva lógica)."""
        if not data_rows:
            return

        codification_rules = self._get_codification_rules()
        assigned_count = 0

        for row_data in data_rows:
            code = self._determine_codification(row_data, codification_rules)
            if code:
                row_data['Código'] = code
                assigned_count += 1
            else:
                row_data['Código'] = ''

        if assigned_count:
            logger.log(
                f"Se asignaron códigos automáticamente a {assigned_count} fila(s) según las reglas configuradas.",
                level="INFO",
            )

    def _get_codification_rules(self) -> Dict[str, List[Tuple[str, str]]]:
        """Obtiene y prepara las reglas de codificación configuradas para el caso."""
        raw_rules = self.config_manager.get_case10_codification_rules()
        prepared: Dict[str, List[Tuple[str, str]]] = {'debit': [], 'credit': []}

        for key in ('debit', 'credit'):
            entries = raw_rules.get(key, [])
            if not isinstance(entries, list):
                continue
            for item in entries:
                if not isinstance(item, dict):
                    continue
                search_text = item.get('search_text', '')
                code = item.get('code', '')
                if not isinstance(search_text, str) or not isinstance(code, str):
                    continue
                normalized_search = self._normalize_text(search_text)
                if normalized_search and code.strip():
                    prepared[key].append((normalized_search, code.strip()))

        return prepared

    def _determine_codification(
            self,
            row_data: Dict[str, Any],
            codification_rules: Dict[str, List[Tuple[str, str]]],
    ) -> str:
        """Determina el código a asignar a la fila según los filtros configurados."""
        description = row_data.get("Descripción")
        if not isinstance(description, str):
            return ''

        normalized_description = self._normalize_text(description)
        if not normalized_description:
            return ''

        credit_amount = self._to_number(row_data.get("Créditos"))
        debit_amount = self._to_number(row_data.get("Débitos"))

        # Primero verificar créditos
        if credit_amount > 0:
            code = self._match_codification(normalized_description, codification_rules.get('credit', []))
            if code:
                return code

        # Luego verificar débitos
        if debit_amount > 0:
            code = self._match_codification(normalized_description, codification_rules.get('debit', []))
            if code:
                return code

        return ''

    def _match_codification(
            self,
            normalized_description: str,
            rules: List[Tuple[str, str]],
    ) -> str:
        """Devuelve el código correspondiente si alguna regla coincide con la descripción."""
        for search_text, code in rules:
            if search_text and code and search_text in normalized_description:
                return code
        return ''

    def _apply_styles(
            self,
            worksheet,
            header_row: int,
            data_start: int,
            num_data_rows: int,
            column_map: Dict[str, int],
            logger,
    ):
        """Aplica estilos celestes al workbook rediseñado."""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        # Fuentes
        title_font = Font(bold=True, size=12)
        label_font = Font(bold=True, size=11)
        header_font = Font(bold=True, color='FFFFFF', size=11)
        regular_font = Font(size=10)

        # Colores celeste
        header_fill = PatternFill(fill_type='solid', fgColor='4BACC6')  # Celeste para encabezados

        # Bordes
        thin_border = Border(
            left=Side(border_style='thin', color='B0B0B0'),
            right=Side(border_style='thin', color='B0B0B0'),
            top=Side(border_style='thin', color='B0B0B0'),
            bottom=Side(border_style='thin', color='B0B0B0'),
        )

        # Estilos para metadata (filas 2-5)
        for row in [2, 4, 5]:
            for col in [1, 3]:  # Columnas de etiquetas
                cell = worksheet.cell(row=row, column=col)
                cell.font = label_font
                cell.alignment = Alignment(horizontal='left', vertical='center')

            for col in [2, 4]:  # Columnas de valores
                cell = worksheet.cell(row=row, column=col)
                cell.font = regular_font
                cell.alignment = Alignment(horizontal='left', vertical='center')

        total_columns = len(column_map)

        # Estilos para encabezados de datos (fila 7)
        for col_idx in range(1, total_columns + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Estilos para datos
        data_end = data_start + num_data_rows - 1

        numeric_columns = {
            column_map.get('Débitos'),
            column_map.get('Créditos'),
        }

        date_columns = {
            column_map.get('Fecha Contable'),
            column_map.get('Fecha de Registro'),
        }

        review_column = column_map.get('Revisar')

        for row_idx in range(data_start, data_end + 1):
            for col_idx in range(1, total_columns + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = regular_font

                if col_idx in numeric_columns:
                    # Formato numérico con separadores de miles y decimales
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx in date_columns:
                    # Formato de fecha
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == column_map.get('Hora de Registro'):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif review_column and col_idx == review_column:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        logger.log("Estilos celestes aplicados correctamente al archivo.", level="INFO")

    def _highlight_rows_by_filters(
            self,
            worksheet,
            column_map: Dict[str, int],
            start_row: int,
            end_row: int,
            total_columns: int,
            logger,
    ) -> None:
        """Resalta filas cuya descripción coincida con los filtros configurados."""
        filters = self.config_manager.get_case10_filters()
        if not filters:
            return

        normalized_filters = [
            self._normalize_text(filter_text)
            for filter_text in filters
            if self._normalize_text(filter_text)
        ]

        if not normalized_filters:
            return

        description_column = column_map.get('Descripción')
        review_column = column_map.get('Revisar')

        if not description_column:
            logger.log(
                "No se encontró la columna de descripción para aplicar los filtros del Caso 10.",
                level="WARNING",
            )
            return

        from openpyxl.styles import Alignment, PatternFill

        highlight_fill = PatternFill(fill_type='solid', fgColor='FFF3B0')
        highlighted_rows = 0

        for row_idx in range(start_row, end_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=description_column).value
            if cell_value in (None, ''):
                continue

            normalized_value = self._normalize_text(str(cell_value))
            if not normalized_value:
                continue

            if any(filter_text in normalized_value for filter_text in normalized_filters):
                for col_idx in range(1, total_columns + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = highlight_fill

                if review_column:
                    review_cell = worksheet.cell(row=row_idx, column=review_column)
                    review_cell.value = 'Revisar'
                    review_cell.alignment = Alignment(horizontal='center', vertical='center')

                highlighted_rows += 1

        if highlighted_rows:
            logger.log(
                f"Se resaltaron {highlighted_rows} fila(s) según los filtros configurados para el Caso 10.",
                level="INFO",
            )

    def _adjust_column_widths(self, worksheet, num_columns: int):
        """Ajusta el ancho de las columnas automáticamente."""
        from openpyxl.utils import get_column_letter

        for col_idx in range(1, num_columns + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for cell in worksheet[column_letter]:
                if cell.value is None:
                    continue

                cell_value = cell.value
                if isinstance(cell_value, (int, float)):
                    text = f"{cell_value:,.2f}"
                elif isinstance(cell_value, datetime):
                    text = cell_value.strftime('%d/%m/%Y')
                else:
                    text = str(cell_value)

                if len(text) > max_length:
                    max_length = len(text)

            # Establecer ancho con un máximo de 50
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)

    def _parse_date_from_value(self, value: Any) -> Optional[datetime]:
        if isinstance(value, datetime):
            return value
        if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
            try:
                return datetime(value.year, value.month, value.day)
            except Exception:
                return None
        if isinstance(value, (int, float)) and value > 0:
            try:
                base_date = datetime(1899, 12, 30)
                converted = base_date + timedelta(days=float(value))
                if 1900 <= converted.year <= 9999:
                    return converted
            except Exception:
                return None
        if isinstance(value, str):
            return self._parse_date_string(value)
        return None

    def _parse_date_string(self, value: str) -> Optional[datetime]:
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.replace('.', '/').replace('-', '/').replace('–', '/')
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(cleaned, fmt)
            except ValueError:
                continue
        return None

    def _to_number(self, value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return 0.0
            cleaned = cleaned.replace(' ', '')
            if ',' in cleaned and '.' in cleaned:
                if cleaned.rfind(',') > cleaned.rfind('.'):
                    cleaned = cleaned.replace('.', '')
                    cleaned = cleaned.replace(',', '.')
                else:
                    cleaned = cleaned.replace(',', '')
            elif ',' in cleaned:
                cleaned = cleaned.replace(',', '.')
            else:
                cleaned = cleaned.replace(',', '')
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return 0.0

    def _normalize_text(self, text: Any) -> str:
        if not isinstance(text, str):
            return ''
        normalized = unicodedata.normalize('NFKD', text)
        without_accents = ''.join(c for c in normalized if not unicodedata.combining(c))
        return without_accents.lower().replace(' ', '')

    def _build_output_filename(
            self,
            original_name: str,
            suffix: str = 'caso10',
            timestamp: Optional[str] = None,
    ) -> str:
        base, _ = os.path.splitext(original_name)
        if not timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base}_{suffix}_{timestamp}.xlsx"