# Archivo: case9.py
# Ubicación: raíz del proyecto
# Descripción: Caso 9 - Genera plantillas CP/CB desde archivos del Caso 7

from __future__ import annotations

import io
import os
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple
from zipfile import BadZipFile

from config_manager import ConfigManager


class MissingRequiredRowsError(Exception):
    """Excepción lanzada cuando no se encuentran filas CP/CB requeridas en el archivo."""


class InvalidFileFormatError(Exception):
    """Excepción lanzada cuando el archivo no tiene el formato esperado del Caso 7."""


class Case:
    """Caso 9 - Procesa archivos del Caso 7 para generar plantillas CP y CB."""

    OUTPUT_HEADERS_CP = [
        "Proveedor",
        "Número",
        "Tipo Documento",
        "Fecha Documento",
        "Fecha Rige",
        "Aplicacion",
        "Monto",
        "Subtotal",
        "Descuento",
        "Impuesto1",
        "Impuesto2",
        "Rubro1",
        "Rubro2",
        "Condición De Pago",
        "Moneda",
        "Cuenta Bancaria",
        "Subtipo Documento",
        "Fecha Vence",
        "Codigo_impuesto",
        "Tipo Asiento",
        "Paquete",
        "Actividad Comercial",
    ]

    OUTPUT_HEADERS_CB = [
        "Cuenta Bancaria",
        "tipo Documento",
        "Numero",
        "Subtipo Documento",
        "Fecha",
        "Fecha Contable",
        "Concepto",
        "Monto",
        "Confirmado/entregado",
        "tipo Asiento",
        "Paquete",
        "Cod_impuesto",
    ]

    REQUIRED_HEADERS = {
        "fecha",
        "documento",
        "descripcion",
        "debitos",
        "creditos",
        "saldo",
        "codigo",
        "revisar",
    }

    def __init__(self) -> None:
        self.name = "Caso 9"
        self.description = (
            "Recibe archivos Excel generados por el Caso 7, identifica las filas marcadas como "
            "CP o CB en las columnas Código/Revisar y transforma la información en plantillas "
            "estándar por cuenta bancaria."
        )
        self.response_message = (
            "Hola,\n\nSe adjunta(n) el/los archivo(s) con la información transformada según la "
            "plantilla solicitada. Quedo atento a cualquier comentario.\n\nSaludos cordiales."
        )
        self.missing_rows_message = (
            "Hola,\n\nNo se pudieron procesar los archivos adjuntos porque no contienen las filas "
            "necesarias marcadas con 'CP' o 'CB' en las columnas 'Código' o 'Revisar'.\n\n"
            "Por favor, envía un archivo generado previamente por el Caso 7 que contenga las filas "
            "requeridas con las marcas 'CP' y/o 'CB'.\n\nSaludos cordiales."
        )
        self.config_manager = ConfigManager()

    # ==================== MÉTODOS PÚBLICOS ====================

    def get_name(self) -> str:
        return self.name

    def get_description(self) -> str:
        return self.description

    def get_search_keywords(self) -> List[str]:
        try:
            config = self.config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso9', '').strip()
            if keyword:
                return [keyword]
            return []
        except Exception as exc:
            print(f"Error al cargar palabras clave para caso9: {exc}")
            return []

    def process_email(self, email_data: Dict[str, Any], logger) -> Optional[Dict[str, Any]]:
        try:
            sender = email_data.get('sender', '')
            subject = email_data.get('subject', '')
            attachments = email_data.get('attachments', [])

            logger.log(
                f"Procesando {self.name} para email de {sender} con asunto: {subject}",
                level="INFO",
            )

            excel_attachments = [
                attachment
                for attachment in attachments
                if self._is_excel_file(attachment.get('filename'))
            ]

            if not excel_attachments:
                logger.log(
                    "No se encontraron adjuntos de Excel para procesar en el correo recibido.",
                    level="WARNING",
                )
                return None

            processed_files: List[Dict[str, Any]] = []
            files_without_rows = 0

            for attachment in excel_attachments:
                try:
                    files = self._process_attachment(attachment, logger)
                    if files:
                        processed_files.extend(files)
                except MissingRequiredRowsError:
                    logger.log(
                        f"El archivo '{attachment.get('filename')}' no contiene filas CP/CB necesarias.",
                        level="WARNING",
                    )
                    files_without_rows += 1
                    continue
                except InvalidFileFormatError as exc:
                    logger.log(str(exc), level="ERROR")
                    continue

            if files_without_rows > 0 and not processed_files:
                logger.log(
                    "Los archivos no contienen filas marcadas como CP o CB. Se enviará respuesta informativa.",
                    level="ERROR",
                )
                return self._build_missing_rows_response(sender, subject)

            if not processed_files:
                logger.log(
                    "No fue posible generar los archivos de plantilla requeridos para los adjuntos.",
                    level="ERROR",
                )
                return None

            response_data = {
                'recipient': sender,
                'subject': f"Re: {subject}",
                'body': self.response_message,
                'attachments': processed_files,
            }

            logger.log(
                f"Respuesta generada para {self.name} con {len(processed_files)} adjunto(s).",
                level="INFO",
            )

            return response_data
        except Exception as exc:
            logger.log(f"Error al procesar email en {self.name}: {exc}", level="ERROR")
            return None

    def get_response_message(self) -> str:
        return self.response_message

    # ==================== MÉTODOS INTERNOS ====================

    def _process_attachment(self, attachment: Dict[str, Any], logger) -> List[Dict[str, Any]]:
        filename = attachment.get('filename') or 'reporte.xlsx'
        content = attachment.get('content')

        if not content:
            logger.log(
                f"El adjunto '{filename}' está vacío o no pudo leerse.",
                level="WARNING",
            )
            return []

        try:
            from openpyxl import load_workbook
            import warnings

            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                workbook = load_workbook(io.BytesIO(content), data_only=True)

            sheet = workbook.active
        except BadZipFile:
            logger.log(
                f"El archivo '{filename}' está corrupto o no tiene el formato XLSX válido.",
                level="ERROR",
            )
            raise InvalidFileFormatError(
                "El archivo está dañado. Solicite al remitente reenviar el archivo guardándolo nuevamente desde Excel."
            )
        except Exception as exc:
            logger.log(
                f"No se pudo abrir el archivo '{filename}' como XLSX: {exc}",
                level="ERROR",
            )
            return []

        account_number = self._extract_account_number(sheet, logger)

        if not account_number:
            logger.log(
                f"No se pudo determinar el número de cuenta en el archivo '{filename}'.",
                level="WARNING",
            )
            return []

        account_name = self.config_manager.find_account_by_code(account_number, case_key='case9')

        if not account_name:
            logger.log(
                f"El número de cuenta '{account_number}' del archivo '{filename}' no coincide con ninguna cuenta configurada.",
                level="WARNING",
            )
            return []

        account_config = self.config_manager.get_case_account_config('case9', account_name)

        if account_config is None:
            logger.log(
                f"No se pudo cargar la configuración de la cuenta '{account_name}' para {self.name}.",
                level="ERROR",
            )
            return []

        extraction_result = self._extract_rows(sheet, logger)

        if not extraction_result:
            logger.log(
                f"No se pudieron extraer datos del archivo '{filename}'.",
                level="WARNING",
            )
            return []

        cp_rows = extraction_result['cp_rows']
        cb_rows = extraction_result['cb_rows']
        currency_value = extraction_result['currency']

        if not cp_rows and not cb_rows:
            raise MissingRequiredRowsError(
                f"El archivo '{filename}' no contiene filas CP/CB necesarias."
            )

        result_files: List[Dict[str, Any]] = []

        if cp_rows:
            logger.log(
                f"Se encontraron {len(cp_rows)} fila(s) con 'CP' en el archivo '{filename}'.",
                level="INFO",
            )
            cp_workbook_bytes = self._build_cp_workbook(
                cp_rows,
                currency_value,
                account_number,
                account_config,
                logger,
            )
            cp_output_name = self._build_output_filename(filename, 'CP', account_name)
            result_files.append({
                'filename': cp_output_name,
                'content': cp_workbook_bytes,
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

        if cb_rows:
            logger.log(
                f"Se encontraron {len(cb_rows)} fila(s) con 'CB' en el archivo '{filename}'.",
                level="INFO",
            )
            cb_workbook_bytes = self._build_cb_workbook(
                cb_rows,
                account_number,
                account_config,
                logger,
            )
            cb_output_name = self._build_output_filename(filename, 'CB', account_name)
            result_files.append({
                'filename': cb_output_name,
                'content': cb_workbook_bytes,
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

        return result_files

    def _extract_rows(self, sheet, logger) -> Optional[Dict[str, Any]]:
        header_row, header_map = self._find_header_row(sheet)

        if not header_row or not header_map:
            raise InvalidFileFormatError(
                "No se localizaron encabezados válidos (Fecha, Documento, Código, Revisar) en el archivo."
            )

        date_column = header_map.get('fecha')
        description_column = header_map.get('descripcion')
        document_column = header_map.get('documento')
        debit_column = header_map.get('debitos')
        credit_column = header_map.get('creditos')
        code_column = header_map.get('codigo')
        review_column = header_map.get('revisar')

        if not code_column and not review_column:
            raise InvalidFileFormatError(
                "Las columnas 'Código' o 'Revisar' no están presentes en el archivo del Caso 7."
            )

        cp_rows: List[Dict[str, Any]] = []
        cb_rows: List[Dict[str, Any]] = []

        max_row = sheet.max_row
        empty_streak = 0
        data_start = header_row + 1

        for row_idx in range(data_start, max_row + 1):
            date_value = sheet.cell(row=row_idx, column=date_column).value if date_column else None
            description_value = sheet.cell(row=row_idx, column=description_column).value if description_column else None
            document_value = sheet.cell(row=row_idx, column=document_column).value if document_column else None
            debit_value = sheet.cell(row=row_idx, column=debit_column).value if debit_column else None
            credit_value = sheet.cell(row=row_idx, column=credit_column).value if credit_column else None
            code_value = sheet.cell(row=row_idx, column=code_column).value if code_column else None
            review_value = sheet.cell(row=row_idx, column=review_column).value if review_column else None

            if self._is_empty_row(
                date_value,
                description_value,
                document_value,
                debit_value,
                credit_value,
                code_value,
                review_value,
            ):
                empty_streak += 1
                if empty_streak >= 3:
                    break
                continue

            empty_streak = 0

            row_type = self._determine_row_type(code_value, review_value)

            if row_type not in {'CP', 'CB'}:
                continue

            parsed_date = self._parse_date_value(date_value)
            debit_amount = self._parse_decimal(debit_value)
            credit_amount = self._parse_decimal(credit_value)

            descripcion = str(description_value).strip() if description_value not in (None, '') else ''
            referencia = str(document_value).strip() if document_value not in (None, '') else ''
            codigo = str(code_value).strip() if code_value not in (None, '') else ''

            row_data = {
                'fecha': parsed_date if parsed_date else date_value,
                'descripcion': descripcion,
                'debito': debit_amount if debit_amount is not None else 0,
                'credito': credit_amount if credit_amount is not None else 0,
                'referencia': referencia,
                'codigo': codigo,
            }

            if row_type == 'CP':
                cp_rows.append(row_data)
            else:
                cb_rows.append(row_data)

        currency = self._extract_currency(sheet, logger)

        return {
            'cp_rows': cp_rows,
            'cb_rows': cb_rows,
            'currency': currency,
        }

    def _find_header_row(self, sheet) -> Tuple[Optional[int], Dict[str, int]]:
        max_row = min(sheet.max_row, 40)
        max_col = sheet.max_column
        best_row: Optional[int] = None
        best_matches = 0
        best_map: Dict[str, int] = {}

        target_headers = {
            'fecha', 'documento', 'descripcion', 'debitos', 'creditos', 'saldo', 'codigo', 'revisar'
        }

        for row_idx in range(1, max_row + 1):
            current_map: Dict[str, int] = {}
            matches = 0
            for col_idx in range(1, max_col + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                normalized = self._normalize_text(cell_value)
                if not normalized:
                    continue
                current_map[normalized] = col_idx
                if normalized in target_headers:
                    matches += 1

            if matches > best_matches:
                best_row = row_idx
                best_matches = matches
                best_map = current_map

            if matches >= len(target_headers) - 1:
                break

        if not best_row or best_matches == 0:
            return None, {}

        return best_row, best_map

    def _determine_row_type(self, code_value: Any, review_value: Any) -> str:
        for value in (code_value, review_value):
            if value in (None, ''):
                continue
            text = str(value).strip().upper()
            if text in {'CP', 'CB'}:
                return text
        return ''

    def _is_empty_row(self, *values: Any) -> bool:
        for value in values:
            if value not in (None, ''):
                if isinstance(value, str) and not value.strip():
                    continue
                return False
        return True

    def _extract_account_number(self, sheet, logger) -> str:
        max_row = min(sheet.max_row, 15)
        max_col = min(sheet.max_column, 6)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                value = sheet.cell(row=row, column=col).value
                if not value:
                    continue
                text = str(value)
                digits = re.findall(r'\d{6,}', text)
                if digits:
                    account = max(digits, key=len)
                    logger.log(
                        f"Número de cuenta detectado: {account} (fila {row}, columna {col})",
                        level="INFO",
                    )
                    return account
        return ''

    def _extract_currency(self, sheet, logger) -> str:
        max_row = min(sheet.max_row, 20)
        max_col = min(sheet.max_column, 8)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                value = sheet.cell(row=row, column=col).value
                if not isinstance(value, str):
                    continue

                normalized = value.lower().strip()
                if 'moneda' in normalized:
                    adjacent = sheet.cell(row=row, column=col + 1).value if col + 1 <= max_col else None
                    if isinstance(adjacent, str) and adjacent.strip():
                        currency = adjacent.strip()
                        logger.log(
                            f"Moneda detectada junto a etiqueta: '{currency}'",
                            level="INFO",
                        )
                        return currency

                    match = re.search(r'moneda\s*[:\-]?\s*(\w+)', normalized)
                    if match:
                        currency = match.group(1).upper()
                        logger.log(
                            f"Moneda detectada en la etiqueta: '{currency}'",
                            level="INFO",
                        )
                        return currency
        return ''

    def _build_cp_workbook(
            self,
            cp_rows: List[Dict[str, Any]],
            currency_value: str,
            account_number: str,
            account_config: Dict[str, Any],
            logger,
    ) -> bytes:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        sheet.append(self.OUTPUT_HEADERS_CP)

        for row_data in cp_rows:
            fecha = row_data.get('fecha')
            descripcion = row_data.get('descripcion', '')
            debito = row_data.get('debito', 0)
            credito = row_data.get('credito', 0)
            referencia = row_data.get('referencia', '')

            provider_code = self._find_provider_code(descripcion, account_config, logger)
            monto = self._get_amount(debito, credito)

            row = [''] * len(self.OUTPUT_HEADERS_CP)
            row[0] = provider_code
            row[1] = referencia
            row[2] = 'TEF'
            row[3] = fecha
            row[4] = fecha
            row[5] = descripcion
            row[6] = monto
            row[7] = monto
            row[8] = 0
            row[9] = 0
            row[10] = 0
            row[11] = 0
            row[12] = 0
            row[13] = 0
            row[14] = currency_value
            row[15] = account_number
            row[16] = 0
            row[17] = fecha
            row[19] = 'CP'
            row[20] = 'CP'
            row[21] = 523906

            sheet.append(row)

        for column_index in (4, 5, 18):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for date_cell in column_cells:
                    if isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'dd/mm/yyyy'

        for column_index in (7, 8):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for numeric_cell in column_cells:
                    if isinstance(numeric_cell.value, (int, float)):
                        numeric_cell.number_format = '#,##0.00'

        logger.log(
            f"Se generó el archivo CP con {len(cp_rows)} fila(s).",
            level="INFO",
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _build_cb_workbook(
            self,
            cb_rows: List[Dict[str, Any]],
            account_number: str,
            account_config: Dict[str, Any],
            logger,
    ) -> bytes:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        sheet.append(self.OUTPUT_HEADERS_CB)

        for row_data in cb_rows:
            fecha = row_data.get('fecha')
            descripcion = row_data.get('descripcion', '')
            debito = row_data.get('debito', 0)
            credito = row_data.get('credito', 0)
            referencia = row_data.get('referencia', '')
            codigo = row_data.get('codigo', '')

            monto = self._get_amount(debito, credito)
            subtype_value = self._find_subtype_value(codigo, descripcion, account_config, logger)

            row = [''] * len(self.OUTPUT_HEADERS_CB)
            row[0] = account_number
            row[1] = codigo
            row[2] = referencia
            row[3] = subtype_value
            row[4] = fecha
            row[5] = fecha
            row[6] = descripcion
            row[7] = monto
            row[8] = ''
            row[9] = 'CB'
            row[10] = 'CB'
            row[11] = 'ND'

            sheet.append(row)

        for column_index in (5, 6):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for date_cell in column_cells:
                    if isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'dd/mm/yyyy'

        for column_index in (8,):
            for column_cells in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=sheet.max_row,
            ):
                for numeric_cell in column_cells:
                    if isinstance(numeric_cell.value, (int, float)):
                        numeric_cell.number_format = '#,##0.00'

        logger.log(
            f"Se generó el archivo CB con {len(cb_rows)} fila(s).",
            level="INFO",
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    def _build_output_filename(self, original_name: str, file_type: str, account_name: str) -> str:
        base, _ = os.path.splitext(original_name)
        safe_account_name = re.sub(r'[^\w\s-]', '', account_name).strip().replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base}_{safe_account_name}_{file_type}_{timestamp}.xlsx"

    def _find_provider_code(self, description: str, account_config: Dict[str, Any], logger) -> str:
        if not description:
            return ''

        providers_list = account_config.get('providers', [])

        if not providers_list:
            return ''

        description_lower = description.lower()

        for provider in providers_list:
            search_text = provider.get('search_text', '').strip().lower()
            provider_code = provider.get('provider_code', '').strip()

            if not search_text or not provider_code:
                continue

            if search_text in description_lower:
                logger.log(
                    f"Proveedor detectado '{provider_code}' para descripción que contiene '{search_text}'.",
                    level="INFO",
                )
                return provider_code

        return ''

    def _find_subtype_value(self, document_type: str, description: str, account_config: Dict[str, Any], logger) -> str:
        if not document_type or not description:
            return ''

        subtypes_list = account_config.get('subtypes', [])

        if not subtypes_list:
            return ''

        document_type_upper = document_type.strip().upper()
        description_lower = description.lower()

        for subtype_rule in subtypes_list:
            rule_doc_type = subtype_rule.get('document_type', '').strip().upper()
            rule_search_text = subtype_rule.get('search_text', '').strip().lower()
            subtype_value = subtype_rule.get('subtype_value', '').strip()

            if not rule_doc_type or not rule_search_text or not subtype_value:
                continue

            if rule_doc_type == document_type_upper and rule_search_text in description_lower:
                logger.log(
                    f"Subtipo detectado '{subtype_value}' para tipo '{document_type_upper}'.",
                    level="INFO",
                )
                return subtype_value

        return ''

    def _get_amount(self, debit: float, credit: float) -> float:
        if debit and debit > 0:
            return debit
        if credit and credit > 0:
            return credit
        return 0

    def _parse_date_value(self, value: Any) -> Optional[datetime]:
        if isinstance(value, datetime):
            return value
        if isinstance(value, (int, float)):
            try:
                base_date = datetime(1899, 12, 30)
                converted = base_date + timedelta(days=float(value))
                if 1900 <= converted.year <= 9999:
                    return converted
            except Exception:
                return None
        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y'):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue
        return None

    def _parse_decimal(self, value: Any) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if not text or text in {'-', '--'}:
                return None
            text = text.replace('\xa0', '').replace(' ', '')
            text = re.sub(r'[^0-9,.-]', '', text)
            if not text:
                return None

            if text.count(',') > 1 and '.' not in text:
                last_comma = text.rfind(',')
                text = text[:last_comma].replace(',', '') + '.' + text[last_comma + 1:]
            elif text.count('.') > 1 and ',' not in text:
                last_dot = text.rfind('.')
                text = text[:last_dot].replace('.', '') + '.' + text[last_dot + 1:]
            elif text.count(',') == 1 and '.' not in text:
                text = text.replace(',', '.')
            elif text.count('.') == 1 and text.count(',') == 1:
                if text.rfind('.') < text.rfind(','):
                    text = text.replace('.', '')
                    text = text.replace(',', '.')
                else:
                    text = text.replace(',', '')

            try:
                return float(text)
            except ValueError:
                return None
        return None

    def _normalize_text(self, text: Any) -> str:
        if not isinstance(text, str):
            return ''
        import unicodedata

        normalized = unicodedata.normalize('NFKD', text)
        normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r'[^\w]', '', normalized)
        return normalized.lower()

    def _is_excel_file(self, filename: Optional[str]) -> bool:
        if not filename:
            return False
        extension = os.path.splitext(filename)[1].lower()
        return extension in {'.xls', '.xlsx'}

    def _build_missing_rows_response(self, sender: str, subject: str) -> Dict[str, Any]:
        return {
            'recipient': sender,
            'subject': f"Re: {subject}",
            'body': self.missing_rows_message,
            'attachments': [],
        }

