# Archivo: case5.py
# Ubicación: raíz del proyecto
# Descripción: Caso 5 - Rediseña el formato del estado de cuenta con filtrado por descripción

import io
import os
import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from case4 import Case as BaseCase
from config_manager import ConfigManager


class Case(BaseCase):
    """Caso 5 - Aplica rediseño del estado de cuenta con filtrado por descripción."""

    HEADERS: Tuple[str, ...] = (
        "Fecha",
        "Código",
        "Descripción",
        "Ref.",
        "Créditos (CR)",
        "Revisar",
        "Ref2",
        "Tipo Tran",
        "Causa",
        "Sucursal",
        "D/C",
        "Cuenta",
    )

    def __init__(self):
        super().__init__()
        self.name = "Caso 5"
        self.description = (
            "Recibe archivos Excel del estado de cuenta, genera una versión con el nuevo diseño "
            "y resalta las filas cuya descripción coincida con las palabras clave configuradas."
        )
        self.response_message = (
            "Hola,\n\nSe adjunta el archivo con el formato actualizado del estado de cuenta "
            "correspondiente al Caso 5, resaltando los movimientos configurados para seguimiento. "
            "Quedo atento a cualquier comentario.\n\nSaludos cordiales."
        )
        self.config_manager = ConfigManager()
        self.config_case_key = 'case5'

    def get_search_keywords(self) -> List[str]:
        """Obtiene la palabra clave configurada para el Caso 5."""
        try:
            config = self.config_manager.load_config()
            search_params = config.get('search_params', {})
            keyword = search_params.get('caso5', '').strip()
            if keyword:
                return [keyword]
            return []
        except Exception as exc:
            print(f"Error al cargar palabras clave para caso5: {exc}")
            return []

    def _redesign_excel_attachment(
            self,
            attachment: Dict[str, Any],
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[List[Dict[str, Any]]]:
        """Genera archivo Excel rediseñado sin archivo resumen."""
        filename = attachment.get('filename') or 'reporte.xlsx'
        content = attachment.get('content')

        if not content:
            logger.log(f"El adjunto '{filename}' está vacío o no pudo leerse.", level="WARNING")
            return None

        try:
            workbook_result = self._create_redesigned_workbook(
                content, filename, logger, date_range
            )

            if not workbook_result:
                return None

            workbook_bytes = workbook_result['workbook_bytes']

            output_name = self._build_output_filename(filename)

            attachments_list: List[Dict[str, Any]] = [
                {
                    'filename': output_name,
                    'content': workbook_bytes,
                    'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                }
            ]

            return attachments_list

        except Exception as exc:
            logger.log(
                f"Error inesperado al rediseñar el archivo '{filename}': {exc}",
                level="ERROR",
            )
            return None

    def _create_redesigned_workbook(
            self,
            file_bytes: bytes,
            original_name: str,
            logger,
            date_range: Optional[Tuple[datetime, datetime]] = None,
    ) -> Optional[Dict[str, Any]]:
        """Crea el nuevo archivo Excel con el encabezado, tabla actualizada y filtrado por descripción."""
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image
        import warnings

        try:
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                source_wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        except Exception as exc:
            logger.log(
                f"No fue posible abrir el archivo '{original_name}' para rediseño: {exc}",
                level="ERROR",
            )
            return None

        source_ws = source_wb.active

        info_values = self._extract_info_fields(source_ws)
        header_row, header_map = self._find_header_row(source_ws)

        if not header_row or not header_map:
            logger.log(
                "No se encontraron encabezados válidos en el archivo fuente para aplicar el rediseño.",
                level="ERROR",
            )
            return None

        target_columns = [header_map.get(self._simplify_header(header)) for header in self.HEADERS]

        if not any(target_columns):
            logger.log(
                "No se pudieron mapear las columnas requeridas para el nuevo diseño.",
                level="ERROR",
            )
            return None

        missing_headers = [
            header for header, column_index in zip(self.HEADERS, target_columns)
            if header not in ("Código", "Revisar") and not column_index
        ]

        if missing_headers:
            logger.log(
                "No se encontraron todas las columnas esperadas. Las columnas faltantes son: "
                + ", ".join(missing_headers),
                level="WARNING",
            )

        data_rows: List[Dict[str, Any]] = []
        blank_streak = 0
        row_idx = header_row + 1
        max_row = source_ws.max_row

        while row_idx <= max_row:
            row_data: Dict[str, Any] = {}
            row_has_value = False

            for header, column_index in zip(self.HEADERS, target_columns):
                value = None
                if column_index:
                    value = source_ws.cell(row=row_idx, column=column_index).value
                row_data[header] = value
                if value not in (None, ''):
                    row_has_value = True

            if "Código" not in row_data or row_data["Código"] in (None, ''):
                row_data["Código"] = ""
            row_data["Revisar"] = ""

            if row_has_value:
                data_rows.append(row_data)
                blank_streak = 0
            else:
                blank_streak += 1
                if blank_streak >= 2:
                    break

            row_idx += 1

        if date_range:
            data_rows = self._filter_data_rows_by_date_range(data_rows, date_range, logger)

        columns_to_remove, description_keywords = self._get_removal_configuration()

        if description_keywords:
            data_rows = self._remove_rows_by_description_keywords(
                data_rows,
                description_keywords,
                logger,
            )

        if data_rows:
            self._assign_codes_by_description(data_rows, logger)

        active_headers = self._get_active_headers(columns_to_remove)
        removed_headers = [header for header in self.HEADERS if header not in active_headers]
        if removed_headers:
            logger.log(
                "Se eliminarán las columnas configuradas: " + ", ".join(removed_headers),
                level="INFO",
            )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Detalle"

        self._insert_logo(worksheet, logger)
        self._populate_header_section(worksheet, info_values)
        self._populate_table_with_headers(worksheet, data_rows, active_headers)
        self._apply_styles_with_headers(worksheet, len(data_rows), active_headers)
        self._highlight_rows_by_filters(
            worksheet,
            data_rows,
            active_headers,
            logger,
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return {
            'workbook_bytes': output.read(),
            'data_rows': data_rows
        }

    def _populate_table_with_headers(
            self,
            worksheet,
            data_rows: List[Dict[str, Any]],
            headers: List[str],
    ) -> None:
        header_row = 13
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=header_row, column=col_idx, value=header)

        data_start = header_row + 1
        for row_offset, row_data in enumerate(data_rows):
            for col_idx, header in enumerate(headers, start=1):
                worksheet.cell(
                    row=data_start + row_offset,
                    column=col_idx,
                    value=row_data.get(header),
                )

    def _apply_styles_with_headers(
            self,
            worksheet,
            data_length: int,
            headers: List[str],
    ) -> None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        header_row = 13
        start_row = header_row + 1
        end_row = start_row + max(data_length - 1, 0)

        title_font = Font(bold=True)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type='solid', fgColor='004C97')
        thin_border = Border(
            left=Side(border_style='thin', color='B0B0B0'),
            right=Side(border_style='thin', color='B0B0B0'),
            top=Side(border_style='thin', color='B0B0B0'),
            bottom=Side(border_style='thin', color='B0B0B0'),
        )

        for offset in range(len(self.INFO_FIELDS)):
            cell = worksheet.cell(row=5 + offset, column=1)
            cell.font = title_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

        for col_idx in range(1, len(headers) + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        numeric_headers = {"Débitos (DR)", "Créditos (CR)"}
        date_column = None

        for idx, header in enumerate(headers, start=1):
            if header == "Fecha":
                date_column = idx
                break

        for row_idx in range(start_row, end_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if header in numeric_headers:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif date_column and col_idx == date_column:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        worksheet.freeze_panes = "A14"

        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                if cell.value is None:
                    continue
                text = str(cell.value)
                if len(text) > max_length:
                    max_length = len(text)
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 40)

    def _get_active_headers(self, columns_to_remove: Optional[List[str]] = None) -> List[str]:
        if columns_to_remove is None:
            columns_to_remove, _ = self._get_removal_configuration()

        simplified_to_remove = [
            self._simplify_header(column_name)
            for column_name in columns_to_remove
            if self._simplify_header(column_name)
        ]

        def should_remove(header: str) -> bool:
            simplified_header = self._simplify_header(header)
            if not simplified_header:
                return False
            return any(
                removal == simplified_header
                for removal in simplified_to_remove
            )

        headers = [
            header
            for header in self.HEADERS
            if not should_remove(header)
        ]

        return headers or list(self.HEADERS)

    def _get_removal_configuration(self) -> Tuple[List[str], List[Tuple[str, str]]]:
        configured = self.config_manager.get_case5_columns_to_remove()
        columns_to_remove: List[str] = []
        keywords: List[Tuple[str, str]] = []

        header_variants: Dict[str, str] = {}
        for header in self.HEADERS:
            simplified = self._simplify_header(header)
            if simplified and simplified not in header_variants:
                header_variants[simplified] = header
            no_parentheses = re.sub(r'\([^)]*\)', '', header)
            simplified_no_parentheses = self._simplify_header(no_parentheses)
            if (
                simplified_no_parentheses
                and simplified_no_parentheses not in header_variants
            ):
                header_variants[simplified_no_parentheses] = header

        for entry in configured:
            if not isinstance(entry, str):
                continue
            simplified_entry = self._simplify_header(entry)
            if simplified_entry and simplified_entry in header_variants:
                header = header_variants[simplified_entry]
                if header not in columns_to_remove:
                    columns_to_remove.append(header)
                continue

            normalized_keyword = self._normalize_text(entry)
            if normalized_keyword:
                keywords.append((entry, normalized_keyword))

        return columns_to_remove, keywords

    def _remove_rows_by_description_keywords(
            self,
            data_rows: List[Dict[str, Any]],
            keywords: List[Tuple[str, str]],
            logger,
    ) -> List[Dict[str, Any]]:
        if not data_rows or not keywords:
            return data_rows

        normalized_keywords: List[str] = []
        logged_keywords: List[str] = []
        seen = set()

        for original, normalized in keywords:
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            normalized_keywords.append(normalized)
            logged_keywords.append(original.strip() or normalized)

        if not normalized_keywords:
            return data_rows

        filtered_rows: List[Dict[str, Any]] = []
        removed_count = 0

        for row_data in data_rows:
            normalized_description = self._normalize_text(row_data.get("Descripción"))
            if (
                normalized_description
                and any(keyword in normalized_description for keyword in normalized_keywords)
            ):
                removed_count += 1
                continue
            filtered_rows.append(row_data)

        if removed_count:
            logger.log(
                (
                    "Se eliminaron "
                    f"{removed_count} fila(s) por coincidencias en la descripción con: "
                    + ", ".join(logged_keywords)
                ),
                level="INFO",
            )

        return filtered_rows

    def _highlight_rows_by_filters(
            self,
            worksheet,
            data_rows: List[Dict[str, Any]],
            headers: List[str],
            logger,
    ) -> None:
        """Resalta las filas cuya descripción coincida con los filtros configurados."""
        filters = self.config_manager.get_case5_filters()

        if not filters:
            return

        normalized_filters = [
            self._normalize_text(filter_text)
            for filter_text in filters
            if self._normalize_text(filter_text)
        ]

        if not normalized_filters:
            return

        try:
            description_idx = headers.index("Descripción") + 1
        except ValueError:
            logger.log(
                "No se encontró la columna Descripción para aplicar filtros del Caso 5.",
                level="WARNING",
            )
            return

        review_idx = None
        if "Revisar" in headers:
            review_idx = headers.index("Revisar") + 1

        from openpyxl.styles import Alignment, PatternFill

        highlight_fill = PatternFill(fill_type='solid', fgColor='FFF3B0')
        header_row = 13
        data_start = header_row + 1
        highlighted_rows = 0

        total_columns = len(headers)

        for row_offset, _ in enumerate(data_rows):
            current_row = data_start + row_offset
            cell_value = worksheet.cell(row=current_row, column=description_idx).value

            if cell_value in (None, ""):
                continue

            normalized_value = self._normalize_text(str(cell_value))

            if not normalized_value:
                continue

            if any(filter_text in normalized_value for filter_text in normalized_filters):
                for col_idx in range(1, total_columns + 1):
                    cell = worksheet.cell(row=current_row, column=col_idx)
                    cell.fill = highlight_fill

                if review_idx:
                    review_cell = worksheet.cell(row=current_row, column=review_idx)
                    review_cell.value = 'Revisar'
                    review_cell.alignment = Alignment(horizontal='center', vertical='center')

                highlighted_rows += 1

        if highlighted_rows:
            logger.log(
                (
                    "Se resaltaron "
                    f"{highlighted_rows} fila(s) que coinciden con los filtros configurados del Caso 5."
                ),
                level="INFO",
            )

    def _normalize_text(self, text: Any) -> str:
        """Normaliza texto eliminando acentos, espacios y caracteres especiales."""
        import unicodedata

        if not isinstance(text, str):
            return ''
        normalized = unicodedata.normalize('NFKD', text)
        normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r'[^\w\s]', '', normalized)
        return normalized.lower().strip()

    def _build_output_filename(self, original_name: str) -> str:
        """Construye el nombre del archivo de salida rediseñado."""
        base, _ = os.path.splitext(original_name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base}_caso5_{timestamp}.xlsx"
